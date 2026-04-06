#!/usr/bin/env python3
from __future__ import annotations

import csv
import difflib
import json
import re
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

import migrate_hodom_csv as base


BASE_DIR = Path(__file__).resolve().parent.parent
DOWNLOADS_DIR = Path.home() / "Downloads"
ENRICHED_DIR = BASE_DIR / "output" / "spreadsheet" / "enriched"
OUTPUT_DIR = BASE_DIR / "output" / "active_patient_packets_2026-04-01"
LEGACY_BASE_DIR = BASE_DIR / "documentacion-legacy" / "Hodom-hsc-dia-1"
LEGACY_TURNO_DIR = LEGACY_BASE_DIR / "ENTREGA TURNO "
LEGACY_FILE_RESPONSES_DIR = LEGACY_BASE_DIR / "Formulario sin título (File responses)" / "DAU -- EPICRISIS (ADJUNTAR) IDENTIFICAR ARCHIVO CON NOMBRE DE USUARIO (File responses)"
LEGACY_OLD_EPICRISIS_DIR = LEGACY_BASE_DIR / "EPICRISIS ANTIGUAS" / "DAU -- EPICRISIS (ADJUNTAR) identificar cada archivo con nombre de usuario (File responses)"
LEGACY_FORM_FILES = [
    LEGACY_BASE_DIR / "Copia de FORMULARIO HODOM HSC (respuestas).xlsx",
    LEGACY_BASE_DIR / "2025 FORMULARIO HODOM.xlsx",
    LEGACY_BASE_DIR / "FORMULARIO 2026 RESP (respuestas).xlsx",
]
DOWNLOAD_TURNO_DIR = DOWNLOADS_DIR / "drive-download-20260401T185340Z-1-001"
DOWNLOAD_KINE_TURNO_FILE = DOWNLOADS_DIR / "Ent. Turno Hodom KINE (1).xlsx"
DOWNLOAD_FORM_2026_FILE = DOWNLOADS_DIR / "FORMULARIO 2026 RESP (respuestas) (2).xlsx"

ROSTER = [
    {"cama": "01", "nombre": "LUIS ALBERTO MENDEZ MUÑOZ", "rut": "9441281-1", "fecha_ingreso_roster": "25-03-2026"},
    {"cama": "02", "nombre": "SONIA GERMANA CRISOSTOMO MAUREIRA", "rut": "5494336-9", "fecha_ingreso_roster": "SIN INGRESO"},
    {"cama": "03", "nombre": "SYLVIA DEL CARMEN PALACIOS MORALES", "rut": "5885688-6", "fecha_ingreso_roster": "01-04-2026"},
    {"cama": "04", "nombre": "MARIA LUZ GONZALEZ CANDIA", "rut": "5497062-5", "fecha_ingreso_roster": "23-03-2026"},
    {"cama": "05", "nombre": "VICTOR HUGO BELMAR BELMAR", "rut": "9888109-3", "fecha_ingreso_roster": "23-03-2026"},
    {"cama": "06", "nombre": "LUIS ROBERTO MALDONADO BURGOS", "rut": "10885193-7", "fecha_ingreso_roster": "12-02-2026"},
    {"cama": "07", "nombre": "NELSON AGUSTIN SEPULVEDA SEPULVEDA", "rut": "6335989-0", "fecha_ingreso_roster": "31-03-2026"},
    {"cama": "08", "nombre": "LUIS ANSELMO CORREA ROJAS", "rut": "5155151-6", "fecha_ingreso_roster": "01-04-2026"},
    {"cama": "09", "nombre": "LUCIA AURORA NORAMBUENA PINOCHET", "rut": "13615312-9", "fecha_ingreso_roster": "20-03-2026"},
    {"cama": "10", "nombre": "TERESA DEL ROSARIO VALENZUELA POBLETE", "rut": "4323133-2", "fecha_ingreso_roster": "31-03-2026"},
    {"cama": "11", "nombre": "ELISA ELIANA GARCIA MOLINA", "rut": "3362648-7", "fecha_ingreso_roster": "31-03-2026"},
    {"cama": "12", "nombre": "LUZMENIA ISABEL SOTO ROMAN", "rut": "4931889-8", "fecha_ingreso_roster": "25-03-2026"},
    {"cama": "13", "nombre": "DANIEL ANTONIO CRISOSTOMO MOLINA", "rut": "12205281-8", "fecha_ingreso_roster": "12-03-2026"},
    {"cama": "14", "nombre": "MARIA LUZ DEL ROSARIO ALARCON INZUNZA", "rut": "10001107-7", "fecha_ingreso_roster": "20-03-2026"},
    {"cama": "15", "nombre": "NORMA ROSA HERNANDEZ CASTILLO", "rut": "6458759-5", "fecha_ingreso_roster": "27-03-2026"},
    {"cama": "18", "nombre": "NESTOR IVAN RIQUELME BASCUR", "rut": "11444532-0", "fecha_ingreso_roster": "12-12-2025"},
    {"cama": "19", "nombre": "MARGARITA DEL CARMEN GOMEZ GUTIERREZ", "rut": "5465758-7", "fecha_ingreso_roster": "27-03-2026"},
    {"cama": "23", "nombre": "ZOILA DEL CARMEN QUEZADA JAQUE", "rut": "6174494-0", "fecha_ingreso_roster": "20-03-2026"},
    {"cama": "24", "nombre": "SERGIO IGNACIO SEPULVEDA RODRIGUEZ", "rut": "5016534-5", "fecha_ingreso_roster": "27-03-2026"},
    {"cama": "25", "nombre": "MARTA ELENA ROMERO GUZMAN", "rut": "9260553-1", "fecha_ingreso_roster": "27-03-2026"},
    {"cama": "26", "nombre": "CECILIA LEON GALLEGOS", "rut": "5268164-2", "fecha_ingreso_roster": "28-03-2026"},
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)


def slugify(text: str) -> str:
    return base.slug_text(text).replace("-", "_")


def parse_roster_date(value: str) -> str:
    return "" if value == "SIN INGRESO" else base.parse_date(value)


def best_name_match(rows: list[dict[str, str]], target_name: str, threshold: float = 0.86) -> dict[str, str] | None:
    best = None
    best_score = 0.0
    for row in rows:
        score = difflib.SequenceMatcher(
            None,
            base.canonical_text(target_name),
            base.canonical_text(row.get("nombre_completo", "")),
        ).ratio()
        if score > best_score:
            best = row
            best_score = score
    if best and best_score >= threshold:
        return best
    return None


def sort_episodes(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    def key(row: dict[str, str]) -> tuple[str, str, str]:
        return (
            row.get("fecha_ingreso", "") or row.get("requested_at", "") or row.get("fecha_egreso", ""),
            row.get("fecha_egreso", ""),
            row.get("episode_id", ""),
        )

    return sorted(rows, key=key, reverse=True)


def extract_docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml").decode("utf-8", errors="ignore")
    except Exception:
        return ""
    text = re.sub(r"<[^>]+>", " ", xml)
    return re.sub(r"\s+", " ", text).strip()


def text_contains_patient(text: str, roster_name: str, canonical_name: str) -> bool:
    normalized = base.canonical_text(text)
    roster_tokens = [token for token in base.canonical_text(roster_name).split() if len(token) >= 4]
    patient_tokens = [token for token in base.canonical_text(canonical_name).split() if len(token) >= 4]
    target_tokens = patient_tokens or roster_tokens
    if not target_tokens:
        return False
    return all(token in normalized for token in target_tokens[:2]) if len(target_tokens) >= 2 else target_tokens[0] in normalized


def text_snippet_for_patient(text: str, roster_name: str, canonical_name: str, max_len: int = 900) -> str:
    normalized = re.sub(r"\s+", " ", text).strip()
    for probe in [canonical_name, roster_name]:
        probe = probe.strip()
        if not probe:
            continue
        idx = normalized.upper().find(probe.upper().split()[0])
        if idx != -1:
            start = max(0, idx - 180)
            end = min(len(normalized), idx + max_len)
            return normalized[start:end]
    return normalized[:max_len]


def collect_legacy_form_rows(roster_name: str, roster_rut: str) -> list[dict[str, str]]:
    hits: list[dict[str, str]] = []
    canonical_roster = base.canonical_text(roster_name)
    for path in LEGACY_FORM_FILES:
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = [base.normalize_whitespace(str(cell or "")) for cell in next(ws.iter_rows(values_only=True))]
        canonical_header = {base.canonical_text(name): idx for idx, name in enumerate(header)}

        def get_value(row: tuple[object, ...], *keys: str) -> str:
            for key in keys:
                idx = canonical_header.get(base.canonical_text(key))
                if idx is not None and idx < len(row):
                    return base.normalize_whitespace(str(row[idx] or ""))
            return ""

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nombre = base.join_non_empty(
                [
                    get_value(row, "NOMBRES", "NOMBRES "),
                    get_value(row, "APELLIDO PATERNO"),
                    get_value(row, "APELLIDO MATERNO"),
                    get_value(row, "APELLIDOS "),
                ]
            )
            rut = get_value(row, "RUT (sin puntos con guión)", "RUT (9999999-9) sin puntos con guión ")
            rut_norm = base.normalize_rut(rut)
            if (rut_norm and rut_norm == roster_rut) or base.canonical_text(nombre) == canonical_roster:
                hits.append(
                    {
                        "source_file": path.name,
                        "source_row_number": str(row_number),
                        "nombre": nombre,
                        "rut_raw": rut,
                        "rut_norm": rut_norm,
                        "fecha_nacimiento": get_value(row, "FECHA DE NACIMIENTO (dd-mm-aaaa)"),
                        "edad": get_value(row, "EDAD"),
                        "servicio_origen_solicitud": get_value(row, "SERVICIO ORIGEN SOLICITUD", "SERVICIO DE ORIGEN SOLICITUD"),
                        "diagnostico": get_value(row, "DIAGNÓSTICO DE EGRESO DE HOSPITALIZACIÓN", "DIAGNÓSTICO DE EGRESO "),
                        "prestacion_solicitada": get_value(row, "PRESTACIÓN SOLICITADA"),
                        "gestora": get_value(row, "NOMBRE POSTULANTE (GESTOR(A))", "GESTORA ENCARGADA"),
                        "direccion": get_value(row, "DIRECCIÓN", "DIRECCIÓN "),
                        "cesfam": get_value(row, "CESFAM INSCRITO"),
                    }
                )
    return hits


def collect_download_form_rows(roster_name: str, roster_rut: str) -> list[dict[str, str]]:
    path = DOWNLOAD_FORM_2026_FILE
    if not path.exists():
        return []
    hits: list[dict[str, str]] = []
    canonical_roster = base.canonical_text(roster_name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [base.normalize_whitespace(str(cell or "")) for cell in next(ws.iter_rows(values_only=True))]
    canonical_header = {base.canonical_text(name): idx for idx, name in enumerate(header)}

    def get_value(row: tuple[object, ...], *keys: str) -> str:
        for key in keys:
            idx = canonical_header.get(base.canonical_text(key))
            if idx is not None and idx < len(row):
                return base.normalize_whitespace(str(row[idx] or ""))
        return ""

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nombres = get_value(row, "NOMBRES")
        apellidos = get_value(row, "APELLIDOS ")
        nombre = base.join_non_empty([nombres, apellidos])
        rut_raw = get_value(row, "RUT (9999999-9) sin puntos con guión ")
        rut_norm = base.normalize_rut(rut_raw)
        if (rut_norm and rut_norm == roster_rut) or base.canonical_text(nombre) == canonical_roster:
            hits.append(
                {
                    "source_file": path.name,
                    "source_row_number": str(row_number),
                    "timestamp": get_value(row, "Columna 1"),
                    "nombre": nombre,
                    "rut_raw": rut_raw,
                    "rut_norm": rut_norm,
                    "fecha_nacimiento": get_value(row, "FECHA DE NACIMIENTO (dd-mm-aaaa)"),
                    "edad": get_value(row, "EDAD"),
                    "servicio_origen_solicitud": get_value(row, "SERVICIO DE ORIGEN SOLICITUD"),
                    "diagnostico": get_value(row, "DIAGNÓSTICO DE EGRESO "),
                    "prestacion_solicitada": get_value(row, "PRESTACIÓN SOLICITADA"),
                    "gestora": get_value(row, "GESTORA ENCARGADA"),
                    "direccion": get_value(row, "DIRECCIÓN"),
                    "cesfam": get_value(row, "CESFAM INSCRITO"),
                    "celular": get_value(row, "CELULAR"),
                }
            )
    return hits


def collect_legacy_turno_hits(roster_name: str, canonical_name: str) -> list[dict[str, str]]:
    hits: list[dict[str, str]] = []
    if not LEGACY_TURNO_DIR.exists():
        return hits
    for path in sorted(LEGACY_TURNO_DIR.glob("*.docx")):
        text = extract_docx_text(path)
        if not text:
            continue
        if text_contains_patient(text, roster_name, canonical_name):
            hits.append(
                {
                    "source_file": path.name,
                    "source_path": str(path),
                    "snippet": text_snippet_for_patient(text, roster_name, canonical_name),
                }
            )
    return hits


def collect_download_turno_hits(roster_name: str, canonical_name: str) -> list[dict[str, str]]:
    hits: list[dict[str, str]] = []
    if not DOWNLOAD_TURNO_DIR.exists():
        return hits
    for path in sorted(DOWNLOAD_TURNO_DIR.glob("*.docx")):
        text = extract_docx_text(path)
        if not text:
            continue
        if text_contains_patient(text, roster_name, canonical_name):
            hits.append(
                {
                    "source_file": path.name,
                    "source_path": str(path),
                    "snippet": text_snippet_for_patient(text, roster_name, canonical_name),
                }
            )
    return hits


def collect_kine_turn_rows(roster_name: str, roster_rut: str, canonical_name: str) -> list[dict[str, str]]:
    path = DOWNLOAD_KINE_TURNO_FILE
    if not path.exists():
        return []
    hits: list[dict[str, str]] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            continue
        header = [base.normalize_whitespace(str(cell or "")) for cell in rows[4]]
        canonical_header = {base.canonical_text(name): idx for idx, name in enumerate(header)}

        def get_value(row: tuple[object, ...], *keys: str) -> str:
            for key in keys:
                idx = canonical_header.get(base.canonical_text(key))
                if idx is not None and idx < len(row):
                    return base.normalize_whitespace(str(row[idx] or ""))
            return ""

        for row in rows[5:]:
            nombre = get_value(row, "NOMBRE PCTE.")
            rut_raw = get_value(row, "RUT")
            rut_norm = base.normalize_rut(rut_raw)
            if (rut_norm and rut_norm == roster_rut) or base.canonical_text(nombre) == base.canonical_text(canonical_name or roster_name):
                hits.append(
                    {
                        "source_file": path.name,
                        "sheet": ws.title,
                        "nombre": nombre,
                        "rut_raw": rut_raw,
                        "rut_norm": rut_norm,
                        "cobertura": get_value(row, "COBERTURA"),
                        "diagnostico": get_value(row, "DIAGNOSTICO"),
                        "observaciones": get_value(row, "OBSERVACIONES"),
                        "hora": get_value(row, "HORA AT./AYUNO"),
                        "registro_fc": get_value(row, "REGISTRO FC"),
                    }
                )
    return hits


def collect_legacy_document_hits(roster_name: str, roster_rut: str, canonical_name: str) -> list[dict[str, str]]:
    hits: list[dict[str, str]] = []
    candidate_dirs = [LEGACY_FILE_RESPONSES_DIR, LEGACY_OLD_EPICRISIS_DIR]
    name_token = base.canonical_text(canonical_name or roster_name)
    for directory in candidate_dirs:
        if not directory.exists():
            continue
        for path in directory.glob("*"):
            if not path.is_file():
                continue
            haystack = base.canonical_text(path.name)
            score = difflib.SequenceMatcher(None, name_token, haystack).ratio()
            if roster_rut.replace("-", "") in haystack or score >= 0.55:
                hits.append(
                    {
                        "source_file": path.name,
                        "source_path": str(path),
                        "match_score": f"{score:.2f}",
                        "extension": path.suffix.lower(),
                    }
                )
    return sorted(hits, key=lambda row: float(row["match_score"]), reverse=True)


def build_findings(
    roster_row: dict[str, str],
    patient: dict[str, str] | None,
    active_episodes: list[dict[str, str]],
    all_episodes: list[dict[str, str]],
    requests: list[dict[str, str]],
    quality: list[dict[str, str]],
    review_queue: list[dict[str, str]],
    forms: list[dict[str, str]],
    address_rows: list[dict[str, str]],
    legacy_forms: list[dict[str, str]],
    legacy_turnos: list[dict[str, str]],
    legacy_docs: list[dict[str, str]],
    download_forms: list[dict[str, str]],
    download_turnos: list[dict[str, str]],
    kine_turn_rows: list[dict[str, str]],
) -> list[str]:
    findings: list[str] = []
    if patient is None:
        findings.append("CRITICAL: No existe match exacto en `patient_master` para este paciente del roster.")
    if patient and not active_episodes:
        findings.append("CRITICAL: El roster lo muestra hospitalizado al 01-04-2026, pero no hay episodio `ACTIVO` en la capa enriquecida.")
    if len(active_episodes) > 1:
        findings.append(f"WARNING: Hay {len(active_episodes)} episodios activos simultáneos para el mismo paciente.")
    if roster_row["fecha_ingreso_roster"] != "SIN INGRESO" and active_episodes:
        roster_date = parse_roster_date(roster_row["fecha_ingreso_roster"])
        active_date = active_episodes[0].get("fecha_ingreso", "")
        if roster_date and active_date and roster_date != active_date:
            findings.append(f"WARNING: La fecha de ingreso del roster ({roster_date}) no coincide con el episodio activo ({active_date}).")
    if patient and patient.get("canonical_patient_key", "").startswith("rut:") and patient.get("rut") and patient["canonical_patient_key"] != f"rut:{patient['rut']}":
        findings.append("WARNING: `canonical_patient_key` no coincide con el RUT actual del paciente.")
    if quality:
        findings.append(f"WARNING: Hay {len(quality)} issues de calidad asociados a este paciente/episodio.")
    if review_queue:
        findings.append(f"WARNING: Hay {len(review_queue)} candidatos en `match_review_queue` para este paciente.")
    if not requests and forms:
        findings.append("WARNING: Hay formularios redundantes, pero no quedaron requests consolidados en `episode_request`.")
    if not forms:
        findings.append("WARNING: No se encontraron formularios redundantes asociados.")
    if legacy_forms:
        findings.append(f"INFO: Se encontraron {len(legacy_forms)} filas en formularios históricos locales.")
    if legacy_turnos:
        findings.append(f"INFO: Se encontraron {len(legacy_turnos)} entregas de turno locales con mención al paciente.")
    if legacy_docs:
        findings.append(f"INFO: Se encontraron {len(legacy_docs)} archivos DAU/Epicrisis locales potencialmente asociados.")
    if download_forms:
        findings.append(f"INFO: Se encontraron {len(download_forms)} filas en FORMULARIO 2026 RESP (respuestas) (2).xlsx.")
    if download_turnos:
        findings.append(f"INFO: Se encontraron {len(download_turnos)} entregas de turno descargadas con mención al paciente.")
    if kine_turn_rows:
        findings.append(f"INFO: Se encontraron {len(kine_turn_rows)} registros en Ent. Turno Hodom KINE.")
    if address_rows:
        unresolved = [row for row in address_rows if row.get("resolution_status") != "RESOLVED"]
        if unresolved:
            findings.append("WARNING: La resolución territorial del paciente no está completamente resuelta.")
    if roster_row["fecha_ingreso_roster"] == "SIN INGRESO":
        findings.append("INFO: El roster indica `SIN INGRESO`; revisar coherencia entre hospitalización activa y registro clínico-administrativo.")
    return findings


def write_summary(
    path: Path,
    roster_row: dict[str, str],
    patient: dict[str, str] | None,
    active_episodes: list[dict[str, str]],
    all_episodes: list[dict[str, str]],
    requests: list[dict[str, str]],
    discharges: list[dict[str, str]],
    quality: list[dict[str, str]],
    review_queue: list[dict[str, str]],
    forms: list[dict[str, str]],
    provenance: list[dict[str, str]],
    address_rows: list[dict[str, str]],
    legacy_forms: list[dict[str, str]],
    legacy_turnos: list[dict[str, str]],
    legacy_docs: list[dict[str, str]],
    download_forms: list[dict[str, str]],
    download_turnos: list[dict[str, str]],
    kine_turn_rows: list[dict[str, str]],
    findings: list[str],
) -> None:
    lines = [
        f"# Paciente {roster_row['cama']} - {roster_row['nombre']}",
        "",
        "## Roster 01-04-2026",
        "",
        f"- Cama: `{roster_row['cama']}`",
        f"- Nombre: `{roster_row['nombre']}`",
        f"- RUT roster: `{roster_row['rut']}`",
        f"- Fecha ingreso roster: `{roster_row['fecha_ingreso_roster']}`",
        "",
        "## Match Enriquecido",
        "",
    ]

    if patient:
        lines.extend(
            [
                f"- patient_id: `{patient.get('patient_id', '')}`",
                f"- nombre canónico: `{patient.get('nombre_completo', '')}`",
                f"- rut canónico: `{patient.get('rut', '')}`",
                f"- resolución identidad: `{patient.get('identity_resolution_status', '')}`",
                f"- strategy: `{patient.get('patient_key_strategy', '')}`",
                f"- fecha nacimiento: `{patient.get('fecha_nacimiento_date', '')}`",
                f"- edad reportada: `{patient.get('edad_reportada', '')}`",
                f"- comuna: `{patient.get('comuna', '')}`",
                f"- cesfam: `{patient.get('cesfam', '')}`",
                f"- episodes históricos: `{patient.get('episode_count', '')}`",
            ]
        )
    else:
        lines.append("- Sin match exacto en `patient_master`.")

    lines.extend(
        [
            "",
            "## Episodios",
            "",
            f"- Episodios activos encontrados: `{len(active_episodes)}`",
            f"- Episodios históricos encontrados: `{len(all_episodes)}`",
            f"- Requests asociados: `{len(requests)}`",
            f"- Discharges asociados: `{len(discharges)}`",
            f"- Formularios redundantes: `{len(forms)}`",
            f"- Formularios históricos locales: `{len(legacy_forms)}`",
            f"- Entregas de turno locales: `{len(legacy_turnos)}`",
            f"- Archivos DAU/Epicrisis locales: `{len(legacy_docs)}`",
            f"- Formularios descargados 2026: `{len(download_forms)}`",
            f"- Entregas de turno descargadas: `{len(download_turnos)}`",
            f"- Registros turno KINE: `{len(kine_turn_rows)}`",
            f"- Quality issues: `{len(quality)}`",
            f"- Review queue rows: `{len(review_queue)}`",
            f"- Field provenance rows: `{len(provenance)}`",
            "",
            "## Auditoría",
            "",
        ]
    )
    if findings:
        for item in findings:
            lines.append(f"- {item}")
    else:
        lines.append("- Sin hallazgos críticos o advertencias relevantes.")

    lines.extend(
        [
            "",
            "## Archivos En La Carpeta",
            "",
            "- `roster_row.json`",
            "- `patient_master.csv`",
            "- `active_episodes.csv`",
            "- `episode_history.csv`",
            "- `episode_requests.csv`",
            "- `episode_discharges.csv`",
            "- `normalized_forms.csv`",
            "- `quality_issues.csv`",
            "- `match_review_queue.csv`",
            "- `legacy_form_rows.csv`",
            "- `legacy_turn_handoffs.csv`",
            "- `legacy_document_hits.csv`",
            "- `download_form_2026_rows.csv`",
            "- `download_turn_handoffs.csv`",
            "- `kine_turn_rows.csv`",
            "- `field_provenance.csv`",
            "- `address_resolution.csv`",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    patients = read_csv(ENRICHED_DIR / "patient_master.csv")
    episodes = read_csv(ENRICHED_DIR / "episode_master.csv")
    requests = read_csv(ENRICHED_DIR / "episode_request.csv")
    discharges = read_csv(ENRICHED_DIR / "episode_discharge.csv")
    quality_issues = read_csv(ENRICHED_DIR / "data_quality_issue.csv")
    review_queue = read_csv(ENRICHED_DIR / "match_review_queue.csv")
    forms = read_csv(ENRICHED_DIR / "normalized_form_submission.csv")
    field_provenance = read_csv(ENRICHED_DIR / "field_provenance.csv")
    address_resolution = read_csv(ENRICHED_DIR / "address_resolution.csv")

    patient_by_rut = {row["rut"]: row for row in patients if row.get("rut")}
    index_rows: list[dict[str, str]] = []

    for roster_row in ROSTER:
        patient = patient_by_rut.get(roster_row["rut"])
        match_method = "rut_exact" if patient else "none"

        if patient is None:
            patient = best_name_match(patients, roster_row["nombre"])
            if patient:
                match_method = "name_similarity"

        patient_id = patient.get("patient_id", "") if patient else ""
        patient_name = patient.get("nombre_completo", "") if patient else ""
        patient_rut = patient.get("rut", "") if patient else roster_row["rut"]

        patient_episodes = sort_episodes([row for row in episodes if row.get("patient_id") == patient_id]) if patient_id else []
        active_episodes = [row for row in patient_episodes if row.get("estado") == "ACTIVO"]
        patient_requests = sort_episodes([row for row in requests if row.get("patient_id") == patient_id]) if patient_id else []
        patient_discharges = sort_episodes([row for row in discharges if row.get("patient_id") == patient_id]) if patient_id else []
        patient_quality = [row for row in quality_issues if row.get("episode_id") in {ep["episode_id"] for ep in patient_episodes}]
        patient_provenance = [
            row
            for row in field_provenance
            if (row.get("entity_type") == "patient" and row.get("entity_id") == patient_id)
            or (row.get("entity_type") == "episode" and row.get("entity_id") in {ep["episode_id"] for ep in patient_episodes})
        ]
        patient_review_queue = [
            row
            for row in review_queue
            if row.get("patient_id") == patient_id
            or (patient_rut and row.get("patient_rut") == patient_rut)
            or (patient_name and base.canonical_text(row.get("patient_name", "")) == base.canonical_text(patient_name))
        ]
        patient_forms = [
            row
            for row in forms
            if row.get("rut_norm") == roster_row["rut"]
            or base.canonical_text(row.get("nombre_completo", "")) == base.canonical_text(roster_row["nombre"])
        ]
        legacy_form_rows = collect_legacy_form_rows(roster_row["nombre"], roster_row["rut"])
        legacy_turnos = collect_legacy_turno_hits(roster_row["nombre"], patient_name or roster_row["nombre"])
        legacy_docs = collect_legacy_document_hits(roster_row["nombre"], roster_row["rut"], patient_name or roster_row["nombre"])
        download_forms = collect_download_form_rows(roster_row["nombre"], roster_row["rut"])
        download_turnos = collect_download_turno_hits(roster_row["nombre"], patient_name or roster_row["nombre"])
        kine_turn_rows = collect_kine_turn_rows(roster_row["nombre"], roster_row["rut"], patient_name or roster_row["nombre"])
        patient_address_rows = [row for row in address_resolution if row.get("patient_id") == patient_id] if patient_id else []

        findings = build_findings(
            roster_row,
            patient,
            active_episodes,
            patient_episodes,
            patient_requests,
            patient_quality,
            patient_review_queue,
            patient_forms,
            patient_address_rows,
            legacy_form_rows,
            legacy_turnos,
            legacy_docs,
            download_forms,
            download_turnos,
            kine_turn_rows,
        )

        folder_name = f"{roster_row['cama']}_{slugify(roster_row['nombre'])}_{roster_row['rut'].replace('-', '_')}"
        packet_dir = OUTPUT_DIR / folder_name
        packet_dir.mkdir(parents=True, exist_ok=True)

        (packet_dir / "roster_row.json").write_text(json.dumps(roster_row, indent=2, ensure_ascii=False), encoding="utf-8")
        write_summary(
            packet_dir / "patient_summary.md",
            roster_row,
            patient,
            active_episodes,
            patient_episodes,
            patient_requests,
            patient_discharges,
            patient_quality,
            patient_review_queue,
            patient_forms,
            patient_provenance,
            patient_address_rows,
            legacy_form_rows,
            legacy_turnos,
            legacy_docs,
            download_forms,
            download_turnos,
            kine_turn_rows,
            findings,
        )
        write_csv(packet_dir / "patient_master.csv", [patient] if patient else [])
        write_csv(packet_dir / "active_episodes.csv", active_episodes)
        write_csv(packet_dir / "episode_history.csv", patient_episodes)
        write_csv(packet_dir / "episode_requests.csv", patient_requests)
        write_csv(packet_dir / "episode_discharges.csv", patient_discharges)
        write_csv(packet_dir / "normalized_forms.csv", patient_forms)
        write_csv(packet_dir / "quality_issues.csv", patient_quality)
        write_csv(packet_dir / "match_review_queue.csv", patient_review_queue)
        write_csv(packet_dir / "legacy_form_rows.csv", legacy_form_rows)
        write_csv(packet_dir / "legacy_turn_handoffs.csv", legacy_turnos)
        write_csv(packet_dir / "legacy_document_hits.csv", legacy_docs)
        write_csv(packet_dir / "download_form_2026_rows.csv", download_forms)
        write_csv(packet_dir / "download_turn_handoffs.csv", download_turnos)
        write_csv(packet_dir / "kine_turn_rows.csv", kine_turn_rows)
        write_csv(packet_dir / "field_provenance.csv", patient_provenance)
        write_csv(packet_dir / "address_resolution.csv", patient_address_rows)

        index_rows.append(
            {
                "cama": roster_row["cama"],
                "roster_nombre": roster_row["nombre"],
                "roster_rut": roster_row["rut"],
                "roster_fecha_ingreso": roster_row["fecha_ingreso_roster"],
                "match_method": match_method,
                "patient_id": patient_id,
                "patient_nombre": patient_name,
                "patient_rut": patient_rut,
                "active_episode_count": str(len(active_episodes)),
                "episode_history_count": str(len(patient_episodes)),
                "requests_count": str(len(patient_requests)),
                "discharges_count": str(len(patient_discharges)),
                "forms_count": str(len(patient_forms)),
                "legacy_forms_count": str(len(legacy_form_rows)),
                "legacy_turnos_count": str(len(legacy_turnos)),
                "legacy_docs_count": str(len(legacy_docs)),
                "download_forms_count": str(len(download_forms)),
                "download_turnos_count": str(len(download_turnos)),
                "kine_turn_rows_count": str(len(kine_turn_rows)),
                "quality_issues_count": str(len(patient_quality)),
                "review_queue_count": str(len(patient_review_queue)),
                "folder": str(packet_dir.relative_to(BASE_DIR)),
            }
        )

    write_csv(OUTPUT_DIR / "index.csv", index_rows)
    (OUTPUT_DIR / "README.md").write_text(
        "# Paquetes de pacientes activos al 2026-04-01\n\nSe generó una carpeta por paciente del roster, con resumen narrativo y extractos CSV de la capa enriquecida.\n",
        encoding="utf-8",
    )
    print(f"Output dir: {OUTPUT_DIR}")
    print(f"Patients processed: {len(index_rows)}")


if __name__ == "__main__":
    main()
