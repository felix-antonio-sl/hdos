#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import difflib
import json
import math
import subprocess
import tempfile
import warnings
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import requests
from requests.exceptions import RequestException, SSLError

import build_hodom_intermediate as core
import migrate_hodom_csv as base

try:
    import pyogrio
except ImportError:  # pragma: no cover - optional runtime dependency
    pyogrio = None


REPO_DIR = Path(__file__).resolve().parents[1]
DOWNLOADS_DIR = Path.home() / "Downloads"

FORM_SOURCE_PATHS = [
    DOWNLOADS_DIR / "2025 FORMULARIO HODOM (1).xlsx",
    DOWNLOADS_DIR / "2025 FORMULARIO HODOM (2).xlsx",
    DOWNLOADS_DIR / "FORMULARIO 2026 RESP (respuestas) (1).xlsx",
]

DISCHARGE_SOURCE_PATH = REPO_DIR / "input" / "reference" / "legacy_imports" / "planilla-altas-2024-2026.xlsx"
GEODB_REFERENCE_PATH = DOWNLOADS_DIR / "Cartografía_censo2024_R16.gdb"
GEODB_DICTIONARY_PATH = DOWNLOADS_DIR / "Diccionario_variables_geograficas_CPV24.xlsx"
DEIS_REFERENCE_URL = "https://www.minsal.cl/wp-content/uploads/2018/12/Listado-Establecimientos-DEIS.pdf"
INE_GEODATA_URL = "https://www.ine.gob.cl/herramientas/portal-de-mapas/geodatos-abiertos/"
INE_CENSO_NUBLE_URL = "https://regiones.ine.gob.cl/nuble/prensa/ine-publica-bases-de-datos-a-nivel-de-manzana-y-cartografia-del-censo-de-poblacion-y-vivienda-2024"
INE_LOCALIDAD_RURAL_NUBLE_URL = "https://geoine-ine-chile.opendata.arcgis.com/datasets/a38eda47692e43999559aea6e0d3f7cc_138"
INE_ENTIDAD_RURAL_NUBLE_URL = "https://geoine-ine-chile.opendata.arcgis.com/datasets/a38eda47692e43999559aea6e0d3f7cc_154"
ENRICHED_CACHE_DIR = Path("output/spreadsheet/enriched")
REFERENCE_SNAPSHOT_DIR = Path("input/reference")
MANUAL_DIR = Path("input/manual")
PUNILLA_COMMUNES = {"SAN CARLOS", "COIHUECO", "NINHUE", "PINTO", "SAN FABIAN", "SAN NICOLAS", "NIQUEN"}
COMMUNE_ALIAS_MAP = {
    "SAN CARLOS": "SAN CARLOS",
    "SAN CAROS": "SAN CARLOS",
    "SAN CARTLOS": "SAN CARLOS",
    "SAN CARLLOS": "SAN CARLOS",
    "SAN CARLOSS": "SAN CARLOS",
    "SSAN CARLOS": "SAN CARLOS",
    "SN CARLOS": "SAN CARLOS",
    "SAN CARCALOS": "SAN CARLOS",
    "SAN NICOLAS": "SAN NICOLAS",
    "SAN NICOLÁS": "SAN NICOLAS",
    "SANNICOLAS": "SAN NICOLAS",
    "NIQUEN": "NIQUEN",
    "ÑIQUEN": "NIQUEN",
    "ÑIQUÉN": "NIQUEN",
    "SAN FABIÁN": "SAN FABIAN",
    "SAN FABIAN": "SAN FABIAN",
    "SAN GREGORIO": "SAN NICOLAS",
    "CACHAPOAL": "SAN CARLOS",
}
ESTABLISHMENT_ALIAS_MAP = {
    "C T BALDECCHI": "117329",
    "C T BALDECHI": "117329",
    "C TERESA BALDECCHI": "117329",
    "C TERESA BALDECHI": "117329",
    "TERESA BALDECHI": "117329",
    "BALDECHI": "117329",
    "C BALDECHI": "117329",
    "C DURAN TRUJILLO": "117311",
    "C D TRUJILLO": "117311",
    "DURAN TRUJILLO": "117311",
    "JOSE DURAN TRUJILLO": "117311",
    "C SAN NICOLAS": "117312",
    "SAN NICOLAS": "117312",
    "PUENTE NUBLE": "117312",
    "PTE NUBLE": "117312",
    "C NIQUEN": "117313",
    "C N IQUEN": "117313",
    "C ÑIQUEN": "117313",
    "NIQUEN": "117313",
    "C SAN GREGORIO C NIQUEN": "117313",
}
LOCALITY_ALIAS_HINTS = {
    "CHACAY": ("CHACAY", "NIQUEN"),
    "RIVERA NUBLE": ("RIVERA DE ÑUBLE", "SAN CARLOS"),
    "RIVERA DE NUBLE": ("RIVERA DE ÑUBLE", "SAN CARLOS"),
    "TORRECILLA": ("TORRECILLAS", "SAN CARLOS"),
    "TORRECILLAS": ("TORRECILLAS", "SAN CARLOS"),
    "MONTE BLANCO": ("MONTE BLANCO", "SAN CARLOS"),
    "LA GLORIA": ("LA GLORIA", "NIQUEN"),
    "ARIZONA": ("ARIZONA", "SAN CARLOS"),
    "EL RINCON": ("EL RINCON", "NINHUE"),
    "BELEN": ("BELEN", "NIQUEN"),
    "ZEMITA": ("ZEMITA", "NIQUEN"),
    "TRABUNCURA": ("TRABUNCURA", "SAN FABIAN"),
    "CARACOL": ("EL CARACOL", "SAN FABIAN"),
    "RECINTO": ("RECINTO", "PINTO"),
    "EL SAUCE": ("EL SAUCE", "NINHUE"),
    "HUENUTIL": ("HUENUTIL", "NIQUEN"),
    "VIRGUIN": ("VIRGUIN", "NIQUEN"),
    "TIUQUILEMU": ("TIUQUILEMU", "NIQUEN"),
    "BUCALEMU": ("BUCALEMU", "NIQUEN"),
    "PAQUE": ("PAQUE NORTE", "NIQUEN"),
    "SAN FERNANDO": ("SAN FERNANDO", "NIQUEN"),
    "EL ESPINAL": ("EL ESPINAL", "NIQUEN"),
    "NINQUIHUE": ("NINQUIHUE", "SAN NICOLAS"),
    "TRAPICHE": ("TRAPICHE", "SAN CARLOS"),
    "GAONA": ("GAONA", "SAN CARLOS"),
    "CACHAPOAL": ("CACHAPOAL", "SAN CARLOS"),
    "DADINCO": ("DADINCO", "SAN NICOLAS"),
    "MONTE LEON": ("MONTE LEON", "SAN NICOLAS"),
    "MEMBRILLO": ("BAJO EL MEMBRILLO", "SAN NICOLAS"),
    "QUINTRALA": ("LA QUINTRALA", "SAN NICOLAS"),
}

RAW_FORM_SUBMISSION_COLUMNS = [
    "form_submission_id",
    "source_file",
    "source_sheet",
    "source_row_number",
    "raw_json",
]

RAW_DISCHARGE_SHEET_COLUMNS = [
    "discharge_row_id",
    "source_file",
    "source_sheet",
    "source_row_number",
    "raw_json",
]

RAW_REFERENCE_SNAPSHOT_COLUMNS = [
    "reference_snapshot_id",
    "reference_type",
    "source_url",
    "fetched_at",
    "status",
    "notes",
]

NORMALIZED_FORM_SUBMISSION_COLUMNS = [
    "form_submission_id",
    "dedupe_key",
    "form_source_count",
    "source_files",
    "source_rows",
    "submission_timestamp",
    "rut_raw",
    "rut_norm",
    "rut_valido",
    "nombres",
    "apellido_paterno",
    "apellido_materno",
    "apellidos",
    "nombre_completo",
    "fecha_nacimiento",
    "edad_reportada",
    "sexo",
    "servicio_origen_solicitud",
    "diagnostico",
    "direccion",
    "nro_casa",
    "cesfam",
    "celular_1",
    "celular_2",
    "prevision",
    "request_prestacion",
    "antecedentes",
    "gestora",
    "attachment_url",
    "usuario_o2",
    "source_authority",
]

NORMALIZED_DISCHARGE_EVENT_COLUMNS = [
    "discharge_event_id",
    "dedupe_key",
    "source_file",
    "source_sheet",
    "source_row_number",
    "fecha_ingreso",
    "fecha_egreso",
    "motivo_egreso",
    "diagnostico",
    "nombre_completo",
    "rut_raw",
    "rut_norm",
    "rut_valido",
    "comuna",
    "direccion_o_comuna",
    "source_authority",
]

NORMALIZED_FORM_EPISODE_CANDIDATE_COLUMNS = [
    "candidate_id",
    "candidate_type",
    "source_id",
    "patient_key",
    "episode_id",
    "match_score",
    "match_status",
    "resolution_status",
    "episode_origin",
    "rescue_priority",
]

NORMALIZED_DISCHARGE_EPISODE_CANDIDATE_COLUMNS = NORMALIZED_FORM_EPISODE_CANDIDATE_COLUMNS

PATIENT_MASTER_COLUMNS = [
    "patient_id",
    "canonical_patient_key",
    "identity_resolution_status",
    "patient_key",
    "patient_key_strategy",
    "rut",
    "rut_valido",
    "rut_raw",
    "nombres",
    "apellido_paterno",
    "apellido_materno",
    "apellidos",
    "nombre_completo",
    "sexo",
    "fecha_nacimiento_date",
    "fecha_nacimiento_raw",
    "edad_reportada",
    "nacionalidad",
    "nro_contacto",
    "domicilio",
    "comuna",
    "cesfam",
    "episode_count",
    "source_files",
    "patient_resolution_status",
    "canonical_address_text",
    "canonical_locality_id",
    "canonical_establishment_id",
]

EPISODE_MASTER_COLUMNS = [
    "episode_id",
    "patient_id",
    "source_episode_key",
    "record_uid",
    "estado",
    "tipo_flujo",
    "fecha_ingreso",
    "fecha_egreso",
    "dias_estadia_reportados",
    "dias_estadia_calculados",
    "motivo_egreso",
    "motivo_derivacion",
    "servicio_origen",
    "prevision",
    "barthel",
    "categorizacion",
    "usuario_o2",
    "requerimiento_o2",
    "diagnostico_principal_texto",
    "episode_status_quality",
    "duplicate_count",
    "episode_origin",
    "resolution_status",
    "match_status",
    "match_score",
    "requested_at",
    "request_prestacion",
    "gestora",
    "form_source_count",
    "codigo_deis",
    "establishment_id",
    "locality_id",
    "rescue_priority",
]

EPISODE_REQUEST_COLUMNS = [
    "episode_request_id",
    "form_submission_id",
    "episode_id",
    "patient_id",
    "submission_timestamp",
    "request_prestacion",
    "gestora",
    "servicio_origen_solicitud",
    "diagnostico",
    "match_score",
    "match_status",
    "resolution_status",
    "episode_origin",
    "source_file",
    "source_rows",
]

EPISODE_DISCHARGE_COLUMNS = [
    "episode_discharge_id",
    "discharge_event_id",
    "episode_id",
    "patient_id",
    "fecha_ingreso",
    "fecha_egreso",
    "motivo_egreso",
    "diagnostico",
    "match_score",
    "match_status",
    "resolution_status",
    "episode_origin",
    "source_file",
    "source_sheet",
    "source_row_number",
]

EPISODE_RESCUE_CANDIDATE_COLUMNS = [
    "rescue_candidate_id",
    "candidate_type",
    "source_id",
    "episode_id",
    "patient_id",
    "requested_at",
    "fecha_ingreso",
    "fecha_egreso",
    "nombre_completo",
    "rut_norm",
    "diagnostico",
    "servicio_origen",
    "motivo_egreso",
    "resolution_status",
    "episode_origin",
    "rescue_priority",
]

FIELD_PROVENANCE_COLUMNS = [
    "field_provenance_id",
    "entity_type",
    "entity_id",
    "field_name",
    "selected_value",
    "selected_source",
    "selected_confidence",
    "competing_values_json",
]

ESTABLISHMENT_REFERENCE_COLUMNS = [
    "establishment_id",
    "codigo_deis",
    "servicio_salud",
    "tipo_establecimiento",
    "dependencia",
    "nombre_oficial",
    "comuna",
    "via",
    "numero",
    "direccion",
    "region",
    "source_url",
    "fetched_at",
]

LOCALITY_REFERENCE_COLUMNS = [
    "locality_id",
    "nombre_oficial",
    "comuna",
    "provincia",
    "region",
    "territory_type",
    "latitud",
    "longitud",
    "crs",
    "source_url",
    "source_priority",
    "fetched_at",
]

LOCALITY_ALIAS_COLUMNS = [
    "locality_alias_id",
    "locality_id",
    "alias",
    "source_url",
]

ADDRESS_RESOLUTION_COLUMNS = [
    "address_resolution_id",
    "patient_id",
    "episode_id",
    "raw_address",
    "raw_comuna",
    "raw_cesfam",
    "resolved_establishment_id",
    "resolved_locality_id",
    "establishment_match_confidence",
    "locality_match_confidence",
    "resolution_status",
]

DATA_QUALITY_ISSUE_COLUMNS = [
    "quality_issue_id",
    "normalized_row_id",
    "episode_id",
    "issue_type",
    "severity",
    "raw_value",
    "suggested_value",
    "status",
]

RECONCILIATION_REPORT_COLUMNS = [
    "month",
    "baseline_episode_count",
    "matched_form_count",
    "matched_discharge_count",
    "rescued_form_count",
    "rescued_discharge_count",
    "enriched_episode_count",
    "delta_vs_baseline",
]

MATCH_REVIEW_QUEUE_COLUMNS = [
    "review_queue_id",
    "form_submission_id",
    "patient_id",
    "patient_name",
    "patient_rut",
    "submission_timestamp",
    "servicio_origen_solicitud",
    "diagnostico_form",
    "request_prestacion",
    "gestora",
    "candidate_rank",
    "candidate_episode_id",
    "candidate_episode_origin",
    "candidate_fecha_ingreso",
    "candidate_fecha_egreso",
    "candidate_servicio_origen",
    "candidate_diagnostico",
    "candidate_score",
    "auto_close_recommended",
]

IDENTITY_REVIEW_QUEUE_COLUMNS = [
    "identity_review_id",
    "issue_type",
    "patient_id",
    "episode_id",
    "patient_name",
    "patient_rut",
    "edad_reportada",
    "fecha_nacimiento_actual",
    "issue_raw_value",
    "suggested_value",
    "fecha_ingreso",
    "fecha_egreso",
    "episode_origin",
    "severity",
    "status",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def read_manual_csv(name: str) -> list[dict[str, str]]:
    path = MANUAL_DIR / name
    if not path.exists():
        return []
    return read_csv(path)


def ensure_intermediate(script_dir: Path) -> None:
    subprocess.run(
        ["python3", str(script_dir / "build_hodom_intermediate.py")],
        check=True,
    )


def parse_datetime_or_empty(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    text = base.normalize_whitespace(str(value))
    if text in {"", ",", "-", "None"}:
        return ""
    parsed = base.parse_date(str(value))
    if parsed:
        return f"{parsed}T00:00:00"
    return text


def parse_date_or_empty(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = base.normalize_whitespace(str(value))
    if text in {"", ",", "-", "None"}:
        return ""
    return base.parse_date(text)


def canonical_name(*parts: str) -> str:
    return base.join_non_empty([part for part in parts if base.normalize_whitespace(part)])


def read_existing_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def fetch_bytes_with_ssl_fallback(url: str, timeout: int = 60) -> tuple[bytes, str]:
    try:
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
        return response.content, "ssl_verified"
    except SSLError:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            response = requests.get(url, timeout=timeout, verify=False)
        response.raise_for_status()
        return response.content, "ssl_unverified"
    except RequestException:
        raise


def normalize_commune(value: str) -> str:
    canonical = base.canonical_text(value)
    return COMMUNE_ALIAS_MAP.get(canonical, base.normalize_whitespace(value).upper())


def normalize_establishment_alias(value: str) -> str:
    return base.canonical_text(value)


def normalize_locality_name(value: str) -> str:
    text = base.canonical_text(value)
    replacements = {
        "ALDEA DE ": "",
        "ALDEA ": "",
        "CASERIO DE ": "",
        "CASERIO ": "",
        "SECTOR ": "",
        "PUEBLO DE ": "",
        "PUEBLO ": "",
        "LOCALIDAD RURAL INDETERMINADA ": "",
        "ENTIDAD RURAL INDETERMINADA ": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text.strip()


def split_phone_candidates(*values: str) -> str:
    phones: list[str] = []
    for raw in values:
        for token in core.split_contact_values(base.normalize_whitespace(raw)):
            if token not in phones:
                phones.append(token)
    return "; ".join(phones)


def flatten_coords(coords: Any) -> list[tuple[float, float]]:
    if not isinstance(coords, list):
        return []
    if len(coords) >= 2 and all(isinstance(v, (int, float)) for v in coords[:2]):
        return [(float(coords[0]), float(coords[1]))]
    points: list[tuple[float, float]] = []
    for item in coords:
        points.extend(flatten_coords(item))
    return points


def centroid_from_geometry(geometry: dict[str, Any]) -> tuple[str, str]:
    if not geometry:
        return "", ""
    geom_type = geometry.get("type")
    coords = geometry.get("coordinates")
    points = flatten_coords(coords)
    if not points:
        return "", ""
    if geom_type == "Point":
        lon, lat = points[0]
        return f"{lat:.6f}", f"{lon:.6f}"
    lon = sum(point[0] for point in points) / len(points)
    lat = sum(point[1] for point in points) / len(points)
    return f"{lat:.6f}", f"{lon:.6f}"


def parse_form_workbook(path: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    header = [base.normalize_whitespace(str(cell or "")) for cell in next(iterator)]
    canonical_header = {base.canonical_text(name): idx for idx, name in enumerate(header)}

    raw_rows: list[dict[str, str]] = []
    normalized_rows: list[dict[str, str]] = []

    def cell(row: tuple[Any, ...], *keys: str) -> str:
        for key in keys:
            idx = canonical_header.get(base.canonical_text(key))
            if idx is not None and idx < len(row):
                return base.normalize_whitespace(str(row[idx] or ""))
        return ""

    for row_number, row in enumerate(iterator, start=2):
        if not any(cell_value not in (None, "") for cell_value in row):
            continue
        raw_json = json.dumps([None if value is None else str(value) for value in row], ensure_ascii=False)
        form_submission_id = core.stable_id("frm", path.name, row_number)
        raw_rows.append(
            {
                "form_submission_id": form_submission_id,
                "source_file": path.name,
                "source_sheet": ws.title,
                "source_row_number": str(row_number),
                "raw_json": raw_json,
            }
        )

        nombres = cell(row, "NOMBRES", "NOMBRES ")
        apellido_paterno = cell(row, "APELLIDO PATERNO")
        apellido_materno = cell(row, "APELLIDO MATERNO")
        apellidos = cell(row, "APELLIDOS ")
        nombre_completo = canonical_name(nombres, apellido_paterno, apellido_materno, apellidos)
        rut_raw = cell(row, "RUT (sin puntos con guión)", "RUT (9999999-9) sin puntos con guión ")
        rut_norm = base.normalize_rut(rut_raw)
        fecha_nacimiento = parse_date_or_empty(
            cell(row, "FECHA DE NACIMIENTO (dd-mm-aaaa)")
        )
        edad = base.parse_int(cell(row, "EDAD"))
        sexo = base.normalize_sexo(cell(row, "SEXO "))
        servicio = base.normalize_servicio_origen(cell(row, "SERVICIO ORIGEN SOLICITUD", "SERVICIO DE ORIGEN SOLICITUD"))
        diagnostico = cell(row, "DIAGNÓSTICO DE EGRESO DE HOSPITALIZACIÓN", "DIAGNÓSTICO DE EGRESO ")
        direccion = cell(row, "DIRECCIÓN", "DIRECCIÓN ")
        nro_casa = cell(row, "NRO. DE CASA")
        cesfam = cell(row, "CESFAM INSCRITO")
        celular_1 = cell(row, "CELULAR 1", "CELULAR")
        celular_2 = cell(row, "CELULAR 2")
        prevision = base.normalize_prevision(cell(row, "PREVISIÓN"))
        request_prestacion = cell(row, "PRESTACIÓN SOLICITADA")
        antecedentes = cell(
            row,
            "Antecedentes complementarios o información clínica relevante que no puede ser expresada en las preguntas previas. Observaciones y detalles de las prestaciones. EN CASO DE EXÁMENES, ADJUNTAR ORDEN AL FINAL DE LA ENCUESTA.",
        )
        gestora = cell(row, "NOMBRE POSTULANTE (GESTOR(A))", "GESTORA ENCARGADA")
        attachment_url = cell(row, "DAU // EPICRISIS (ADJUNTAR) IDENTIFICAR ARCHIVO CON NOMBRE DE USUARIO", "EPICRISIS CON IND MÉDICA")
        usuario_o2 = cell(row, "USUARIO DE O2")
        submission_timestamp = parse_datetime_or_empty(row[0] if row else "")
        contact_norm = split_phone_candidates(celular_1, celular_2)

        dedupe_key_seed = "|".join(
            [
                submission_timestamp,
                rut_norm or base.canonical_text(nombre_completo),
                fecha_nacimiento,
                servicio,
                base.canonical_text(diagnostico),
                base.canonical_text(request_prestacion),
            ]
        )
        dedupe_key = core.stable_id("fdup", dedupe_key_seed)

        normalized_rows.append(
            {
                "form_submission_id": form_submission_id,
                "dedupe_key": dedupe_key,
                "form_source_count": "1",
                "source_files": path.name,
                "source_rows": f"{path.name}:{row_number}",
                "submission_timestamp": submission_timestamp,
                "rut_raw": rut_raw,
                "rut_norm": rut_norm,
                "rut_valido": "1" if rut_norm else "0",
                "nombres": nombres,
                "apellido_paterno": apellido_paterno,
                "apellido_materno": apellido_materno,
                "apellidos": apellidos or canonical_name(apellido_paterno, apellido_materno),
                "nombre_completo": nombre_completo,
                "fecha_nacimiento": fecha_nacimiento,
                "edad_reportada": edad,
                "sexo": sexo,
                "servicio_origen_solicitud": servicio,
                "diagnostico": diagnostico,
                "direccion": direccion,
                "nro_casa": nro_casa,
                "cesfam": cesfam,
                "celular_1": celular_1,
                "celular_2": celular_2,
                "prevision": prevision,
                "request_prestacion": request_prestacion,
                "antecedentes": antecedentes,
                "gestora": gestora,
                "attachment_url": attachment_url,
                "usuario_o2": usuario_o2,
                "source_authority": "formulario_hodom",
                "contact_norm": contact_norm,
            }
        )

    return raw_rows, normalized_rows


def dedupe_form_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row["dedupe_key"]].append(row)

    deduped: list[dict[str, str]] = []
    for key, group in grouped.items():
        ordered = sorted(group, key=lambda row: (row["submission_timestamp"], row["form_submission_id"]))
        selected = dict(ordered[0])
        selected["form_source_count"] = str(len(group))
        selected["source_files"] = "; ".join(sorted({row["source_files"] for row in group}))
        selected["source_rows"] = "; ".join(sorted({row["source_rows"] for row in group}))
        deduped.append(selected)
    return deduped


def parse_discharge_workbook(path: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    raw_rows: list[dict[str, str]] = []
    normalized_rows: list[dict[str, str]] = []

    def looks_like_discharge_header(values: tuple[Any, ...] | list[Any]) -> bool:
        normalized = [base.canonical_text("" if value is None else str(value)) for value in values]
        non_empty = [value for value in normalized if value]
        if not non_empty:
            return False
        has_fecha_ingreso = any(value in {"FECHA INGRESO", "FECHA DE INGRESO"} for value in non_empty)
        has_fecha_egreso = any(value in {"FECHA EGRESO", "FECHA DE EGRESO"} for value in non_empty)
        has_rut = "RUT" in non_empty
        has_name = any(value in {"NOMBRE", "NOMBRE USUARIO", "APELLIDOS", "0"} for value in non_empty)
        second_column_blank = len(normalized) > 1 and normalized[1] == ""
        return has_fecha_ingreso and has_fecha_egreso and has_rut and (has_name or second_column_blank)

    for ws in wb.worksheets:
        preview = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
        header_row_number = None
        header_cells: list[str] = []
        for idx, preview_row in enumerate(preview, start=1):
            if looks_like_discharge_header(preview_row):
                header_row_number = idx
                header_cells = [base.normalize_whitespace("" if value is None else str(value)) for value in preview_row]
                break
        if header_row_number is None:
            continue

        canonical_header = {base.canonical_text(name): idx for idx, name in enumerate(header_cells)}

        def get_value(row: tuple[Any, ...], *keys: str) -> str:
            for key in keys:
                idx = canonical_header.get(base.canonical_text(key))
                if idx is not None and idx < len(row):
                    return base.normalize_whitespace(str(row[idx] or ""))
            return ""

        def get_raw(row: tuple[Any, ...], *keys: str) -> Any:
            for key in keys:
                idx = canonical_header.get(base.canonical_text(key))
                if idx is not None and idx < len(row):
                    return row[idx]
            return None

        blank_streak = 0
        for row_number, row in enumerate(ws.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
            if not any(value not in (None, "") for value in row):
                blank_streak += 1
                if blank_streak >= 3:
                    break
                continue
            blank_streak = 0
            raw_json = json.dumps([None if value is None else str(value) for value in row], ensure_ascii=False)
            discharge_row_id = core.stable_id("rawalta", path.name, ws.title, row_number)
            raw_rows.append(
                {
                    "discharge_row_id": discharge_row_id,
                    "source_file": path.name,
                    "source_sheet": ws.title,
                    "source_row_number": str(row_number),
                    "raw_json": raw_json,
                }
            )

            fecha_ingreso = parse_date_or_empty(get_raw(row, "FECHA DE INGRESO", "FECHA INGRESO"))
            fecha_egreso = parse_date_or_empty(get_raw(row, "FECHA DE EGRESO", "FECHA EGRESO"))
            motivo_egreso = get_value(row, "MOTIVO DE EGRESO", "OBSERVACIONES")
            diagnostico = get_value(row, "DIAGNÓSTICO", "DIAGNOSTICO DE EGRESO", "DIAGNÓSTICO DE EGRESO")
            nombre = get_value(row, "NOMBRE USUARIO", "NOMBRE", "0")
            apellidos = get_value(row, "APELLIDOS")
            if not nombre and len(row) > 2:
                guessed_name = base.normalize_whitespace("" if row[1] is None else str(row[1]))
                guessed_rut = base.normalize_whitespace("" if row[2] is None else str(row[2]))
                if guessed_name and guessed_rut:
                    nombre = guessed_name
            nombre_completo = canonical_name(nombre, "", "", apellidos)
            rut_raw = get_value(row, "RUT")
            comuna = normalize_commune(get_value(row, "COMUNA", "CONSULTORIO DERIVACION"))
            direccion = get_value(row, "DIRECCIÓN", "DOMICILIO")
            rut_norm = base.normalize_rut(rut_raw)

            if not any([fecha_ingreso, fecha_egreso, nombre_completo, rut_raw]):
                continue

            dedupe_key_seed = "|".join(
                [
                    fecha_ingreso,
                    fecha_egreso,
                    base.canonical_text(nombre_completo),
                    rut_norm or base.normalize_whitespace(rut_raw),
                    base.canonical_text(motivo_egreso),
                    base.canonical_text(diagnostico),
                ]
            )
            dedupe_key = core.stable_id("adup", dedupe_key_seed)

            normalized_rows.append(
                {
                    "discharge_event_id": core.stable_id("alta", path.name, ws.title, row_number),
                    "dedupe_key": dedupe_key,
                    "source_file": path.name,
                    "source_sheet": ws.title,
                    "source_row_number": str(row_number),
                    "fecha_ingreso": fecha_ingreso,
                    "fecha_egreso": fecha_egreso,
                    "motivo_egreso": motivo_egreso,
                    "diagnostico": diagnostico,
                    "nombre_completo": nombre_completo,
                    "rut_raw": rut_raw,
                    "rut_norm": rut_norm,
                    "rut_valido": "1" if rut_norm else "0",
                    "comuna": comuna,
                    "direccion_o_comuna": direccion or comuna,
                    "source_authority": "planilla_altas",
                }
            )

    grouped: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    for row in normalized_rows:
        grouped[row["dedupe_key"]].append(row)
    deduped: list[dict[str, str]] = []
    for _, group in grouped.items():
        selected = sorted(group, key=lambda row: (row["fecha_egreso"], row["discharge_event_id"]))[0]
        deduped.append(selected)

    return raw_rows, deduped


def fetch_establishment_reference() -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    fetched_at = datetime.now().isoformat(timespec="seconds")
    reference_snapshots: list[dict[str, str]] = []

    with tempfile.NamedTemporaryFile(prefix="deis_", delete=False) as tmp:
        pdf_path = Path(tmp.name)
    try:
        content, mode = fetch_bytes_with_ssl_fallback(DEIS_REFERENCE_URL, timeout=60)
        pdf_path.write_bytes(content)
        reference_snapshots.append(
            {
                "reference_snapshot_id": core.stable_id("ref", "deis_establecimientos"),
                "reference_type": "establishment_pdf",
                "source_url": DEIS_REFERENCE_URL,
                "fetched_at": fetched_at,
                "status": "OK",
                "notes": "Fuente oficial DEIS parseada con pdftotext -layout." if mode == "ssl_verified" else "Fuente DEIS obtenida con fallback verify=False por problema de certificados locales.",
            }
        )
    except RequestException as exc:
        cached_rows = read_existing_csv_rows(ENRICHED_CACHE_DIR / "establishment_reference.csv")
        if cached_rows:
            reference_snapshots.append(
                {
                    "reference_snapshot_id": core.stable_id("ref", "deis_establecimientos_cache"),
                    "reference_type": "establishment_csv_cache",
                    "source_url": str(ENRICHED_CACHE_DIR / "establishment_reference.csv"),
                    "fetched_at": fetched_at,
                    "status": "CACHE_FALLBACK",
                    "notes": f"Fallo descarga DEIS ({type(exc).__name__}); se reutiliza snapshot local existente.",
                }
            )
            return reference_snapshots, cached_rows
        raise

    proc = subprocess.run(
        ["pdftotext", "-layout", str(pdf_path), "-"],
        check=True,
        capture_output=True,
        text=True,
    )
    lines = proc.stdout.splitlines()

    def split_dependencia_y_nombre(raw_value: str) -> tuple[str, str]:
        for prefix in ("Municipal", "Servicio de Salud", "Particular Subvencionado", "Particular", "Servicio"):
            if raw_value == prefix:
                return raw_value, ""
            if raw_value.startswith(prefix + " "):
                return prefix, raw_value[len(prefix) :].strip()
        return "", raw_value

    def split_location_parts(parts: list[str]) -> tuple[str, str, str, str]:
        via_tokens = ["Calle", "Avenida", "Camino", "Pasaje", "Ruta"]
        if len(parts) >= 4:
            return parts[0], parts[1], parts[2], " ".join(parts[3:])
        if len(parts) == 3:
            return parts[0], parts[1], parts[2], ""
        if len(parts) == 2:
            combined = parts[0]
            for token in via_tokens:
                marker = f" {token}"
                if marker in combined:
                    comuna, via = combined.split(marker, 1)
                    return comuna.strip(), token, parts[1], via.strip()
            return parts[0], "", parts[1], ""
        if len(parts) == 1:
            return parts[0], "", "", ""
        return "", "", "", ""

    rows: list[dict[str, str]] = []
    for line in lines:
        stripped = line.strip()
        if not stripped.startswith("117"):
            continue
        parts = [part.strip() for part in base.re.split(r"\s{2,}", stripped) if part.strip()]
        if len(parts) < 5:
            continue
        code_service = parts[0]
        tokens = code_service.split(maxsplit=1)
        if len(tokens) != 2:
            continue
        codigo_deis, servicio_salud = tokens
        tipo_establecimiento = parts[1] if len(parts) > 1 else ""
        tail = parts[2:]
        dependencia = ""
        nombre_oficial = ""
        if tail:
            dependencia, nombre_oficial = split_dependencia_y_nombre(tail[0])
            if dependencia and nombre_oficial:
                tail = tail[1:]
            elif dependencia and not nombre_oficial:
                nombre_oficial = tail[1] if len(tail) > 1 else ""
                tail = tail[2:]
            else:
                nombre_oficial = tail[0]
                tail = tail[1:]
        comuna, via, numero, direccion = split_location_parts(tail)
        comuna = normalize_commune(comuna)
        rows.append(
            {
                "establishment_id": core.stable_id("est", codigo_deis),
                "codigo_deis": codigo_deis,
                "servicio_salud": servicio_salud,
                "tipo_establecimiento": tipo_establecimiento,
                "dependencia": dependencia,
                "nombre_oficial": nombre_oficial,
                "comuna": comuna,
                "via": via,
                "numero": numero,
                "direccion": direccion,
                "region": "Nuble",
                "source_url": DEIS_REFERENCE_URL,
                "fetched_at": fetched_at,
            }
        )

    return reference_snapshots, rows


def build_locality_reference(
    address_rows: list[dict[str, str]],
    form_rows: list[dict[str, str]],
    discharge_rows: list[dict[str, str]],
    establishment_rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    fetched_at = datetime.now().isoformat(timespec="seconds")
    entries: dict[tuple[str, str], dict[str, str]] = {}
    aliases: list[dict[str, str]] = []

    def register_locality(name: str, comuna: str, territory_type: str, source_url: str, alias: str = "") -> None:
        locality_name = base.normalize_whitespace(name)
        comuna_name = base.normalize_whitespace(comuna)
        if not locality_name:
            return
        key = (base.canonical_text(locality_name), base.canonical_text(comuna_name))
        if key not in entries:
            entries[key] = {
                "locality_id": core.stable_id("loc", locality_name, comuna_name),
                "nombre_oficial": locality_name,
                "comuna": comuna_name,
                "provincia": "Punilla" if comuna_name in PUNILLA_COMMUNES else "",
                "region": "Nuble",
                "territory_type": territory_type,
                "latitud": "",
                "longitud": "",
                "crs": "EPSG:4326",
                "source_url": source_url,
                "source_priority": "official_page",
                "fetched_at": fetched_at,
            }
        if alias and base.canonical_text(alias) != base.canonical_text(locality_name):
            aliases.append(
                {
                    "locality_alias_id": core.stable_id("lca", locality_name, comuna_name, alias),
                    "locality_id": entries[key]["locality_id"],
                    "alias": base.normalize_whitespace(alias),
                    "source_url": source_url,
                }
            )

    for row in establishment_rows:
        address_hint = row["direccion"]
        if any(token in address_hint.upper() for token in ["ALDEA", "CASERIO", "SECTOR", "PUEBLO"]):
            locality_name = (
                address_hint.replace("Aldea de", "")
                .replace("Aldea", "")
                .replace("Caserio de", "")
                .replace("Caserio", "")
                .replace("Sector", "")
                .replace("Pueblo de", "")
                .replace("Pueblo", "")
            )
            register_locality(locality_name, row["comuna"], "RURAL", DEIS_REFERENCE_URL, address_hint)
        register_locality(row["comuna"], row["comuna"], "URBANO", INE_CENSO_NUBLE_URL)

    for row in address_rows:
        register_locality(row.get("comuna", ""), row.get("comuna", ""), row.get("territory_type", ""), INE_GEODATA_URL)

    for row in form_rows:
        register_locality(row.get("cesfam", ""), "", "SANITARY", INE_GEODATA_URL, row.get("direccion", ""))

    for row in discharge_rows:
        register_locality(row.get("comuna", ""), row.get("comuna", ""), "URBANO", INE_CENSO_NUBLE_URL, row.get("direccion_o_comuna", ""))

    return list(entries.values()), aliases


def find_snapshot_file(stem: str) -> Path | None:
    for ext in (".geojson", ".json", ".csv"):
        candidate = REFERENCE_SNAPSHOT_DIR / f"{stem}{ext}"
        if candidate.exists():
            return candidate
    fallback_candidates = sorted(
        [
            path
            for path in REFERENCE_SNAPSHOT_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in {".geojson", ".json", ".csv"}
        ]
    ) if REFERENCE_SNAPSHOT_DIR.exists() else []
    if len(fallback_candidates) == 1:
        return fallback_candidates[0]
    for candidate in fallback_candidates:
        if candidate.stem.lower() == "features":
            return candidate
    return None


def read_manual_snapshot(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ";"
            rows = list(csv.DictReader(handle, delimiter=delimiter))
        normalized: list[dict[str, str]] = []
        for row in rows:
            locality_name = (
                row.get("nombre")
                or row.get("nombre_oficial")
                or row.get("localidad")
                or row.get("LOCALIDAD")
                or row.get("entidad")
                or row.get("ENTIDAD")
                or row.get("name")
                or ""
            )
            comuna = row.get("comuna") or row.get("COMUNA") or row.get("NOMBRE COMUNA") or row.get("nom_comuna") or row.get("municipio") or ""
            provincia = row.get("provincia") or row.get("PROVINCIA") or ""
            region = row.get("region") or row.get("REGION") or ""
            lat = row.get("latitud") or row.get("lat") or row.get("y") or ""
            lon = row.get("longitud") or row.get("lon") or row.get("lng") or row.get("x") or ""
            normalized.append(
                {
                    "nombre": base.normalize_whitespace(locality_name),
                    "comuna": normalize_commune(comuna),
                    "provincia": base.normalize_whitespace(provincia),
                    "region": base.normalize_whitespace(region),
                    "localidad": base.normalize_whitespace(row.get("LOCALIDAD") or row.get("localidad") or ""),
                    "entidad": base.normalize_whitespace(row.get("ENTIDAD") or row.get("entidad") or ""),
                    "categoria": base.normalize_whitespace(row.get("CATEGORIA") or row.get("categoria") or ""),
                    "latitud": base.normalize_whitespace(str(lat)),
                    "longitud": base.normalize_whitespace(str(lon)),
                    "source_file": path.name,
                }
            )
        return normalized

    data = json.loads(path.read_text(encoding="utf-8"))
    features = data.get("features", [])
    normalized = []
    for feature in features:
        props = feature.get("properties", {}) or {}
        locality_name = (
            props.get("nombre")
            or props.get("NOMBRE")
            or props.get("nom_local")
            or props.get("NOM_LOC")
            or props.get("localidad")
            or props.get("entidad")
            or props.get("NOM_ENTIDAD")
            or props.get("nom_entidad")
            or ""
        )
        comuna = (
            props.get("comuna")
            or props.get("COMUNA")
            or props.get("nom_comuna")
            or props.get("NOM_COMUNA")
            or ""
        )
        lat, lon = centroid_from_geometry(feature.get("geometry") or {})
        normalized.append(
            {
                "nombre": base.normalize_whitespace(locality_name),
                "comuna": normalize_commune(comuna),
                "provincia": base.normalize_whitespace(props.get("PROVINCIA") or props.get("provincia") or ""),
                "region": base.normalize_whitespace(props.get("REGION") or props.get("region") or ""),
                "localidad": base.normalize_whitespace(props.get("LOCALIDAD") or props.get("localidad") or ""),
                "entidad": base.normalize_whitespace(props.get("ENTIDAD") or props.get("entidad") or ""),
                "categoria": base.normalize_whitespace(props.get("CATEGORIA") or props.get("categoria") or ""),
                "latitud": lat,
                "longitud": lon,
                "source_file": path.name,
            }
        )
    return normalized


def apply_manual_locality_snapshots(
    localities: list[dict[str, str]],
    aliases: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    by_name, by_alias = build_locality_lookup(localities, aliases)
    snapshots = [
        ("ine_localidad_rural_nuble", INE_LOCALIDAD_RURAL_NUBLE_URL),
        ("ine_entidad_rural_nuble", INE_ENTIDAD_RURAL_NUBLE_URL),
    ]
    reference_rows: list[dict[str, str]] = []

    for stem, source_url in snapshots:
        snapshot = find_snapshot_file(stem)
        if not snapshot:
            reference_rows.append(
                {
                    "reference_snapshot_id": core.stable_id("ref", stem, "manual"),
                    "reference_type": f"{stem}_manual_snapshot",
                    "source_url": source_url,
                    "fetched_at": datetime.now().isoformat(timespec="seconds"),
                    "status": "MISSING_SNAPSHOT",
                    "notes": f"No existe snapshot manual en {REFERENCE_SNAPSHOT_DIR / (stem + '.[geojson|json|csv]')}",
                }
            )
            continue

        reference_rows.append(
            {
                "reference_snapshot_id": core.stable_id("ref", stem, "manual"),
                "reference_type": f"{stem}_manual_snapshot",
                "source_url": source_url,
                "fetched_at": datetime.now().isoformat(timespec="seconds"),
                "status": "IMPORTED",
                "notes": f"Snapshot manual importado desde {snapshot}",
            }
        )

        for row in read_manual_snapshot(snapshot):
            official_name = normalize_locality_name(row.get("entidad") or row.get("nombre") or row.get("localidad"))
            comuna_key = base.canonical_text(normalize_commune(row["comuna"]))
            key = (official_name, comuna_key)
            locality = by_name.get(key) or by_alias.get(key)
            if not locality:
                locality = {
                    "locality_id": core.stable_id("loc", official_name, comuna_key),
                    "nombre_oficial": base.normalize_whitespace(row.get("entidad") or row.get("nombre") or row.get("localidad")),
                    "comuna": normalize_commune(row.get("comuna", "")),
                    "provincia": base.normalize_whitespace(row.get("provincia") or ("Punilla" if normalize_commune(row.get("comuna", "")) in PUNILLA_COMMUNES else "")),
                    "region": base.normalize_whitespace(row.get("region") or "REGIÓN DE ÑUBLE"),
                    "territory_type": "RURAL",
                    "latitud": row.get("latitud", ""),
                    "longitud": row.get("longitud", ""),
                    "crs": "EPSG:4326",
                    "source_url": source_url,
                    "source_priority": "manual_official_snapshot",
                    "fetched_at": datetime.now().isoformat(timespec="seconds"),
                }
                localities.append(locality)
                by_name[key] = locality
            elif row["latitud"] and row["longitud"]:
                locality["latitud"] = row["latitud"]
                locality["longitud"] = row["longitud"]
                locality["source_priority"] = "manual_official_snapshot"
                locality["source_url"] = source_url

            for alias_value in {row.get("localidad", ""), row.get("entidad", ""), row.get("nombre", "")}:
                alias_text = base.normalize_whitespace(alias_value)
                if not alias_text:
                    continue
                alias_key = (base.canonical_text(alias_text), comuna_key)
                if alias_key not in by_alias and alias_key not in by_name:
                    alias_row = {
                        "locality_alias_id": core.stable_id("lca", locality["locality_id"], alias_text),
                        "locality_id": locality["locality_id"],
                        "alias": alias_text,
                        "source_url": source_url,
                    }
                    aliases.append(alias_row)
                    by_alias[alias_key] = locality

    return localities, reference_rows


def load_census24_geodb_localities() -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    fetched_at = datetime.now().isoformat(timespec="seconds")
    if pyogrio is None or not GEODB_REFERENCE_PATH.exists():
        status = "MISSING_GDB" if not GEODB_REFERENCE_PATH.exists() else "MISSING_PYOGRIO"
        notes = (
            f"No existe {GEODB_REFERENCE_PATH}"
            if not GEODB_REFERENCE_PATH.exists()
            else "pyogrio no está disponible en el runtime actual"
        )
        return [], [], [
            {
                "reference_snapshot_id": core.stable_id("ref", "censo24_geodb"),
                "reference_type": "censo24_gdb",
                "source_url": str(GEODB_REFERENCE_PATH),
                "fetched_at": fetched_at,
                "status": status,
                "notes": notes,
            }
        ]

    layers = {
        "Localidades_CPV24": {
            "name_field": "LOCALIDAD",
            "alias_fields": ["LOCALIDAD"],
            "territory_type": "RURAL",
        },
        "Aldeas_CPV24": {
            "name_field": "ENTIDAD",
            "alias_fields": ["ENTIDAD"],
            "territory_type": "RURAL",
        },
        "Entidades_CPV24": {
            "name_field": "ENTIDAD",
            "alias_fields": ["LOCALIDAD", "ENTIDAD", "DISTRITO"],
            "territory_type": "RURAL",
        },
    }

    locality_rows: list[dict[str, str]] = []
    locality_alias_rows: list[dict[str, str]] = []
    snapshots = [
        {
            "reference_snapshot_id": core.stable_id("ref", "censo24_gdb"),
            "reference_type": "censo24_gdb",
            "source_url": str(GEODB_REFERENCE_PATH),
            "fetched_at": fetched_at,
            "status": "IMPORTED",
            "notes": "Cartografía Censo 2024 R16 leída desde FileGDB oficial.",
        }
    ]
    if GEODB_DICTIONARY_PATH.exists():
        snapshots.append(
            {
                "reference_snapshot_id": core.stable_id("ref", "censo24_dictionary"),
                "reference_type": "censo24_dictionary",
                "source_url": str(GEODB_DICTIONARY_PATH),
                "fetched_at": fetched_at,
                "status": "IMPORTED",
                "notes": "Diccionario de variables geográficas Censo 2024 disponible.",
            }
        )

    seen_localities: set[tuple[str, str]] = set()
    seen_aliases: set[tuple[str, str, str]] = set()

    for layer_name, config in layers.items():
        gdf = pyogrio.read_dataframe(str(GEODB_REFERENCE_PATH), layer=layer_name)
        rep_points = gdf.geometry.representative_point()
        for idx, row in gdf.iterrows():
            name = base.normalize_whitespace(str(row.get(config["name_field"], "") or ""))
            comuna = normalize_commune(str(row.get("COMUNA", "") or ""))
            provincia = base.normalize_whitespace(str(row.get("PROVINCIA", "") or ""))
            region = base.normalize_whitespace(str(row.get("REGION", "") or ""))
            if not name or not comuna:
                continue
            key = (normalize_locality_name(name), base.canonical_text(comuna))
            if key not in seen_localities:
                point = rep_points.iloc[idx]
                locality_rows.append(
                    {
                        "locality_id": core.stable_id("loc", key[0], key[1]),
                        "nombre_oficial": name,
                        "comuna": comuna,
                        "provincia": provincia,
                        "region": region,
                        "territory_type": config["territory_type"],
                        "latitud": f"{point.y:.6f}" if point is not None else "",
                        "longitud": f"{point.x:.6f}" if point is not None else "",
                        "crs": str(gdf.crs) if gdf.crs else "EPSG:4674",
                        "source_url": str(GEODB_REFERENCE_PATH),
                        "source_priority": "official_gdb",
                        "fetched_at": fetched_at,
                    }
                )
                seen_localities.add(key)

            locality_id = core.stable_id("loc", key[0], key[1])
            for alias_field in config["alias_fields"]:
                alias_value = base.normalize_whitespace(str(row.get(alias_field, "") or ""))
                alias_key = (base.canonical_text(alias_value), base.canonical_text(comuna), locality_id)
                if alias_value and alias_key not in seen_aliases and base.canonical_text(alias_value) != key[0]:
                    locality_alias_rows.append(
                        {
                            "locality_alias_id": core.stable_id("lca", locality_id, alias_value),
                            "locality_id": locality_id,
                            "alias": alias_value,
                            "source_url": str(GEODB_REFERENCE_PATH),
                        }
                    )
                    seen_aliases.add(alias_key)

    return locality_rows, locality_alias_rows, snapshots


def build_establishment_lookup(establishments: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    for row in establishments:
        keys = {
            base.canonical_text(row["nombre_oficial"]),
            base.canonical_text(f"{row['nombre_oficial']} {row['comuna']}"),
        }
        for key in keys:
            if key:
                lookup[key] = row
    for alias, code in ESTABLISHMENT_ALIAS_MAP.items():
        establishment = next((row for row in establishments if row["codigo_deis"] == code), None)
        if establishment:
            lookup[alias] = establishment
    return lookup


def build_locality_establishment_map(establishments: list[dict[str, str]]) -> tuple[dict[tuple[str, str], dict[str, str]], dict[str, dict[str, str]]]:
    locality_map: dict[tuple[str, str], dict[str, str]] = {}
    commune_default: dict[str, dict[str, str]] = {}
    commune_candidates: defaultdict[str, list[dict[str, str]]] = defaultdict(list)

    for row in establishments:
        comuna = row["comuna"]
        name = row["nombre_oficial"]
        name_key = base.canonical_text(name)
        comuna_key = base.canonical_text(comuna)
        if "POSTA DE SALUD RURAL" in name_key:
            locality_name = (
                name_key.replace("POSTA DE SALUD RURAL", "")
                .replace("( NIQUEN )", "")
                .replace("( NINHUE )", "")
                .strip()
            )
            if locality_name:
                locality_map[(locality_name, comuna_key)] = row
        elif "CENTRO DE SALUD FAMILIAR" in name_key:
            commune_candidates[comuna_key].append(row)

    for comuna_key, candidates in commune_candidates.items():
        if len(candidates) == 1:
            commune_default[comuna_key] = candidates[0]

    return locality_map, commune_default


def resolve_establishment(raw_cesfam: str, lookup: dict[str, dict[str, str]]) -> tuple[str, str]:
    key = normalize_establishment_alias(raw_cesfam)
    if not key or key in {"OTRO", "SIN REFERENCIA", "SIN REFERENCIAS", "NO", "000", "00", "-"}:
        return "", "0.0"
    exact = lookup.get(key)
    if exact:
        return exact["establishment_id"], "1.0"
    best_id = ""
    best_score = 0.0
    for alias_key, row in lookup.items():
        score = similarity(alias_key, key)
        if score > best_score:
            best_score = score
            best_id = row["establishment_id"]
    if best_score >= 0.88:
        return best_id, f"{best_score:.2f}"
    return "", "0.0"


def resolve_establishment_from_locality(
    locality_id: str,
    locality_rows: list[dict[str, str]],
    locality_aliases: list[dict[str, str]],
    locality_est_map: dict[tuple[str, str], dict[str, str]],
    commune_default: dict[str, dict[str, str]],
) -> tuple[str, str]:
    if not locality_id:
        return "", "0.0"
    locality_by_id = {row["locality_id"]: row for row in locality_rows}
    locality = locality_by_id.get(locality_id)
    if not locality:
        return "", "0.0"
    key = (normalize_locality_name(locality["nombre_oficial"]), base.canonical_text(locality["comuna"]))
    match = locality_est_map.get(key)
    if match:
        return match["establishment_id"], "0.9"
    comuna_key = base.canonical_text(locality["comuna"])
    default_est = commune_default.get(comuna_key)
    if default_est and locality.get("territory_type") == "URBANO":
        return default_est["establishment_id"], "0.6"
    return "", "0.0"


def build_locality_lookup(localities: list[dict[str, str]], aliases: list[dict[str, str]]) -> tuple[dict[tuple[str, str], dict[str, str]], dict[tuple[str, str], dict[str, str]]]:
    by_name: dict[tuple[str, str], dict[str, str]] = {}
    by_alias: dict[tuple[str, str], dict[str, str]] = {}
    locality_by_id = {row["locality_id"]: row for row in localities}
    for row in localities:
        key = (base.canonical_text(row["nombre_oficial"]), base.canonical_text(row["comuna"]))
        by_name[key] = row
    for alias in aliases:
        locality = locality_by_id.get(alias["locality_id"])
        if not locality:
            continue
        key = (base.canonical_text(alias["alias"]), base.canonical_text(locality["comuna"]))
        by_alias[key] = locality
    return by_name, by_alias


def resolve_locality(raw_address: str, raw_comuna: str, by_name: dict[tuple[str, str], dict[str, str]], by_alias: dict[tuple[str, str], dict[str, str]]) -> tuple[str, str, str]:
    comuna = normalize_commune(raw_comuna)
    address_text = base.canonical_text(raw_address)
    comuna_key = base.canonical_text(comuna)

    for token, (locality_name, locality_comuna) in LOCALITY_ALIAS_HINTS.items():
        if token in address_text and (not locality_comuna or locality_comuna == comuna or not comuna):
            key = (base.canonical_text(locality_name), base.canonical_text(locality_comuna or comuna))
            locality = by_name.get(key) or by_alias.get(key)
            if locality:
                return locality["locality_id"], "1.0", "RESOLVED"

    for (alias_name, alias_comuna), locality in by_alias.items():
        if alias_comuna == comuna_key and alias_name and alias_name in address_text:
            return locality["locality_id"], "0.9", "RESOLVED"

    for (name_key, name_comuna), locality in by_name.items():
        if name_comuna == comuna_key and name_key and name_key in address_text:
            return locality["locality_id"], "0.8", "RESOLVED"

    comuna_locality = by_name.get((comuna_key, comuna_key))
    if comuna_locality:
        return comuna_locality["locality_id"], "0.3", "PARTIAL"
    return "", "0.0", "UNRESOLVED"


def load_intermediate_outputs(base_dir: Path) -> dict[str, list[dict[str, str]]]:
    names = [
        "patient_master.csv",
        "episode.csv",
        "patient_address.csv",
        "data_quality_issue.csv",
        "normalized_row.csv",
    ]
    return {name[:-4]: read_csv(base_dir / name) for name in names}


def build_baseline_episode_lookup(
    episodes: list[dict[str, str]],
    patients: list[dict[str, str]],
    addresses: list[dict[str, str]],
) -> list[dict[str, str]]:
    patient_by_id = {row["patient_id"]: row for row in patients}
    address_by_patient: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    for address in addresses:
        address_by_patient[address["patient_id"]].append(address)

    rows: list[dict[str, str]] = []
    for episode in episodes:
        patient = patient_by_id.get(episode["patient_id"], {})
        primary_address = address_by_patient.get(episode["patient_id"], [{}])[0]
        rows.append(
            {
                **episode,
                "patient_rut": patient.get("rut", ""),
                "patient_name": patient.get("nombre_completo", ""),
                "patient_birth": patient.get("fecha_nacimiento_date", ""),
                "patient_sexo": patient.get("sexo", ""),
                "patient_comuna": patient.get("comuna", "") or primary_address.get("comuna", ""),
                "patient_cesfam": patient.get("cesfam", "") or primary_address.get("cesfam", ""),
                "patient_address": patient.get("domicilio", "") or primary_address.get("full_address_raw", ""),
            }
        )
    return rows


def build_episode_indices(episodes: list[dict[str, str]]) -> dict[str, dict[str, list[dict[str, str]]]]:
    indices: dict[str, defaultdict[str, list[dict[str, str]]]] = {
        "rut": defaultdict(list),
        "name": defaultdict(list),
        "service": defaultdict(list),
        "month": defaultdict(list),
    }
    for episode in episodes:
        if episode.get("patient_rut"):
            indices["rut"][episode["patient_rut"]].append(episode)
        if episode.get("patient_name"):
            indices["name"][base.canonical_text(episode["patient_name"])].append(episode)
        if episode.get("servicio_origen"):
            indices["service"][episode["servicio_origen"]].append(episode)
        for field in ("fecha_ingreso", "fecha_egreso"):
            date_value = episode.get(field, "")
            if date_value:
                indices["month"][date_value[:7]].append(episode)
    return indices


def similarity(a: str, b: str) -> float:
    return difflib.SequenceMatcher(None, base.canonical_text(a), base.canonical_text(b)).ratio()


def request_flags(text: str) -> set[str]:
    t = (text or "").upper()
    flags: set[str] = set()
    if "ENDOVENOSO" in t or "TTO EV" in t:
        flags.add("TTO_EV")
    if "CURACIÓN" in t or "CURACION" in t:
        flags.add("CURACIONES")
    if "EXÁMEN" in t or "EXAMEN" in t:
        flags.add("TOMA_MUESTRAS")
    if "KNT" in t or "KINES" in t or "REHABILITACIÓN" in t or "REHABILITACION" in t:
        flags.add("KNT")
    if "FONOAUDIO" in t:
        flags.add("FONO")
    if "EVALUACIÓN MÉDICA" in t or "EVALUACION MEDICA" in t:
        flags.add("MEDICO")
    if "DISPOSITIVOS" in t or "S. FOLEY" in t or "SNG" in t or "NELATON" in t:
        flags.add("INVASIVOS")
    return flags


def prestation_matches_episode(request_text: str, episode_id: str, reqs_by_episode: dict[str, set[str]], profs_by_episode: dict[str, set[str]]) -> bool:
    flags = request_flags(request_text)
    reqs = reqs_by_episode.get(episode_id, set())
    profs = profs_by_episode.get(episode_id, set())
    if "TTO_EV" in flags and "TTO_EV" in reqs:
        return True
    if "CURACIONES" in flags and "CURACIONES" in reqs:
        return True
    if "TOMA_MUESTRAS" in flags and "TOMA_MUESTRAS" in reqs:
        return True
    if "INVASIVOS" in flags and "ELEMENTOS_INVASIVOS" in reqs:
        return True
    if "KNT" in flags and ("KNT" in profs or "KINESIOLOGIA" in profs):
        return True
    if "FONO" in flags and ("FONO" in profs or "FONOAUDIOLOGIA" in profs):
        return True
    if "MEDICO" in flags and "MEDICO" in profs:
        return True
    return False


def candidate_patient_key(name: str, rut_norm: str, birth_date: str, contact_norm: str) -> tuple[str, str]:
    if rut_norm:
        return f"rut:{rut_norm}", "rut"
    name_slug = base.slug_text(name)
    if name_slug and birth_date:
        return f"nombre-fecha:{name_slug}:{birth_date.replace('-', '')}", "nombre_fecha"
    if name_slug and contact_norm:
        first_phone = contact_norm.split(";")[0].strip()
        return f"nombre-contacto:{name_slug}:{first_phone}", "nombre_contacto"
    return f"nombre:{name_slug}" if name_slug else f"legacy:{core.stable_id('legacy', name, birth_date)}", "legacy"


def candidate_episode_pool(
    candidate: dict[str, str],
    indices: dict[str, dict[str, list[dict[str, str]]]],
    candidate_type: str,
    fallback_episodes: list[dict[str, str]],
) -> list[dict[str, str]]:
    pool: list[dict[str, str]] = []
    seen: set[str] = set()

    def add(items: list[dict[str, str]]) -> None:
        for item in items:
            if item["episode_id"] in seen:
                continue
            seen.add(item["episode_id"])
            pool.append(item)

    if candidate.get("rut_norm"):
        add(indices["rut"].get(candidate["rut_norm"], []))
    if candidate.get("nombre_completo"):
        add(indices["name"].get(base.canonical_text(candidate["nombre_completo"]), []))
    if candidate_type == "form" and candidate.get("servicio_origen_solicitud"):
        add(indices["service"].get(candidate["servicio_origen_solicitud"], []))

    date_value = candidate.get("submission_timestamp", "")[:7] if candidate_type == "form" else (candidate.get("fecha_egreso", "")[:7] or candidate.get("fecha_ingreso", "")[:7])
    if date_value:
        add(indices["month"].get(date_value, []))

    return pool or fallback_episodes


def score_form_against_episode(candidate: dict[str, str], episode: dict[str, str]) -> int:
    score = 0
    if candidate["rut_norm"] and candidate["rut_norm"] == episode.get("patient_rut"):
        score += 70
    if candidate["fecha_nacimiento"] and candidate["fecha_nacimiento"] == episode.get("patient_birth"):
        score += 25
    if base.canonical_text(candidate["nombre_completo"]) == base.canonical_text(episode.get("patient_name", "")):
        score += 20
    if candidate["servicio_origen_solicitud"] and candidate["servicio_origen_solicitud"] == episode.get("servicio_origen"):
        score += 10
    if similarity(candidate["diagnostico"], episode.get("diagnostico_principal_texto", "")) >= 0.60:
        score += 10
    if candidate["cesfam"] and candidate["cesfam"] == episode.get("patient_cesfam"):
        score += 10
    elif candidate["direccion"] and episode.get("patient_comuna") and base.canonical_text(candidate["direccion"]).find(base.canonical_text(episode.get("patient_comuna", ""))) >= 0:
        score += 5
    requested_at = candidate["submission_timestamp"][:10] if candidate["submission_timestamp"] else ""
    if requested_at and episode.get("fecha_ingreso"):
        try:
            delta = abs((datetime.strptime(requested_at, "%Y-%m-%d").date() - datetime.strptime(episode["fecha_ingreso"], "%Y-%m-%d").date()).days)
        except ValueError:
            delta = None
        if delta is not None:
            if delta <= 7:
                score += 10
            elif delta <= 14:
                score += 5
    return score


def build_form_candidate_suggestions(
    candidate: dict[str, str],
    episodes: list[dict[str, str]],
    indices: dict[str, dict[str, list[dict[str, str]]]],
) -> list[tuple[int, dict[str, str]]]:
    suggestions = []
    for episode in candidate_episode_pool(candidate, indices, "form", episodes):
        suggestions.append((score_form_against_episode(candidate, episode), episode))
    suggestions.sort(key=lambda item: (item[0], item[1].get("fecha_ingreso", ""), item[1]["episode_id"]), reverse=True)
    return suggestions


def match_form_to_episode(candidate: dict[str, str], episodes: list[dict[str, str]], indices: dict[str, dict[str, list[dict[str, str]]]]) -> dict[str, Any]:
    suggestions = build_form_candidate_suggestions(candidate, episodes, indices)
    if suggestions:
        best_score, best_episode = suggestions[0]
    else:
        best_score, best_episode = -1, None

    match_status = "unresolved"
    if best_score >= 85:
        match_status = "matched_exact"
    elif best_score >= 65:
        match_status = "matched_probable"
    return {"episode": best_episode, "score": best_score, "status": match_status}


def match_discharge_to_episode(candidate: dict[str, str], episodes: list[dict[str, str]], indices: dict[str, dict[str, list[dict[str, str]]]]) -> dict[str, Any]:
    best_score = -1
    best_episode: dict[str, str] | None = None
    for episode in candidate_episode_pool(candidate, indices, "discharge", episodes):
        score = 0
        if candidate["rut_norm"] and candidate["rut_norm"] == episode.get("patient_rut"):
            score += 70
        if base.canonical_text(candidate["nombre_completo"]) == base.canonical_text(episode.get("patient_name", "")):
            score += 20
        if candidate["fecha_egreso"] and candidate["fecha_egreso"] == episode.get("fecha_egreso"):
            score += 15
        if candidate["fecha_ingreso"] and candidate["fecha_ingreso"] == episode.get("fecha_ingreso"):
            score += 10
        if similarity(candidate["diagnostico"], episode.get("diagnostico_principal_texto", "")) >= 0.60:
            score += 10
        if candidate["comuna"] and base.canonical_text(candidate["comuna"]) == base.canonical_text(episode.get("patient_comuna", "")):
            score += 10
        if score > best_score:
            best_score = score
            best_episode = episode

    match_status = "unresolved"
    if best_score >= 85:
        match_status = "matched_exact"
    elif best_score >= 65:
        match_status = "matched_probable"
    return {"episode": best_episode, "score": best_score, "status": match_status}


def rescue_priority(date_iso: str) -> str:
    if not date_iso:
        return "LOW"
    if "2025-10" <= date_iso[:7] <= "2025-12":
        return "HIGH"
    if date_iso[:7] == "2026-01":
        return "MEDIUM"
    return "LOW"


def should_materialize_rescue(date_iso: str) -> bool:
    return rescue_priority(date_iso) in {"HIGH", "MEDIUM"}


def should_materialize_form_rescue(form: dict[str, str], patient_exists: bool) -> bool:
    requested_date = form["submission_timestamp"][:10] if form["submission_timestamp"] else ""
    if should_materialize_rescue(requested_date):
        return True
    if patient_exists and (
        form.get("request_prestacion")
        or form.get("diagnostico")
        or form.get("servicio_origen_solicitud")
    ):
        return True
    return False


def build_enriched_outputs(
    intermediate_dir: Path,
    output_dir: Path,
    form_paths: list[Path],
    discharge_path: Path,
) -> dict[str, list[dict[str, str]]]:
    intermediate = load_intermediate_outputs(intermediate_dir)
    baseline_patients = intermediate["patient_master"]
    baseline_episodes = intermediate["episode"]
    baseline_addresses = intermediate["patient_address"]
    baseline_quality = intermediate["data_quality_issue"]
    baseline_normalized = intermediate["normalized_row"]
    episode_lookup = build_baseline_episode_lookup(baseline_episodes, baseline_patients, baseline_addresses)
    episode_indices = build_episode_indices(episode_lookup)

    raw_form_rows: list[dict[str, str]] = []
    normalized_form_rows: list[dict[str, str]] = []
    for path in form_paths:
        if not path.exists():
            continue
        raw_rows, normalized_rows = parse_form_workbook(path)
        raw_form_rows.extend(raw_rows)
        normalized_form_rows.extend(normalized_rows)
    normalized_form_rows = dedupe_form_rows(normalized_form_rows)

    raw_discharge_rows, normalized_discharge_rows = parse_discharge_workbook(discharge_path)
    reference_snapshots, establishment_rows = fetch_establishment_reference()
    locality_rows, locality_alias_rows = build_locality_reference(
        baseline_addresses,
        normalized_form_rows,
        normalized_discharge_rows,
        establishment_rows,
    )
    gdb_locality_rows, gdb_locality_alias_rows, gdb_snapshots = load_census24_geodb_localities()
    if gdb_locality_rows:
        existing = {(normalize_locality_name(row["nombre_oficial"]), base.canonical_text(row["comuna"])): row for row in locality_rows}
        for row in gdb_locality_rows:
            key = (normalize_locality_name(row["nombre_oficial"]), base.canonical_text(row["comuna"]))
            if key in existing:
                existing[key].update(
                    {
                        "latitud": row["latitud"],
                        "longitud": row["longitud"],
                        "crs": row["crs"],
                        "source_url": row["source_url"],
                        "source_priority": row["source_priority"],
                        "provincia": row["provincia"] or existing[key].get("provincia", ""),
                    }
                )
            else:
                locality_rows.append(row)
        locality_alias_rows.extend(gdb_locality_alias_rows)
    locality_rows, manual_snapshot_rows = apply_manual_locality_snapshots(locality_rows, locality_alias_rows)

    reqs_by_episode: defaultdict[str, set[str]] = defaultdict(set)
    profs_by_episode: defaultdict[str, set[str]] = defaultdict(set)
    req_path = intermediate_dir / "episode_care_requirement.csv"
    prof_path = intermediate_dir / "episode_professional_need.csv"
    if req_path.exists():
        with req_path.open("r", encoding="utf-8", newline="") as handle:
            for row in csv.DictReader(handle):
                if row.get("is_active") == "1":
                    reqs_by_episode[row["episode_id"]].add(row["requirement_type"])
    if prof_path.exists():
        with prof_path.open("r", encoding="utf-8", newline="") as handle:
            for row in csv.DictReader(handle):
                profs_by_episode[row["episode_id"]].add(row["professional_type"])

    enriched_patients = {row["patient_id"]: dict(row) for row in baseline_patients}
    enriched_episodes = {row["episode_id"]: dict(row) for row in baseline_episodes}

    def ensure_enriched_patient(
        patient_id: str,
        patient_key: str,
        patient_key_strategy: str,
        seed: dict[str, str],
    ) -> dict[str, str]:
        patient = enriched_patients.get(patient_id)
        if patient is not None:
            return patient
        patient = {
            "patient_id": patient_id,
            "canonical_patient_key": patient_key,
            "identity_resolution_status": "AUTO" if seed.get("rut") else "AMBIGUOUS",
            "patient_key": patient_key,
            "patient_key_strategy": patient_key_strategy,
            "rut": seed.get("rut", ""),
            "rut_valido": "1" if seed.get("rut") else "0",
            "rut_raw": seed.get("rut_raw", ""),
            "nombres": seed.get("nombres", ""),
            "apellido_paterno": seed.get("apellido_paterno", ""),
            "apellido_materno": seed.get("apellido_materno", ""),
            "apellidos": seed.get("apellidos", ""),
            "nombre_completo": seed.get("nombre_completo", ""),
            "sexo": seed.get("sexo", ""),
            "fecha_nacimiento_date": seed.get("fecha_nacimiento_date", ""),
            "fecha_nacimiento_raw": seed.get("fecha_nacimiento_raw", ""),
            "edad_reportada": seed.get("edad_reportada", ""),
            "nacionalidad": seed.get("nacionalidad", ""),
            "nro_contacto": seed.get("nro_contacto", ""),
            "domicilio": seed.get("domicilio", ""),
            "comuna": seed.get("comuna", ""),
            "cesfam": seed.get("cesfam", ""),
            "episode_count": seed.get("episode_count", "1"),
            "source_files": seed.get("source_files", ""),
            "patient_resolution_status": "AUTO" if seed.get("rut") else "AMBIGUOUS",
            "canonical_address_text": seed.get("canonical_address_text", seed.get("domicilio", "")),
            "canonical_locality_id": seed.get("canonical_locality_id", ""),
            "canonical_establishment_id": seed.get("canonical_establishment_id", ""),
        }
        enriched_patients[patient_id] = patient
        return patient

    for episode in enriched_episodes.values():
        episode["episode_origin"] = "raw"
        episode["resolution_status"] = "AUTO"
        episode["match_status"] = "baseline"
        episode["match_score"] = ""
        episode["requested_at"] = ""
        episode["request_prestacion"] = ""
        episode["gestora"] = ""
        episode["form_source_count"] = "0"
        episode["codigo_deis"] = ""
        episode["establishment_id"] = ""
        episode["locality_id"] = ""
        episode["rescue_priority"] = ""

    # Aplicar correcciones manuales de manual_resolution.csv
    for res in read_manual_csv("manual_resolution.csv"):
        if res.get("applied") == "True":
            continue
        if res.get("action") != "correct":
            continue
        field = res.get("field_corrected", "")
        new_val = res.get("new_value", "")
        entity_id = res.get("item_id", "")
        if not field or not new_val or not entity_id:
            continue
        if entity_id in enriched_patients and field in enriched_patients[entity_id]:
            enriched_patients[entity_id][field] = new_val
        if entity_id in enriched_episodes and field in enriched_episodes[entity_id]:
            enriched_episodes[entity_id][field] = new_val

    field_provenance_rows: list[dict[str, str]] = []
    episode_request_rows: list[dict[str, str]] = []
    episode_discharge_rows: list[dict[str, str]] = []
    rescue_candidate_rows: list[dict[str, str]] = []
    form_candidate_rows: list[dict[str, str]] = []
    discharge_candidate_rows: list[dict[str, str]] = []
    match_review_queue_rows: list[dict[str, str]] = []
    quality_rows = list(baseline_quality)

    def record_provenance(entity_type: str, entity_id: str, field_name: str, selected_value: str, source: str, competing: dict[str, str], confidence: str) -> None:
        field_provenance_rows.append(
            {
                "field_provenance_id": core.stable_id("fp", entity_type, entity_id, field_name, source),
                "entity_type": entity_type,
                "entity_id": entity_id,
                "field_name": field_name,
                "selected_value": selected_value,
                "selected_source": source,
                "selected_confidence": confidence,
                "competing_values_json": json.dumps(competing, ensure_ascii=False),
            }
        )

    for form in normalized_form_rows:
        patient_key, patient_key_strategy = candidate_patient_key(
            form["nombre_completo"],
            form["rut_norm"],
            form["fecha_nacimiento"],
            form["contact_norm"],
        )
        matched = match_form_to_episode(form, episode_lookup, episode_indices)
        best_episode = matched["episode"]
        status = matched["status"]
        score = matched["score"]

        linked_episode_id = best_episode["episode_id"] if best_episode and status in {"matched_exact", "matched_probable"} else ""
        linked_patient_id = best_episode["patient_id"] if best_episode and status in {"matched_exact", "matched_probable"} else core.patient_id(patient_key)
        origin = "merged" if linked_episode_id else "form_rescued"
        resolution = "AUTO" if status == "matched_exact" else "REVIEW_REQUIRED" if status == "matched_probable" else "PROVISIONAL"
        requested_date = form["submission_timestamp"][:10] if form["submission_timestamp"] else ""
        rescue_prio = rescue_priority(requested_date)

        form_candidate_rows.append(
            {
                "candidate_id": core.stable_id("fc", form["form_submission_id"]),
                "candidate_type": "form_submission",
                "source_id": form["form_submission_id"],
                "patient_key": patient_key,
                "episode_id": linked_episode_id,
                "match_score": str(score),
                "match_status": status,
                "resolution_status": resolution,
                "episode_origin": origin,
                "rescue_priority": rescue_prio,
            }
        )

        if linked_episode_id:
            episode = enriched_episodes[linked_episode_id]
            episode["episode_origin"] = "merged"
            episode["resolution_status"] = resolution
            episode["match_status"] = status
            episode["match_score"] = str(score)
            if form["submission_timestamp"]:
                episode["requested_at"] = form["submission_timestamp"]
            if form["request_prestacion"]:
                competing = {"baseline": episode.get("request_prestacion", ""), "form": form["request_prestacion"]}
                episode["request_prestacion"] = form["request_prestacion"]
                record_provenance("episode", linked_episode_id, "request_prestacion", form["request_prestacion"], "formulario_hodom", competing, f"{score}")
            if form["gestora"]:
                competing = {"baseline": episode.get("gestora", ""), "form": form["gestora"]}
                episode["gestora"] = form["gestora"]
                record_provenance("episode", linked_episode_id, "gestora", form["gestora"], "formulario_hodom", competing, f"{score}")
            episode["form_source_count"] = form["form_source_count"]

            patient = ensure_enriched_patient(
                linked_patient_id,
                patient_key,
                patient_key_strategy,
                {
                    "rut": form["rut_norm"],
                    "rut_raw": form["rut_raw"],
                    "nombres": form["nombres"],
                    "apellido_paterno": form["apellido_paterno"],
                    "apellido_materno": form["apellido_materno"],
                    "apellidos": form["apellidos"],
                    "nombre_completo": form["nombre_completo"],
                    "sexo": form["sexo"],
                    "fecha_nacimiento_date": form["fecha_nacimiento"],
                    "fecha_nacimiento_raw": form["fecha_nacimiento"],
                    "edad_reportada": form["edad_reportada"],
                    "nro_contacto": form["contact_norm"],
                    "domicilio": form["direccion"],
                    "cesfam": form["cesfam"],
                    "source_files": form["source_files"],
                    "canonical_address_text": form["direccion"],
                },
            )
            for field_name, form_value in {
                "rut": form["rut_norm"],
                "rut_raw": form["rut_raw"],
                "nombres": form["nombres"],
                "apellido_paterno": form["apellido_paterno"],
                "apellido_materno": form["apellido_materno"],
                "apellidos": form["apellidos"],
                "nombre_completo": form["nombre_completo"],
                "sexo": form["sexo"],
                "fecha_nacimiento_date": form["fecha_nacimiento"],
                "edad_reportada": form["edad_reportada"],
                "nro_contacto": form["contact_norm"],
                "domicilio": form["direccion"],
                "cesfam": form["cesfam"],
            }.items():
                if not form_value:
                    continue
                baseline_value = patient.get(field_name, "")
                if baseline_value != form_value:
                    patient[field_name] = form_value
                    record_provenance("patient", linked_patient_id, field_name, form_value, "formulario_hodom", {"baseline": baseline_value, "form": form_value}, f"{score}")
            patient["patient_resolution_status"] = "AUTO" if patient.get("rut") else "AMBIGUOUS"
        else:
            patient_exists = any(
                patient.get("rut") == form["rut_norm"] for patient in enriched_patients.values() if form["rut_norm"]
            ) or any(
                base.canonical_text(patient.get("nombre_completo", "")) == base.canonical_text(form["nombre_completo"])
                for patient in enriched_patients.values()
            )
            materialize_rescue = should_materialize_form_rescue(form, patient_exists)
            suggestions = build_form_candidate_suggestions(form, episode_lookup, episode_indices)[:5]
            auto_close_recommended = False
            auto_close_episode_id = ""
            if suggestions:
                top_score, top_episode = suggestions[0]
                top_diag_similarity = similarity(form["diagnostico"], top_episode.get("diagnostico_principal_texto", ""))
                same_service = form["servicio_origen_solicitud"] == top_episode.get("servicio_origen", "")
                same_gestora = bool(form["gestora"]) and bool(top_episode.get("gestora")) and base.canonical_text(form["gestora"]) == base.canonical_text(top_episode.get("gestora", ""))
                if len(suggestions) == 1 and top_score >= 30:
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
                elif len(suggestions) > 1 and top_score >= 40 and top_score - suggestions[1][0] >= 15:
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
                elif (
                    top_score >= 30
                    and same_service
                    and prestation_matches_episode(form["request_prestacion"], top_episode["episode_id"], reqs_by_episode, profs_by_episode)
                ):
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
                elif (
                    top_score >= 20
                    and same_gestora
                    and same_service
                ):
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
                elif (
                    top_score >= 20
                    and form["submission_timestamp"]
                    and top_episode.get("fecha_ingreso")
                    and form["submission_timestamp"][:7] == top_episode.get("fecha_ingreso", "")[:7]
                    and same_service
                ):
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
                elif same_service and top_diag_similarity >= 0.95:
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
                elif same_service and top_diag_similarity >= 0.8:
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
                elif same_service and same_gestora and top_diag_similarity >= 0.4:
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
                elif same_service and same_gestora and top_diag_similarity >= 0.6:
                    auto_close_recommended = True
                    auto_close_episode_id = top_episode["episode_id"]
            for rank, (cand_score, cand_episode) in enumerate(suggestions, start=1):
                match_review_queue_rows.append(
                    {
                        "review_queue_id": core.stable_id("rvw", form["form_submission_id"], rank),
                        "form_submission_id": form["form_submission_id"],
                        "patient_id": linked_patient_id,
                        "patient_name": form["nombre_completo"],
                        "patient_rut": form["rut_norm"],
                        "submission_timestamp": form["submission_timestamp"],
                        "servicio_origen_solicitud": form["servicio_origen_solicitud"],
                        "diagnostico_form": form["diagnostico"],
                        "request_prestacion": form["request_prestacion"],
                        "gestora": form["gestora"],
                        "candidate_rank": str(rank),
                        "candidate_episode_id": cand_episode["episode_id"],
                        "candidate_episode_origin": cand_episode.get("episode_origin", ""),
                        "candidate_fecha_ingreso": cand_episode.get("fecha_ingreso", ""),
                        "candidate_fecha_egreso": cand_episode.get("fecha_egreso", ""),
                        "candidate_servicio_origen": cand_episode.get("servicio_origen", ""),
                        "candidate_diagnostico": cand_episode.get("diagnostico_principal_texto", ""),
                        "candidate_score": str(cand_score),
                        "auto_close_recommended": "1" if auto_close_recommended and cand_episode["episode_id"] == auto_close_episode_id else "0",
                    }
                )
            if auto_close_recommended and auto_close_episode_id:
                linked_episode_id = auto_close_episode_id
                linked_patient_id = enriched_episodes[linked_episode_id]["patient_id"]
                status = "matched_review_auto"
                score = max(score, next((cand_score for cand_score, cand_episode in suggestions if cand_episode["episode_id"] == auto_close_episode_id), score))
                resolution = "AUTO"
                origin = "merged"
                episode = enriched_episodes[linked_episode_id]
                episode["episode_origin"] = "merged"
                episode["resolution_status"] = resolution
                episode["match_status"] = status
                episode["match_score"] = str(score)
                if form["submission_timestamp"]:
                    episode["requested_at"] = form["submission_timestamp"]
                if form["request_prestacion"]:
                    competing = {"baseline": episode.get("request_prestacion", ""), "form": form["request_prestacion"]}
                    episode["request_prestacion"] = form["request_prestacion"]
                    record_provenance("episode", linked_episode_id, "request_prestacion", form["request_prestacion"], "formulario_hodom_auto_review", competing, f"{score}")
                if form["gestora"]:
                    competing = {"baseline": episode.get("gestora", ""), "form": form["gestora"]}
                    episode["gestora"] = form["gestora"]
                    record_provenance("episode", linked_episode_id, "gestora", form["gestora"], "formulario_hodom_auto_review", competing, f"{score}")
                episode["form_source_count"] = form["form_source_count"]

                patient = ensure_enriched_patient(
                    linked_patient_id,
                    patient_key,
                    patient_key_strategy,
                    {
                        "rut": form["rut_norm"],
                        "rut_raw": form["rut_raw"],
                        "nombres": form["nombres"],
                        "apellido_paterno": form["apellido_paterno"],
                        "apellido_materno": form["apellido_materno"],
                        "apellidos": form["apellidos"],
                        "nombre_completo": form["nombre_completo"],
                        "sexo": form["sexo"],
                        "fecha_nacimiento_date": form["fecha_nacimiento"],
                        "fecha_nacimiento_raw": form["fecha_nacimiento"],
                        "edad_reportada": form["edad_reportada"],
                        "nro_contacto": form["contact_norm"],
                        "domicilio": form["direccion"],
                        "cesfam": form["cesfam"],
                        "source_files": form["source_files"],
                        "canonical_address_text": form["direccion"],
                    },
                )
                for field_name, form_value in {
                    "rut": form["rut_norm"],
                    "rut_raw": form["rut_raw"],
                    "nombres": form["nombres"],
                    "apellido_paterno": form["apellido_paterno"],
                    "apellido_materno": form["apellido_materno"],
                    "apellidos": form["apellidos"],
                    "nombre_completo": form["nombre_completo"],
                    "sexo": form["sexo"],
                    "fecha_nacimiento_date": form["fecha_nacimiento"],
                    "edad_reportada": form["edad_reportada"],
                    "nro_contacto": form["contact_norm"],
                    "domicilio": form["direccion"],
                    "cesfam": form["cesfam"],
                }.items():
                    if not form_value:
                        continue
                    baseline_value = patient.get(field_name, "")
                    if baseline_value != form_value:
                        patient[field_name] = form_value
                        record_provenance("patient", linked_patient_id, field_name, form_value, "formulario_hodom_auto_review", {"baseline": baseline_value, "form": form_value}, f"{score}")
                patient["patient_resolution_status"] = "AUTO" if patient.get("rut") else "AMBIGUOUS"

                quality_rows.append(
                    {
                        "quality_issue_id": core.stable_id("dq", form["form_submission_id"], "UNMATCHED_FORM_SUBMISSION"),
                        "normalized_row_id": form["form_submission_id"],
                        "episode_id": linked_episode_id,
                        "issue_type": "UNMATCHED_FORM_SUBMISSION",
                        "severity": "MEDIUM",
                        "raw_value": form["nombre_completo"],
                        "suggested_value": linked_episode_id,
                        "status": "RESOLVED_AUTO",
                    }
                )
                episode_request_rows.append(
                    {
                        "episode_request_id": core.stable_id("erq", form["form_submission_id"]),
                        "form_submission_id": form["form_submission_id"],
                        "episode_id": linked_episode_id,
                        "patient_id": linked_patient_id,
                        "submission_timestamp": form["submission_timestamp"],
                        "request_prestacion": form["request_prestacion"],
                        "gestora": form["gestora"],
                        "servicio_origen_solicitud": form["servicio_origen_solicitud"],
                        "diagnostico": form["diagnostico"],
                        "match_score": str(score),
                        "match_status": status,
                        "resolution_status": resolution,
                        "episode_origin": origin,
                        "source_file": form["source_files"],
                        "source_rows": form["source_rows"],
                    }
                )
                continue
            rescue_id = core.stable_id("rescue", "form", form["form_submission_id"])
            rescue_candidate_rows.append(
                {
                    "rescue_candidate_id": rescue_id,
                    "candidate_type": "form_submission",
                    "source_id": form["form_submission_id"],
                    "episode_id": rescue_id if materialize_rescue else "",
                    "patient_id": linked_patient_id,
                    "requested_at": form["submission_timestamp"],
                    "fecha_ingreso": requested_date,
                    "fecha_egreso": "",
                    "nombre_completo": form["nombre_completo"],
                    "rut_norm": form["rut_norm"],
                    "diagnostico": form["diagnostico"],
                    "servicio_origen": form["servicio_origen_solicitud"],
                    "motivo_egreso": "",
                    "resolution_status": "PROVISIONAL",
                    "episode_origin": "form_rescued",
                    "rescue_priority": rescue_prio,
                }
            )
            ensure_enriched_patient(
                linked_patient_id,
                patient_key,
                patient_key_strategy,
                {
                    "rut": form["rut_norm"],
                    "rut_raw": form["rut_raw"],
                    "nombres": form["nombres"],
                    "apellido_paterno": form["apellido_paterno"],
                    "apellido_materno": form["apellido_materno"],
                    "apellidos": form["apellidos"],
                    "nombre_completo": form["nombre_completo"],
                    "sexo": form["sexo"],
                    "fecha_nacimiento_date": form["fecha_nacimiento"],
                    "fecha_nacimiento_raw": form["fecha_nacimiento"],
                    "edad_reportada": form["edad_reportada"],
                    "nro_contacto": form["contact_norm"],
                    "domicilio": form["direccion"],
                    "cesfam": form["cesfam"],
                    "source_files": form["source_files"],
                    "canonical_address_text": form["direccion"],
                },
            )
            if materialize_rescue:
                enriched_episodes[rescue_id] = {
                    "episode_id": rescue_id,
                    "patient_id": linked_patient_id,
                    "source_episode_key": rescue_id,
                    "record_uid": form["form_submission_id"],
                    "estado": "SOLICITADO",
                    "tipo_flujo": "FORMULARIO",
                    "fecha_ingreso": requested_date,
                    "fecha_egreso": "",
                    "dias_estadia_reportados": "",
                    "dias_estadia_calculados": "",
                    "motivo_egreso": "",
                    "motivo_derivacion": "",
                    "servicio_origen": form["servicio_origen_solicitud"],
                    "prevision": form["prevision"],
                    "barthel": "",
                    "categorizacion": "",
                    "usuario_o2": form["usuario_o2"],
                    "requerimiento_o2": "",
                    "diagnostico_principal_texto": form["diagnostico"],
                    "episode_status_quality": "REVIEW",
                    "duplicate_count": "1",
                    "episode_origin": "form_rescued",
                    "resolution_status": "PROVISIONAL",
                    "match_status": "rescued_provisional",
                    "match_score": str(score),
                    "requested_at": form["submission_timestamp"],
                    "request_prestacion": form["request_prestacion"],
                    "gestora": form["gestora"],
                    "form_source_count": form["form_source_count"],
                    "codigo_deis": "",
                    "establishment_id": "",
                    "locality_id": "",
                    "rescue_priority": rescue_prio,
                }
            quality_rows.append(
                {
                    "quality_issue_id": core.stable_id("dq", form["form_submission_id"], "UNMATCHED_FORM_SUBMISSION"),
                    "normalized_row_id": form["form_submission_id"],
                    "episode_id": linked_episode_id,
                    "issue_type": "UNMATCHED_FORM_SUBMISSION",
                    "severity": "HIGH" if rescue_prio == "HIGH" else "MEDIUM",
                    "raw_value": form["nombre_completo"],
                    "suggested_value": "",
                    "status": "RESCUED_PROVISIONAL" if materialize_rescue else "REVIEW_REQUIRED",
                }
            )

        episode_request_rows.append(
            {
                "episode_request_id": core.stable_id("erq", form["form_submission_id"]),
                "form_submission_id": form["form_submission_id"],
                "episode_id": linked_episode_id or rescue_id if not linked_episode_id and materialize_rescue else "",
                "patient_id": linked_patient_id,
                "submission_timestamp": form["submission_timestamp"],
                "request_prestacion": form["request_prestacion"],
                "gestora": form["gestora"],
                "servicio_origen_solicitud": form["servicio_origen_solicitud"],
                "diagnostico": form["diagnostico"],
                "match_score": str(score),
                "match_status": status,
                "resolution_status": resolution,
                "episode_origin": origin if linked_episode_id else "form_rescued",
                "source_file": form["source_files"],
                "source_rows": form["source_rows"],
            }
        )

    # Rebuild the episode lookup after form processing so discharge rows can close
    # provisional episodes rescued from forms in the same period.
    episode_lookup = build_baseline_episode_lookup(
        list(enriched_episodes.values()),
        list(enriched_patients.values()),
        baseline_addresses,
    )
    episode_indices = build_episode_indices(episode_lookup)

    for discharge in normalized_discharge_rows:
        patient_key, patient_key_strategy = candidate_patient_key(
            discharge["nombre_completo"],
            discharge["rut_norm"],
            "",
            "",
        )
        matched = match_discharge_to_episode(discharge, episode_lookup, episode_indices)
        best_episode = matched["episode"]
        status = matched["status"]
        score = matched["score"]
        linked_episode_id = best_episode["episode_id"] if best_episode and status in {"matched_exact", "matched_probable"} else ""
        linked_patient_id = best_episode["patient_id"] if best_episode and status in {"matched_exact", "matched_probable"} else core.patient_id(patient_key)
        event_date = discharge["fecha_egreso"] or discharge["fecha_ingreso"]
        rescue_prio = rescue_priority(event_date)

        discharge_candidate_rows.append(
            {
                "candidate_id": core.stable_id("dc", discharge["discharge_event_id"]),
                "candidate_type": "discharge_event",
                "source_id": discharge["discharge_event_id"],
                "patient_key": patient_key,
                "episode_id": linked_episode_id,
                "match_score": str(score),
                "match_status": status,
                "resolution_status": "AUTO" if status == "matched_exact" else "REVIEW_REQUIRED" if status == "matched_probable" else "PROVISIONAL",
                "episode_origin": "merged" if linked_episode_id else "alta_rescued",
                "rescue_priority": rescue_prio,
            }
        )

        rescue_episode_id = core.stable_id("rescue", "alta", discharge["discharge_event_id"])
        if linked_episode_id:
            episode = enriched_episodes[linked_episode_id]
            episode["episode_origin"] = "merged"
            episode["resolution_status"] = "AUTO" if status == "matched_exact" else "REVIEW_REQUIRED"
            episode["match_status"] = status
            episode["match_score"] = str(score)
            for field_name, value in {
                "fecha_ingreso": discharge["fecha_ingreso"],
                "fecha_egreso": discharge["fecha_egreso"],
                "motivo_egreso": discharge["motivo_egreso"],
                "diagnostico_principal_texto": discharge["diagnostico"],
            }.items():
                if not value:
                    continue
                baseline_value = episode.get(field_name, "")
                if baseline_value != value:
                    episode[field_name] = value
                    record_provenance("episode", linked_episode_id, field_name, value, "planilla_altas", {"baseline": baseline_value, "alta": value}, f"{score}")
            patient = ensure_enriched_patient(
                linked_patient_id,
                patient_key,
                patient_key_strategy,
                {
                    "rut": discharge["rut_norm"],
                    "rut_raw": discharge["rut_raw"],
                    "nombre_completo": discharge["nombre_completo"],
                    "domicilio": discharge["direccion_o_comuna"],
                    "comuna": discharge["comuna"],
                    "source_files": discharge["source_file"],
                    "canonical_address_text": discharge["direccion_o_comuna"],
                },
            )
            for field_name, value in {
                "rut": discharge["rut_norm"],
                "rut_raw": discharge["rut_raw"],
                "nombre_completo": discharge["nombre_completo"],
                "comuna": discharge["comuna"],
            }.items():
                if not value:
                    continue
                baseline_value = patient.get(field_name, "")
                if baseline_value != value:
                    patient[field_name] = value
                    record_provenance("patient", linked_patient_id, field_name, value, "planilla_altas", {"baseline": baseline_value, "alta": value}, f"{score}")
        else:
            rescue_candidate_rows.append(
                {
                    "rescue_candidate_id": rescue_episode_id,
                    "candidate_type": "discharge_event",
                    "source_id": discharge["discharge_event_id"],
                    "episode_id": rescue_episode_id if should_materialize_rescue(event_date) else "",
                    "patient_id": linked_patient_id,
                    "requested_at": "",
                    "fecha_ingreso": discharge["fecha_ingreso"],
                    "fecha_egreso": discharge["fecha_egreso"],
                    "nombre_completo": discharge["nombre_completo"],
                    "rut_norm": discharge["rut_norm"],
                    "diagnostico": discharge["diagnostico"],
                    "servicio_origen": "",
                    "motivo_egreso": discharge["motivo_egreso"],
                    "resolution_status": "PROVISIONAL",
                    "episode_origin": "alta_rescued",
                    "rescue_priority": rescue_prio,
                }
            )
            ensure_enriched_patient(
                linked_patient_id,
                patient_key,
                patient_key_strategy,
                {
                    "rut": discharge["rut_norm"],
                    "rut_raw": discharge["rut_raw"],
                    "nombre_completo": discharge["nombre_completo"],
                    "domicilio": discharge["direccion_o_comuna"],
                    "comuna": discharge["comuna"],
                    "source_files": discharge["source_file"],
                    "canonical_address_text": discharge["direccion_o_comuna"],
                },
            )
            enriched_episodes[rescue_episode_id] = {
                "episode_id": rescue_episode_id,
                "patient_id": linked_patient_id,
                "source_episode_key": rescue_episode_id,
                "record_uid": discharge["discharge_event_id"],
                "estado": "EGRESADO",
                "tipo_flujo": "ALTA",
                "fecha_ingreso": discharge["fecha_ingreso"],
                "fecha_egreso": discharge["fecha_egreso"],
                "dias_estadia_reportados": "",
                "dias_estadia_calculados": core.days_between(discharge["fecha_ingreso"], discharge["fecha_egreso"]),
                "motivo_egreso": discharge["motivo_egreso"],
                "motivo_derivacion": "",
                "servicio_origen": "",
                "prevision": "",
                "barthel": "",
                "categorizacion": "",
                "usuario_o2": "",
                "requerimiento_o2": "",
                "diagnostico_principal_texto": discharge["diagnostico"],
                "episode_status_quality": "REVIEW",
                "duplicate_count": "1",
                "episode_origin": "alta_rescued",
                "resolution_status": "PROVISIONAL",
                "match_status": "rescued_provisional",
                "match_score": str(score),
                "requested_at": "",
                "request_prestacion": "",
                "gestora": "",
                "form_source_count": "0",
                "codigo_deis": "",
                "establishment_id": "",
                "locality_id": "",
                "rescue_priority": rescue_prio,
            }
            quality_rows.append(
                {
                    "quality_issue_id": core.stable_id("dq", discharge["discharge_event_id"], "MISSING_EPISODE_IN_RAW"),
                    "normalized_row_id": discharge["discharge_event_id"],
                    "episode_id": linked_episode_id,
                    "issue_type": "MISSING_EPISODE_IN_RAW",
                    "severity": "HIGH" if rescue_prio == "HIGH" else "MEDIUM",
                    "raw_value": discharge["nombre_completo"],
                    "suggested_value": "",
                    "status": "RESCUED_PROVISIONAL",
                }
            )

        episode_discharge_rows.append(
            {
                "episode_discharge_id": core.stable_id("edc", discharge["discharge_event_id"]),
                "discharge_event_id": discharge["discharge_event_id"],
                "episode_id": linked_episode_id or rescue_episode_id if not linked_episode_id else "",
                "patient_id": linked_patient_id,
                "fecha_ingreso": discharge["fecha_ingreso"],
                "fecha_egreso": discharge["fecha_egreso"],
                "motivo_egreso": discharge["motivo_egreso"],
                "diagnostico": discharge["diagnostico"],
                "match_score": str(score),
                "match_status": status,
                "resolution_status": "AUTO" if status == "matched_exact" else "REVIEW_REQUIRED" if status == "matched_probable" else "PROVISIONAL",
                "episode_origin": "merged" if linked_episode_id else "alta_rescued",
                "source_file": discharge["source_file"],
                "source_sheet": discharge["source_sheet"],
                "source_row_number": discharge["source_row_number"],
            }
        )

    patient_rows_by_id = {row["patient_id"]: row for row in enriched_patients.values()}
    redundant_form_rescue_targets: dict[str, str] = {}
    episodes_by_identity_and_ingreso: defaultdict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for episode in enriched_episodes.values():
        patient = patient_rows_by_id.get(episode.get("patient_id", ""), {})
        identity_key = patient.get("rut", "") or base.canonical_text(patient.get("nombre_completo", ""))
        key = (identity_key, episode.get("fecha_ingreso", ""))
        if key[0] and key[1]:
            episodes_by_identity_and_ingreso[key].append(episode)

    for grouped_episodes in episodes_by_identity_and_ingreso.values():
        closed_candidates = [row for row in grouped_episodes if row.get("fecha_egreso")]
        redundant_forms = [
            row
            for row in grouped_episodes
            if row.get("episode_origin") == "form_rescued" and not row.get("fecha_egreso")
        ]
        if not closed_candidates or not redundant_forms:
            continue
        closed_candidates.sort(
            key=lambda row: (
                0 if row.get("episode_origin") == "merged" else 1 if row.get("episode_origin") == "alta_rescued" else 2,
                row.get("fecha_egreso", ""),
                row["episode_id"],
            )
        )
        target_episode = closed_candidates[0]
        for redundant in redundant_forms:
            if redundant["episode_id"] != target_episode["episode_id"]:
                redundant_form_rescue_targets[redundant["episode_id"]] = target_episode["episode_id"]

    if redundant_form_rescue_targets:
        for request_row in episode_request_rows:
            target_episode_id = redundant_form_rescue_targets.get(request_row.get("episode_id", ""))
            if not target_episode_id:
                continue
            request_row["episode_id"] = target_episode_id
            request_row["episode_origin"] = "merged"
            request_row["resolution_status"] = "AUTO"
            request_row["match_status"] = "matched_auto_discharge_close"

        for rescue_row in rescue_candidate_rows:
            target_episode_id = redundant_form_rescue_targets.get(rescue_row.get("episode_id", ""))
            if not target_episode_id:
                continue
            rescue_row["episode_id"] = target_episode_id
            rescue_row["episode_origin"] = "merged"
            rescue_row["resolution_status"] = "AUTO"

        for quality_row in quality_rows:
            target_episode_id = redundant_form_rescue_targets.get(quality_row.get("episode_id", ""))
            if not target_episode_id:
                continue
            quality_row["episode_id"] = target_episode_id
            if quality_row.get("issue_type") == "UNMATCHED_FORM_SUBMISSION":
                quality_row["suggested_value"] = target_episode_id
                quality_row["status"] = "RESOLVED_AUTO"

        for redundant_episode_id in redundant_form_rescue_targets:
            enriched_episodes.pop(redundant_episode_id, None)

    address_resolution_rows: list[dict[str, str]] = []
    establishment_lookup = build_establishment_lookup(establishment_rows)
    locality_establishment_map, commune_default_establishment = build_locality_establishment_map(establishment_rows)
    locality_by_name, locality_by_alias = build_locality_lookup(locality_rows, locality_alias_rows)
    for patient in enriched_patients.values():
        raw_address = patient.get("domicilio", "")
        raw_comuna = normalize_commune(patient.get("comuna", ""))
        raw_cesfam = patient.get("cesfam", "")
        establishment_id, establishment_confidence = resolve_establishment(raw_cesfam, establishment_lookup)
        locality_id, locality_confidence, locality_status = resolve_locality(raw_address, raw_comuna, locality_by_name, locality_by_alias)
        if not establishment_id and locality_id:
            establishment_id, establishment_confidence = resolve_establishment_from_locality(
                locality_id,
                locality_rows,
                locality_alias_rows,
                locality_establishment_map,
                commune_default_establishment,
            )
        resolution_status = "RESOLVED" if establishment_id and locality_status == "RESOLVED" else "PARTIAL" if (establishment_id or locality_status in {"RESOLVED", "PARTIAL"}) else "UNRESOLVED"
        address_resolution_rows.append(
            {
                "address_resolution_id": core.stable_id("adr", patient["patient_id"]),
                "patient_id": patient["patient_id"],
                "episode_id": "",
                "raw_address": raw_address,
                "raw_comuna": raw_comuna,
                "raw_cesfam": raw_cesfam,
                "resolved_establishment_id": establishment_id,
                "resolved_locality_id": locality_id,
                "establishment_match_confidence": establishment_confidence,
                "locality_match_confidence": locality_confidence,
                "resolution_status": resolution_status,
            }
        )
        patient["comuna"] = raw_comuna
        patient["canonical_establishment_id"] = establishment_id
        patient["canonical_locality_id"] = locality_id
        patient["canonical_address_text"] = raw_address
    episode_request_rows, rescue_candidate_rows, quality_rows, enriched_episodes = apply_match_review_overrides(
        episode_request_rows,
        rescue_candidate_rows,
        quality_rows,
        enriched_episodes,
        normalized_form_rows,
    )

    for patient in enriched_patients.values():
        patient["episode_count"] = str(sum(1 for episode in enriched_episodes.values() if episode["patient_id"] == patient["patient_id"]))
        if int(patient["episode_count"]) == 0:
            patient["patient_resolution_status"] = "SUBMISSION_ONLY"
            patient["identity_resolution_status"] = "SUBMISSION_ONLY"
            quality_rows.append(
                {
                    "quality_issue_id": core.stable_id("dq", patient["patient_id"], "PATIENT_WITHOUT_EPISODE"),
                    "normalized_row_id": "",
                    "episode_id": "",
                    "issue_type": "PATIENT_WITHOUT_EPISODE",
                    "severity": "MEDIUM",
                    "raw_value": patient.get("nombre_completo", ""),
                    "suggested_value": "",
                    "status": "REVIEW_REQUIRED",
                }
            )
        elif patient.get("rut"):
            patient["patient_resolution_status"] = "AUTO"
            patient["identity_resolution_status"] = "AUTO"
        else:
            patient["patient_resolution_status"] = "AMBIGUOUS"
            patient["identity_resolution_status"] = "AMBIGUOUS"

    address_resolution_rows = apply_address_review_overrides(address_resolution_rows, enriched_patients)
    enriched_patients, enriched_episodes = propagate_territorial_resolution(
        enriched_patients,
        enriched_episodes,
        establishment_rows,
    )
    quality_rows = rebuild_territorial_quality_issues(
        quality_rows,
        address_resolution_rows,
        locality_rows,
    )

    quality_rows = remediate_quality_issues(
        quality_rows,
        baseline_normalized,
        enriched_episodes,
        enriched_patients,
        normalized_form_rows,
    )
    quality_rows = apply_issue_review_overrides(quality_rows)
    quality_rows = dedupe_quality_issues(quality_rows)
    identity_review_queue_rows = build_identity_review_queue(
        quality_rows,
        enriched_episodes,
        enriched_patients,
    )

    months_of_interest = ["2025-10", "2025-11", "2025-12", "2026-01"]
    baseline_counts = Counter((row.get("fecha_ingreso") or row.get("fecha_egreso") or "")[:7] for row in baseline_episodes if (row.get("fecha_ingreso") or row.get("fecha_egreso")))
    matched_forms = Counter()
    matched_discharges = Counter()
    rescued_forms = Counter()
    rescued_discharges = Counter()
    for row in episode_request_rows:
        month = row["submission_timestamp"][:7] if row["submission_timestamp"] else ""
        if row["match_status"] in {"matched_exact", "matched_probable", "matched_manual"}:
            matched_forms[month] += 1
        elif row["episode_origin"] == "form_rescued":
            rescued_forms[month] += 1
    for row in episode_discharge_rows:
        month = row["fecha_egreso"][:7] if row["fecha_egreso"] else row["fecha_ingreso"][:7]
        if row["match_status"] in {"matched_exact", "matched_probable", "matched_manual"}:
            matched_discharges[month] += 1
        elif row["episode_origin"] == "alta_rescued":
            rescued_discharges[month] += 1
    enriched_counts = Counter((row.get("fecha_ingreso") or row.get("fecha_egreso") or "")[:7] for row in enriched_episodes.values() if (row.get("fecha_ingreso") or row.get("fecha_egreso")))
    reconciliation_rows = []
    for month in months_of_interest:
        baseline_count = baseline_counts.get(month, 0)
        enriched_count = enriched_counts.get(month, 0)
        reconciliation_rows.append(
            {
                "month": month,
                "baseline_episode_count": str(baseline_count),
                "matched_form_count": str(matched_forms.get(month, 0)),
                "matched_discharge_count": str(matched_discharges.get(month, 0)),
                "rescued_form_count": str(rescued_forms.get(month, 0)),
                "rescued_discharge_count": str(rescued_discharges.get(month, 0)),
                "enriched_episode_count": str(enriched_count),
                "delta_vs_baseline": str(enriched_count - baseline_count),
            }
        )

    outputs = {
        "raw_form_submission": raw_form_rows,
        "raw_discharge_sheet": raw_discharge_rows,
        "raw_reference_snapshot": reference_snapshots + gdb_snapshots + [
            {
                "reference_snapshot_id": core.stable_id("ref", "ine_geodata"),
                "reference_type": "locality_official_page",
                "source_url": INE_GEODATA_URL,
                "fetched_at": datetime.now().isoformat(timespec="seconds"),
                "status": "OK",
                "notes": "Pagina oficial de geodatos abiertos INE usada como fuente estructurante de referencia territorial.",
            },
            {
                "reference_snapshot_id": core.stable_id("ref", "ine_censo_nuble"),
                "reference_type": "censo_nuble_page",
                "source_url": INE_CENSO_NUBLE_URL,
                "fetched_at": datetime.now().isoformat(timespec="seconds"),
                "status": "OK",
                "notes": "Comunicado oficial INE sobre cartografia y base manzana-entidad Censo 2024.",
            },
            {
                "reference_snapshot_id": core.stable_id("ref", "ine_localidad_rural_nuble"),
                "reference_type": "localidad_rural_nuble_dataset_page",
                "source_url": INE_LOCALIDAD_RURAL_NUBLE_URL,
                "fetched_at": datetime.now().isoformat(timespec="seconds"),
                "status": "DISCOVERED",
                "notes": "Dataset oficial INE identificado, pero no se extrajeron aun coordenadas descargables de forma automatica.",
            },
            {
                "reference_snapshot_id": core.stable_id("ref", "ine_entidad_rural_nuble"),
                "reference_type": "entidad_rural_nuble_dataset_page",
                "source_url": INE_ENTIDAD_RURAL_NUBLE_URL,
                "fetched_at": datetime.now().isoformat(timespec="seconds"),
                "status": "DISCOVERED",
                "notes": "Dataset oficial INE identificado, pero no se extrajeron aun coordenadas descargables de forma automatica.",
            },
        ] + manual_snapshot_rows,
        "normalized_form_submission": normalized_form_rows,
        "normalized_discharge_event": normalized_discharge_rows,
        "normalized_establishment_reference": establishment_rows,
        "normalized_locality_reference": locality_rows,
        "normalized_form_episode_candidate": form_candidate_rows,
        "normalized_discharge_episode_candidate": discharge_candidate_rows,
        "patient_master": list(enriched_patients.values()),
        "episode_master": list(enriched_episodes.values()),
        "episode_request": episode_request_rows,
        "episode_discharge": episode_discharge_rows,
        "episode_rescue_candidate": rescue_candidate_rows,
        "field_provenance": field_provenance_rows,
        "match_review_queue": match_review_queue_rows,
        "identity_review_queue": identity_review_queue_rows,
        "establishment_reference": establishment_rows,
        "locality_reference": locality_rows,
        "locality_alias": locality_alias_rows,
        "address_resolution": address_resolution_rows,
        "data_quality_issue": quality_rows,
        "reconciliation_report": reconciliation_rows,
    }
    return outputs


def remediate_quality_issues(
    quality_rows: list[dict[str, str]],
    normalized_rows: list[dict[str, str]],
    episodes: dict[str, dict[str, str]],
    patients: dict[str, dict[str, str]],
    form_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    normalized_by_id = {row["normalized_row_id"]: row for row in normalized_rows}
    form_by_id = {row["form_submission_id"]: row for row in form_rows}
    forms_by_identity: defaultdict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in form_rows:
        forms_by_identity[(row.get("rut_norm", ""), base.canonical_text(row.get("nombre_completo", "")))].append(row)
    remediated: list[dict[str, str]] = []
    for issue in quality_rows:
        issue = dict(issue)
        episode = episodes.get(issue.get("episode_id", ""), {})
        patient = patients.get(episode.get("patient_id", ""), {})

        if issue["issue_type"] == "BIRTHDATE_AGE_MISMATCH" and patient:
            birth = patient.get("fecha_nacimiento_date", "")
            age = patient.get("edad_reportada", "")
            ref_date = episode.get("fecha_ingreso") or episode.get("fecha_egreso") or ""
            if birth and age and ref_date and base.age_matches_birth_date(birth, age, ref_date):
                issue["status"] = "RESOLVED_AUTO"
                issue["suggested_value"] = birth
            else:
                parsed_raw = base.parse_date(issue.get("raw_value", ""))
                if parsed_raw and age and ref_date and base.age_matches_birth_date(parsed_raw, age, ref_date):
                    patient["fecha_nacimiento_date"] = parsed_raw
                    issue["status"] = "RESOLVED_AUTO"
                    issue["suggested_value"] = parsed_raw
                else:
                    identity_key = (patient.get("rut", ""), base.canonical_text(patient.get("nombre_completo", "")))
                    candidate_births = sorted({row.get("fecha_nacimiento", "") for row in forms_by_identity.get(identity_key, []) if row.get("fecha_nacimiento")})
                    if len(candidate_births) == 1:
                        candidate_birth = candidate_births[0]
                        if ref_date:
                            try:
                                birth_dt = datetime.strptime(candidate_birth, "%Y-%m-%d").date()
                                ref_dt = datetime.strptime(ref_date, "%Y-%m-%d").date()
                                calc_age = ref_dt.year - birth_dt.year - ((ref_dt.month, ref_dt.day) < (birth_dt.month, birth_dt.day))
                            except ValueError:
                                calc_age = None
                            if calc_age is not None and 0 <= calc_age <= 120:
                                patient["fecha_nacimiento_date"] = candidate_birth
                                patient["edad_reportada"] = str(calc_age)
                                issue["status"] = "RESOLVED_AUTO"
                                issue["suggested_value"] = f"{candidate_birth}|edad={calc_age}"

        elif issue["issue_type"] == "DATE_PARSE_FAILED":
            normalized = normalized_by_id.get(issue.get("normalized_row_id", ""), {})
            if normalized:
                if issue["raw_value"] == normalized.get("fecha_nacimiento_raw") and patient.get("fecha_nacimiento_date"):
                    issue["status"] = "RESOLVED_AUTO"
                    issue["suggested_value"] = patient["fecha_nacimiento_date"]
                elif issue["raw_value"] == normalized.get("fecha_ingreso_raw") and episode.get("fecha_ingreso"):
                    issue["status"] = "RESOLVED_AUTO"
                    issue["suggested_value"] = episode["fecha_ingreso"]
                elif issue["raw_value"] == normalized.get("fecha_egreso_raw") and episode.get("fecha_egreso"):
                    issue["status"] = "RESOLVED_AUTO"
                    issue["suggested_value"] = episode["fecha_egreso"]
                else:
                    parsed_raw = base.parse_date(issue.get("raw_value", ""))
                    ref_date = episode.get("fecha_ingreso") or episode.get("fecha_egreso") or ""
                    age = patient.get("edad_reportada", "")
                    if parsed_raw and age and ref_date and base.age_matches_birth_date(parsed_raw, age, ref_date):
                        patient["fecha_nacimiento_date"] = parsed_raw
                        issue["status"] = "RESOLVED_AUTO"
                        issue["suggested_value"] = parsed_raw
            if issue["status"] == "OPEN" and patient.get("fecha_nacimiento_date"):
                issue["status"] = "RESOLVED_AUTO"
                issue["suggested_value"] = patient["fecha_nacimiento_date"]

        elif issue["issue_type"] == "UNMATCHED_FORM_SUBMISSION":
            form = form_by_id.get(issue.get("normalized_row_id", ""), {})
            if form:
                patient_known = any(
                    p.get("rut") == form.get("rut_norm") for p in patients.values() if form.get("rut_norm")
                ) or any(
                    base.canonical_text(p.get("nombre_completo", "")) == base.canonical_text(form.get("nombre_completo", ""))
                    for p in patients.values()
                )
                if patient_known and issue["status"] == "OPEN":
                    issue["status"] = "REVIEW_REQUIRED"
                    issue["suggested_value"] = "Paciente conocido, episodio no confirmado"

        remediated.append(issue)
    return remediated


def dedupe_quality_issues(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple[str, str, str, str, str]] = set()
    deduped: list[dict[str, str]] = []
    for row in rows:
        key = (
            row.get("issue_type", ""),
            row.get("episode_id", ""),
            row.get("normalized_row_id", ""),
            row.get("raw_value", ""),
            row.get("status", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def build_identity_review_queue(
    quality_rows: list[dict[str, str]],
    episodes: dict[str, dict[str, str]],
    patients: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    queue_rows: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for issue in quality_rows:
        if issue.get("status") != "OPEN":
            continue
        if issue.get("issue_type") not in {"DATE_PARSE_FAILED", "BIRTHDATE_AGE_MISMATCH"}:
            continue
        episode = episodes.get(issue.get("episode_id", ""), {})
        patient = patients.get(episode.get("patient_id", ""), {})
        key = (
            issue.get("issue_type", ""),
            issue.get("episode_id", ""),
            patient.get("patient_id", ""),
            issue.get("raw_value", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        queue_rows.append(
            {
                "identity_review_id": core.stable_id("idr", *key),
                "issue_type": issue.get("issue_type", ""),
                "patient_id": patient.get("patient_id", ""),
                "episode_id": issue.get("episode_id", ""),
                "patient_name": patient.get("nombre_completo", ""),
                "patient_rut": patient.get("rut", ""),
                "edad_reportada": patient.get("edad_reportada", ""),
                "fecha_nacimiento_actual": patient.get("fecha_nacimiento_date", ""),
                "issue_raw_value": issue.get("raw_value", ""),
                "suggested_value": issue.get("suggested_value", ""),
                "fecha_ingreso": episode.get("fecha_ingreso", ""),
                "fecha_egreso": episode.get("fecha_egreso", ""),
                "episode_origin": episode.get("episode_origin", ""),
                "severity": issue.get("severity", ""),
                "status": issue.get("status", ""),
            }
        )
    return queue_rows


def apply_issue_review_overrides(quality_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    overrides = {row.get("quality_issue_id", ""): row for row in read_manual_csv("issue_review_overrides.csv")}
    if not overrides:
        return quality_rows
    updated: list[dict[str, str]] = []
    for issue in quality_rows:
        row = dict(issue)
        override = overrides.get(row.get("quality_issue_id", ""))
        if override:
            decision = base.normalize_whitespace(override.get("review_decision"))
            approved = base.normalize_whitespace(override.get("approved_value"))
            notes = base.normalize_whitespace(override.get("review_notes"))
            if decision == "accept_suggestion":
                row["status"] = "RESOLVED_MANUAL"
                if row.get("suggested_value"):
                    row["suggested_value"] = row["suggested_value"]
            elif decision == "manual_fix":
                row["status"] = "RESOLVED_MANUAL"
                if approved:
                    row["suggested_value"] = approved
            elif decision == "ignore":
                row["status"] = "WAIVED"
            elif decision == "needs_clinical_review":
                row["status"] = "REVIEW_REQUIRED"
            if notes:
                row["suggested_value"] = approved or row.get("suggested_value", "")
                row["raw_value"] = row.get("raw_value", "")
        updated.append(row)
    return updated


def apply_address_review_overrides(
    address_rows: list[dict[str, str]],
    patients: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    overrides = {row.get("address_resolution_id", ""): row for row in read_manual_csv("address_review_overrides.csv")}
    if not overrides:
        return address_rows
    updated: list[dict[str, str]] = []
    for address in address_rows:
        row = dict(address)
        override = overrides.get(row.get("address_resolution_id", ""))
        if override:
            est_override = base.normalize_whitespace(override.get("override_establishment_id"))
            loc_override = base.normalize_whitespace(override.get("override_locality_id"))
            status_override = base.normalize_whitespace(override.get("override_status"))
            if est_override:
                row["resolved_establishment_id"] = est_override
                row["establishment_match_confidence"] = "1.0"
            if loc_override:
                row["resolved_locality_id"] = loc_override
                row["locality_match_confidence"] = "1.0"
            if status_override == "resolved_manual":
                row["resolution_status"] = "RESOLVED"
            elif status_override == "partial_manual":
                row["resolution_status"] = "PARTIAL"
            elif status_override == "discard_establishment":
                row["resolved_establishment_id"] = ""
                row["establishment_match_confidence"] = "0.0"
            elif status_override == "discard_locality":
                row["resolved_locality_id"] = ""
                row["locality_match_confidence"] = "0.0"
            patient = patients.get(row.get("patient_id", ""))
            if patient:
                if row.get("resolved_establishment_id"):
                    patient["canonical_establishment_id"] = row["resolved_establishment_id"]
                if row.get("resolved_locality_id"):
                    patient["canonical_locality_id"] = row["resolved_locality_id"]
        updated.append(row)
    return updated


def apply_match_review_overrides(
    episode_request_rows: list[dict[str, str]],
    rescue_candidate_rows: list[dict[str, str]],
    quality_rows: list[dict[str, str]],
    enriched_episodes: dict[str, dict[str, str]],
    normalized_form_rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]], dict[str, dict[str, str]]]:
    overrides = read_manual_csv("match_review_decisions.csv")
    if not overrides:
        return episode_request_rows, rescue_candidate_rows, quality_rows, enriched_episodes

    requests_by_form = {row["form_submission_id"]: row for row in episode_request_rows}
    rescues_by_source = {row["source_id"]: row for row in rescue_candidate_rows}
    forms_by_id = {row["form_submission_id"]: row for row in normalized_form_rows}

    for override in overrides:
        decision = base.normalize_whitespace(override.get("review_decision"))
        if not decision:
            continue
        form_id = base.normalize_whitespace(override.get("form_submission_id"))
        if not form_id:
            continue
        request_row = requests_by_form.get(form_id)
        form_row = forms_by_id.get(form_id, {})
        if not request_row:
            continue

        if decision == "link_candidate":
            selected_episode_id = base.normalize_whitespace(
                override.get("selected_episode_id") or override.get("candidate_episode_id")
            )
            if not selected_episode_id or selected_episode_id not in enriched_episodes:
                continue
            episode = enriched_episodes[selected_episode_id]
            request_row["episode_id"] = selected_episode_id
            request_row["patient_id"] = episode["patient_id"]
            request_row["match_status"] = "matched_manual"
            request_row["resolution_status"] = "MANUAL_REVIEW"
            request_row["episode_origin"] = "merged"
            request_row["match_score"] = request_row.get("match_score") or override.get("candidate_score", "")
            episode["episode_origin"] = "merged"
            episode["resolution_status"] = "MANUAL_REVIEW"
            episode["match_status"] = "matched_manual"
            if form_row.get("submission_timestamp"):
                episode["requested_at"] = form_row["submission_timestamp"]
            if form_row.get("request_prestacion"):
                episode["request_prestacion"] = form_row["request_prestacion"]
            if form_row.get("gestora"):
                episode["gestora"] = form_row["gestora"]
            if form_id in rescues_by_source:
                rescue = rescues_by_source[form_id]
                rescue["resolution_status"] = "CLOSED_MANUAL"
            for issue in quality_rows:
                if issue.get("normalized_row_id") == form_id and issue.get("issue_type") == "UNMATCHED_FORM_SUBMISSION":
                    issue["status"] = "RESOLVED_MANUAL"
                    issue["suggested_value"] = selected_episode_id

        elif decision == "create_rescue":
            rescue = rescues_by_source.get(form_id)
            rescue_id = rescue.get("episode_id", "") if rescue else ""
            if not rescue_id:
                rescue_id = core.stable_id("rescue", "form", form_id)
            if rescue_id not in enriched_episodes:
                enriched_episodes[rescue_id] = {
                    "episode_id": rescue_id,
                    "patient_id": request_row["patient_id"],
                    "source_episode_key": rescue_id,
                    "record_uid": form_id,
                    "estado": "SOLICITADO",
                    "tipo_flujo": "FORMULARIO",
                    "fecha_ingreso": (form_row.get("submission_timestamp", "") or "")[:10],
                    "fecha_egreso": "",
                    "dias_estadia_reportados": "",
                    "dias_estadia_calculados": "",
                    "motivo_egreso": "",
                    "motivo_derivacion": "",
                    "servicio_origen": form_row.get("servicio_origen_solicitud", ""),
                    "prevision": form_row.get("prevision", ""),
                    "barthel": "",
                    "categorizacion": "",
                    "usuario_o2": form_row.get("usuario_o2", ""),
                    "requerimiento_o2": "",
                    "diagnostico_principal_texto": form_row.get("diagnostico", ""),
                    "episode_status_quality": "REVIEW",
                    "duplicate_count": "1",
                    "episode_origin": "form_rescued",
                    "resolution_status": "PROVISIONAL",
                    "match_status": "rescued_manual",
                    "match_score": request_row.get("match_score", ""),
                    "requested_at": form_row.get("submission_timestamp", ""),
                    "request_prestacion": form_row.get("request_prestacion", ""),
                    "gestora": form_row.get("gestora", ""),
                    "form_source_count": form_row.get("form_source_count", "1"),
                    "codigo_deis": "",
                    "establishment_id": "",
                    "locality_id": "",
                    "rescue_priority": rescue_priority((form_row.get("submission_timestamp", "") or "")[:10]),
                }
            request_row["match_status"] = "rescued_manual"
            request_row["resolution_status"] = "PROVISIONAL"
            request_row["episode_origin"] = "form_rescued"
            request_row["episode_id"] = rescue_id
            if rescue:
                rescue["episode_id"] = rescue_id
                rescue["resolution_status"] = "PROVISIONAL"
            for issue in quality_rows:
                if issue.get("normalized_row_id") == form_id and issue.get("issue_type") == "UNMATCHED_FORM_SUBMISSION":
                    issue["status"] = "RESCUED_MANUAL"

        elif decision == "ignore":
            request_row["resolution_status"] = "WAIVED"
            for issue in quality_rows:
                if issue.get("normalized_row_id") == form_id and issue.get("issue_type") == "UNMATCHED_FORM_SUBMISSION":
                    issue["status"] = "WAIVED"

    return list(requests_by_form.values()), list(rescues_by_source.values()), quality_rows, enriched_episodes


def rebuild_territorial_quality_issues(
    quality_rows: list[dict[str, str]],
    address_rows: list[dict[str, str]],
    localities: list[dict[str, str]],
) -> list[dict[str, str]]:
    filtered = [
        row
        for row in quality_rows
        if row.get("issue_type") not in {"LOCALITY_UNRESOLVED", "ESTABLISHMENT_UNRESOLVED", "LOCALITY_COORDINATES_MISSING"}
    ]

    for address in address_rows:
        raw_cesfam = normalize_establishment_alias(address.get("raw_cesfam", ""))
        if (
            not address.get("resolved_establishment_id")
            and raw_cesfam
            and raw_cesfam not in {"OTRO", "SIN REFERENCIA", "SIN REFERENCIAS", "NO", "000", "00", "-"}
        ):
            filtered.append(
                {
                    "quality_issue_id": core.stable_id("dq", address["patient_id"], "ESTABLISHMENT_UNRESOLVED"),
                    "normalized_row_id": "",
                    "episode_id": "",
                    "issue_type": "ESTABLISHMENT_UNRESOLVED",
                    "severity": "MEDIUM",
                    "raw_value": address.get("raw_cesfam", ""),
                    "suggested_value": "",
                    "status": "OPEN",
                }
            )
        if not address.get("resolved_locality_id"):
            filtered.append(
                {
                    "quality_issue_id": core.stable_id("dq", address["patient_id"], "LOCALITY_UNRESOLVED"),
                    "normalized_row_id": "",
                    "episode_id": "",
                    "issue_type": "LOCALITY_UNRESOLVED",
                    "severity": "MEDIUM",
                    "raw_value": address.get("raw_address", "") or address.get("raw_comuna", ""),
                    "suggested_value": "",
                    "status": "OPEN",
                }
            )

    for locality in localities:
        if not locality.get("latitud") or not locality.get("longitud"):
            filtered.append(
                {
                    "quality_issue_id": core.stable_id("dq", locality["locality_id"], "LOCALITY_COORDINATES_MISSING"),
                    "normalized_row_id": "",
                    "episode_id": "",
                    "issue_type": "LOCALITY_COORDINATES_MISSING",
                    "severity": "MEDIUM",
                    "raw_value": locality["nombre_oficial"],
                    "suggested_value": "",
                    "status": "OPEN",
                }
            )

    return filtered


def propagate_territorial_resolution(
    patients: dict[str, dict[str, str]],
    episodes: dict[str, dict[str, str]],
    establishments: list[dict[str, str]],
) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    establishment_by_id = {row["establishment_id"]: row for row in establishments}
    for episode in episodes.values():
        patient = patients.get(episode.get("patient_id", ""), {})
        establishment_id = patient.get("canonical_establishment_id", "")
        locality_id = patient.get("canonical_locality_id", "")
        episode["establishment_id"] = establishment_id
        episode["locality_id"] = locality_id
        episode["codigo_deis"] = establishment_by_id.get(establishment_id, {}).get("codigo_deis", "")
    return patients, episodes


def write_sql(path: Path) -> None:
    sql = """CREATE TABLE IF NOT EXISTS raw_form_submission (
    form_submission_id TEXT PRIMARY KEY,
    source_file TEXT NOT NULL,
    source_sheet TEXT NOT NULL,
    source_row_number INTEGER NOT NULL,
    raw_json JSONB NOT NULL
);

CREATE TABLE IF NOT EXISTS raw_discharge_sheet (
    discharge_row_id TEXT PRIMARY KEY,
    source_file TEXT NOT NULL,
    source_sheet TEXT NOT NULL,
    source_row_number INTEGER NOT NULL,
    raw_json JSONB NOT NULL
);

CREATE TABLE IF NOT EXISTS raw_reference_snapshot (
    reference_snapshot_id TEXT PRIMARY KEY,
    reference_type TEXT NOT NULL,
    source_url TEXT NOT NULL,
    fetched_at TIMESTAMP NOT NULL,
    status TEXT NOT NULL,
    notes TEXT
);

CREATE TABLE IF NOT EXISTS normalized_form_submission (
    form_submission_id TEXT PRIMARY KEY,
    dedupe_key TEXT NOT NULL,
    form_source_count INTEGER NOT NULL,
    source_files TEXT NOT NULL,
    source_rows TEXT NOT NULL,
    submission_timestamp TIMESTAMP,
    rut_raw TEXT,
    rut_norm TEXT,
    rut_valido BOOLEAN,
    nombres TEXT,
    apellido_paterno TEXT,
    apellido_materno TEXT,
    apellidos TEXT,
    nombre_completo TEXT,
    fecha_nacimiento DATE,
    edad_reportada INTEGER,
    sexo TEXT,
    servicio_origen_solicitud TEXT,
    diagnostico TEXT,
    direccion TEXT,
    nro_casa TEXT,
    cesfam TEXT,
    celular_1 TEXT,
    celular_2 TEXT,
    prevision TEXT,
    request_prestacion TEXT,
    antecedentes TEXT,
    gestora TEXT,
    attachment_url TEXT,
    usuario_o2 TEXT,
    source_authority TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS normalized_discharge_event (
    discharge_event_id TEXT PRIMARY KEY,
    dedupe_key TEXT NOT NULL,
    source_file TEXT NOT NULL,
    source_sheet TEXT NOT NULL,
    source_row_number INTEGER NOT NULL,
    fecha_ingreso DATE,
    fecha_egreso DATE,
    motivo_egreso TEXT,
    diagnostico TEXT,
    nombre_completo TEXT,
    rut_raw TEXT,
    rut_norm TEXT,
    rut_valido BOOLEAN,
    comuna TEXT,
    direccion_o_comuna TEXT,
    source_authority TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS episode_master (
    episode_id TEXT PRIMARY KEY,
    patient_id TEXT NOT NULL,
    source_episode_key TEXT NOT NULL,
    record_uid TEXT NOT NULL,
    estado TEXT,
    tipo_flujo TEXT,
    fecha_ingreso DATE,
    fecha_egreso DATE,
    dias_estadia_reportados INTEGER,
    dias_estadia_calculados INTEGER,
    motivo_egreso TEXT,
    motivo_derivacion TEXT,
    servicio_origen TEXT,
    prevision TEXT,
    barthel TEXT,
    categorizacion TEXT,
    usuario_o2 TEXT,
    requerimiento_o2 TEXT,
    diagnostico_principal_texto TEXT,
    episode_status_quality TEXT,
    duplicate_count INTEGER,
    episode_origin TEXT,
    resolution_status TEXT,
    match_status TEXT,
    match_score NUMERIC,
    requested_at TIMESTAMP,
    request_prestacion TEXT,
    gestora TEXT,
    form_source_count INTEGER,
    codigo_deis TEXT,
    establishment_id TEXT,
    locality_id TEXT,
    rescue_priority TEXT
);
"""
    path.write_text(sql, encoding="utf-8")


def write_readme(path: Path, outputs: dict[str, list[dict[str, str]]]) -> None:
    lines = [
        "# Pipeline Enriquecido HODOM",
        "",
        "Integra formularios 2025/2026, planilla de altas y referencias oficiales para reconciliar, rescatar y preparar migracion nominal.",
        "",
        "## Fuentes adicionales",
        "",
        "- Formularios HODOM 2025/2026 desde `~/Downloads` o desde otras rutas entregadas al runner.",
        "- Planilla de altas 2024-2026 desde `input/reference/legacy_imports/planilla-altas-2024-2026.xlsx`.",
        "- Referencia oficial de establecimientos desde el PDF DEIS.",
        "- Referencia territorial estructurante desde las paginas oficiales INE/Censo 2024.",
        "",
        "## Conteos",
        "",
    ]
    for name, rows in outputs.items():
        lines.append(f"- `{name}.csv`: `{len(rows)}` filas")
    lines.extend(
        [
            "",
            "## Fuentes web usadas",
            "",
            f"- Establecimientos DEIS: {DEIS_REFERENCE_URL}",
            f"- INE Geodatos Abiertos: {INE_GEODATA_URL}",
            f"- INE Censo 2024 Nuble: {INE_CENSO_NUBLE_URL}",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_manifest(path: Path, outputs: dict[str, list[dict[str, str]]]) -> None:
    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "tables": {name: len(rows) for name, rows in outputs.items()},
    }
    path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Enriquece HODOM con formularios, altas y referencias oficiales.")
    parser.add_argument(
        "--intermediate-dir",
        type=Path,
        default=Path("output/spreadsheet/intermediate"),
        help="Directorio base del pipeline intermedio actual.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output/spreadsheet/enriched"),
        help="Directorio de salida del pipeline enriquecido.",
    )
    args = parser.parse_args()

    ensure_intermediate(Path(__file__).resolve().parent)
    outputs = build_enriched_outputs(args.intermediate_dir, args.output_dir, FORM_SOURCE_PATHS, DISCHARGE_SOURCE_PATH)

    column_map = {
        "raw_form_submission": RAW_FORM_SUBMISSION_COLUMNS,
        "raw_discharge_sheet": RAW_DISCHARGE_SHEET_COLUMNS,
        "raw_reference_snapshot": RAW_REFERENCE_SNAPSHOT_COLUMNS,
        "normalized_form_submission": NORMALIZED_FORM_SUBMISSION_COLUMNS,
        "normalized_discharge_event": NORMALIZED_DISCHARGE_EVENT_COLUMNS,
        "normalized_establishment_reference": ESTABLISHMENT_REFERENCE_COLUMNS,
        "normalized_locality_reference": LOCALITY_REFERENCE_COLUMNS,
        "normalized_form_episode_candidate": NORMALIZED_FORM_EPISODE_CANDIDATE_COLUMNS,
        "normalized_discharge_episode_candidate": NORMALIZED_DISCHARGE_EPISODE_CANDIDATE_COLUMNS,
        "patient_master": PATIENT_MASTER_COLUMNS,
        "episode_master": EPISODE_MASTER_COLUMNS,
        "episode_request": EPISODE_REQUEST_COLUMNS,
        "episode_discharge": EPISODE_DISCHARGE_COLUMNS,
        "episode_rescue_candidate": EPISODE_RESCUE_CANDIDATE_COLUMNS,
        "field_provenance": FIELD_PROVENANCE_COLUMNS,
        "match_review_queue": MATCH_REVIEW_QUEUE_COLUMNS,
        "identity_review_queue": IDENTITY_REVIEW_QUEUE_COLUMNS,
        "establishment_reference": ESTABLISHMENT_REFERENCE_COLUMNS,
        "locality_reference": LOCALITY_REFERENCE_COLUMNS,
        "locality_alias": LOCALITY_ALIAS_COLUMNS,
        "address_resolution": ADDRESS_RESOLUTION_COLUMNS,
        "data_quality_issue": DATA_QUALITY_ISSUE_COLUMNS,
        "reconciliation_report": RECONCILIATION_REPORT_COLUMNS,
    }

    args.output_dir.mkdir(parents=True, exist_ok=True)
    for name, rows in outputs.items():
        core.write_csv(args.output_dir / f"{name}.csv", rows, column_map[name])
    write_sql(args.output_dir / "hodom_enriched_postgres.sql")
    write_manifest(args.output_dir / "manifest.json", outputs)
    write_readme(args.output_dir / "README.md", outputs)

    print(f"Output dir: {args.output_dir.resolve()}")
    for name, rows in outputs.items():
        print(f"{name}: {len(rows)}")


if __name__ == "__main__":
    main()
