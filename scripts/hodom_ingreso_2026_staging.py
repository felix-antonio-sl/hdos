"""Parse and load INGRESOS 2026 DRIVE Excel into staging.hodom_ingreso_2026."""

import hashlib
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / ".tmp" / "drive-consolidation" / "raw"
EXCEL_FILE = RAW_DIR / "INGRESOS_2026_DRIVE.xlsx"
DRIVE_ID = "1qynoVzgF5a5qMdTVXhfM35zQC4QCtVqh0MNaSHD2_aQ"
OUT_SQL = ROOT / ".tmp" / "drive-migration" / "hodom-ingreso-2026-load.sql"

HEADER_ROW = 1
COLUMNS = [
    "row_num",
    "estado",
    "fecha_ingreso",
    "fecha_egreso",
    "dias_estada",
    "motivo_egreso",
    "nombres",
    "apellidos",
    "sexo",
    "edad",
    "fecha_nacimiento",
    "rut",
    "barthel",
    "prevision",
    "servicio_origen",
    "usuario_o2",
    "requerimiento_hodom_o2",
    "categorizacion",
    "diagnostico_egreso",
    "domicilio",
    "comuna",
    "cesfam",
    "nro_contacto",
    "nacionalidad",
    "enfermeria",
    "kinesiologia",
    "fonoaudiologia",
]
EXTRA_COLS = 2  # Some sheets have extra trailing columns


def norm_text(value):
    """Strip, collapse whitespace, return None if empty."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def norm_rut(raw):
    """Normalize Chilean RUT: remove dots, commas, spaces; extract digits+K+verifier.

    Returns (rut_normalizado, calidad_flag) or (None, flag).
    """
    if raw is None or str(raw).strip() == "":
        return None, "RUT_VACIO"

    s = str(raw).strip().upper()
    # Remove dots, commas, spaces (but keep K as verifier)
    cleaned = re.sub(r"[.\s,]+", "", s)
    # Remove leading/trailing junk but keep digit+verifier pattern
    match = re.search(r"(\d{6,8}-?[\dkK])", cleaned)
    if not match:
        # Try format with embedded hyphen
        match = re.search(r"(\d+[.]?\d*[.]?\d*-[\dkK])", s)
    if match:
        digits_body = match.group(1)
        # Ensure proper format: XXXXXXX-X
        norm = digits_body.replace(".", "").replace(",", "").replace(" ", "").upper()
        if "-" not in norm:
            if len(norm) > 1:
                norm = norm[:-1] + "-" + norm[-1]
        if not re.match(r"^\d{6,8}-[\dkK]$", norm):
            return norm, "RUT_FORMATO_IRREGULAR"
        return norm, None
    return cleaned, "RUT_NO_PARSEABLE"


def parse_date(val):
    """Parse a date from datetime, string (dd-mm-yyyy), or formula."""
    if val is None:
        return None, None

    if isinstance(val, datetime):
        return val.date(), None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val, None
    if isinstance(val, (int, float)):
        # Excel serial date numbers
        try:
            from openpyxl.utils import from_excel
            if 1 < val < 100000:
                return from_excel(val), None
        except Exception:
            pass
        return None, "FECHA_NUMERICA_NO_PARSEABLE"

    s = str(val).strip()
    if not s or s.startswith("="):
        return None, "FECHA_FORMULA"

    # Try dd-mm-yyyy
    for fmt in ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y"]:
        try:
            return datetime.strptime(s, fmt).date(), None
        except ValueError:
            continue

    # Garbage date detection (e.g. "15-01-19458")
    if re.match(r"\d{1,2}-\d{1,2}-\d{4,6}", s):
        return None, "FECHA_ANO_ANOMALO"

    return None, "FECHA_NO_PARSEABLE"


def parse_sexo(val):
    """Normalize sexo to M/F or None."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ("M", "MASCULINO", "HOMBRE", "VARON"):
        return "M"
    if s in ("F", "FEMENINO", "MUJER"):
        return "F"
    return "X"


def parse_bool(val):
    """Parse SI/NO boolean."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ("SI", "SÍ", "YES", "1", "TRUE", "VERDADERO"):
        return True
    if s in ("NO", "0", "FALSE", "FALSO"):
        return False
    return None


def flag_estado(val):
    """Detect typos in estado field."""
    if val is None:
        return "ESTADO_VACIO"
    s = str(val).strip().upper()
    if s == "EGRESADO":
        return None
    if s in ("EGRESADC", "EGRESADO ", "EGRESAO", "EGRESADOO"):
        return "ESTADO_TYPO"
    if s == "FALLECIDO" or s == "FALLECIDA":
        return None
    if s == "FALLECIDA":
        return None
    return "ESTADO_NO_RECONOCIDO"


def flag_categorizacion(val):
    """Detect typos in categorizacion."""
    if val is None:
        return "CATEGORIZACION_VACIA"
    s = str(val).strip().upper()
    if s in ("COMPLEJO", "INTERMEDIO"):
        return None
    if s in ("CMPLEJO", "COMPLEJA", "INTERMEDIOI", "COMPLEJO "):
        return "CATEGORIZACION_TYPO"
    return None


def parse_edad(val):
    """Parse edad as integer."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        v = int(val)
        return v if v > 0 and v < 120 else None
    s = str(val).strip()
    try:
        v = int(s)
        return v if v > 0 and v < 120 else None
    except ValueError:
        return None


def compute_row_hash(row_dict):
    """SHA256 of raw_record for dedup/traceability."""
    raw_json = json.dumps(row_dict, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(raw_json.encode("utf-8")).hexdigest()


def parse_sheet(ws, sheet_name):
    """Parse one sheet, returning list of row dicts."""
    rows = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1
    ):
        vals = list(row[: len(COLUMNS)])
        while len(vals) < len(COLUMNS):
            vals.append(None)

        if all(v is None for v in vals):
            continue

        raw = dict(zip(COLUMNS, vals))
        row_num = row_idx

        rut, rut_flag = norm_rut(raw.get("rut"))
        fecha_ing, fig_flag = parse_date(raw.get("fecha_ingreso"))
        fecha_egr, feg_flag = parse_date(raw.get("fecha_egreso"))
        fecha_nac, fn_flag = parse_date(raw.get("fecha_nacimiento"))

        dias_calc = None
        if fecha_ing and fecha_egr:
            dias_calc = (fecha_egr - fecha_ing).days

        calidad = []
        if rut_flag:
            calidad.append(rut_flag)
        if fig_flag:
            calidad.append("INGRESO_" + fig_flag)
        if feg_flag:
            calidad.append("EGRESO_" + feg_flag)
        if fn_flag:
            calidad.append("NACIMIENTO_" + fn_flag)

        est_flag = flag_estado(raw.get("estado"))
        if est_flag:
            calidad.append(est_flag)

        cat_flag = flag_categorizacion(raw.get("categorizacion"))
        if cat_flag:
            calidad.append(cat_flag)

        # Fecha improbable check
        if fecha_ing and fecha_egr and fecha_ing > fecha_egr:
            calidad.append("FECHA_INGRESO_POSTERIOR_EGRESO")
        if fecha_nac and fecha_ing and fecha_nac > fecha_ing:
            calidad.append("FECHA_NACIMIENTO_POSTERIOR_INGRESO")

        # Edad vs fecha_nac discrepancy (>2 years off)
        edad = parse_edad(raw.get("edad"))
        if edad is not None and fecha_nac and fecha_ing:
            expected_age = fecha_ing.year - fecha_nac.year
            if abs(edad - expected_age) > 2:
                calidad.append("EDAD_VS_NACIMIENTO_DISCREPANCIA")

        # Nacionalidad typos
        nac = norm_text(raw.get("nacionalidad"))
        if nac and nac.upper() not in ("CHILENO", "CHILENA", "EXTRANJERO", "EXTRANJERA"):
            calidad.append("NACIONALIDAD_TYPO")

        raw_record = {
            k: (str(v) if isinstance(v, (datetime, date)) else v)
            for k, v in raw.items()
        }

        rows.append({
            "ingreso_id": (
                "ing_" + hashlib.md5(
                    f"{DRIVE_ID}|{sheet_name}|{row_num}".encode()
                ).hexdigest()[:16]
            ),
            "drive_id": DRIVE_ID,
            "sheet_name": sheet_name,
            "source_row_number": row_num,
            "estado": norm_text(raw.get("estado")),
            "fecha_ingreso": fecha_ing,
            "fecha_egreso": fecha_egr,
            "dias_estada_declarado": norm_text(raw.get("dias_estada")),
            "dias_estada_calculado": dias_calc,
            "motivo_egreso": norm_text(raw.get("motivo_egreso")),
            "nombres": norm_text(raw.get("nombres")),
            "apellidos": norm_text(raw.get("apellidos")),
            "sexo": parse_sexo(raw.get("sexo")),
            "edad": edad,
            "fecha_nacimiento": fecha_nac,
            "fecha_nacimiento_raw": norm_text(raw.get("fecha_nacimiento")),
            "rut_raw": norm_text(raw.get("rut")),
            "rut_normalizado": rut,
            "prevision": norm_text(raw.get("prevision")),
            "servicio_origen": norm_text(raw.get("servicio_origen")),
            "usuario_o2": parse_bool(raw.get("usuario_o2")),
            "requerimiento_hodom_o2": parse_bool(raw.get("requerimiento_hodom_o2")),
            "categorizacion": norm_text(raw.get("categorizacion")),
            "diagnostico_egreso": norm_text(raw.get("diagnostico_egreso")),
            "domicilio": norm_text(raw.get("domicilio")),
            "comuna": norm_text(raw.get("comuna")),
            "cesfam": norm_text(raw.get("cesfam")),
            "nro_contacto": norm_text(raw.get("nro_contacto")),
            "nacionalidad": nac,
            "requiere_enfermeria": parse_bool(raw.get("enfermeria")),
            "requiere_kinesiologia": parse_bool(raw.get("kinesiologia")),
            "requiere_fonoaudiologia": parse_bool(raw.get("fonoaudiologia")),
            "raw_record": raw_record,
            "calidad_flags": calidad,
            "source_hash": compute_row_hash(raw_record),
        })
    return rows


def escape_sql(val):
    """Properly quote/escape a value for SQL."""
    if val is None:
        return "NULL"
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, date):
        return f"'{val.isoformat()}'"
    s = str(val).replace("'", "''")
    return f"'{s}'"


def generate_sql(rows, output_path):
    """Generate idempotent INSERT SQL from parsed rows."""
    col_names = [
        "ingreso_id", "drive_id", "sheet_name", "source_row_number",
        "estado", "fecha_ingreso", "fecha_egreso",
        "dias_estada_declarado", "dias_estada_calculado",
        "motivo_egreso", "nombres", "apellidos",
        "sexo", "edad", "fecha_nacimiento", "fecha_nacimiento_raw",
        "rut_raw", "rut_normalizado", "prevision", "servicio_origen",
        "usuario_o2", "requerimiento_hodom_o2", "categorizacion",
        "diagnostico_egreso", "domicilio", "comuna", "cesfam",
        "nro_contacto", "nacionalidad",
        "requiere_enfermeria", "requiere_kinesiologia", "requiere_fonoaudiologia",
        "raw_record", "calidad_flags", "source_hash",
    ]

    lines = [
        "-- HODOM INGRESOS 2026 DRIVE staging load",
        "-- Generated from: INGRESOS_2026_DRIVE.xlsx",
        f"-- Sheets: 5. Drive ID: {DRIVE_ID}",
        "-- Nominal data (PII) — not for versioned docs.",
        "",
        "BEGIN;",
        "",
    ]

    for r in rows:
        kval = escape_sql(json.dumps(r["calidad_flags"], ensure_ascii=False))
        rrec = escape_sql(json.dumps(r["raw_record"], ensure_ascii=False, default=str))
        vals_list = [
            escape_sql(r["ingreso_id"]),
            escape_sql(r["drive_id"]),
            escape_sql(r["sheet_name"]),
            str(r["source_row_number"]),
            escape_sql(r["estado"]),
            escape_sql(r["fecha_ingreso"]),
            escape_sql(r["fecha_egreso"]),
            escape_sql(r["dias_estada_declarado"]),
            escape_sql(r["dias_estada_calculado"]),
            escape_sql(r["motivo_egreso"]),
            escape_sql(r["nombres"]),
            escape_sql(r["apellidos"]),
            escape_sql(r["sexo"]),
            escape_sql(r["edad"]),
            escape_sql(r["fecha_nacimiento"]),
            escape_sql(r["fecha_nacimiento_raw"]),
            escape_sql(r["rut_raw"]),
            escape_sql(r["rut_normalizado"]),
            escape_sql(r["prevision"]),
            escape_sql(r["servicio_origen"]),
            escape_sql(r["usuario_o2"]),
            escape_sql(r["requerimiento_hodom_o2"]),
            escape_sql(r["categorizacion"]),
            escape_sql(r["diagnostico_egreso"]),
            escape_sql(r["domicilio"]),
            escape_sql(r["comuna"]),
            escape_sql(r["cesfam"]),
            escape_sql(r["nro_contacto"]),
            escape_sql(r["nacionalidad"]),
            escape_sql(r["requiere_enfermeria"]),
            escape_sql(r["requiere_kinesiologia"]),
            escape_sql(r["requiere_fonoaudiologia"]),
            rrec,
            kval,
            escape_sql(r["source_hash"]),
        ]

        columns = ", ".join(col_names)
        values = ", ".join(vals_list)
        lines.append(
            f"INSERT INTO staging.hodom_ingreso_2026 ({columns})"
            f"\nVALUES ({values})"
            f"\nON CONFLICT (drive_id, sheet_name, source_row_number) DO UPDATE SET"
            f"\n  estado = EXCLUDED.estado,"
            f"\n  fecha_ingreso = EXCLUDED.fecha_ingreso,"
            f"\n  fecha_egreso = EXCLUDED.fecha_egreso,"
            f"\n  dias_estada_declarado = EXCLUDED.dias_estada_declarado,"
            f"\n  dias_estada_calculado = EXCLUDED.dias_estada_calculado,"
            f"\n  motivo_egreso = EXCLUDED.motivo_egreso,"
            f"\n  nombres = EXCLUDED.nombres,"
            f"\n  apellidos = EXCLUDED.apellidos,"
            f"\n  sexo = EXCLUDED.sexo,"
            f"\n  edad = EXCLUDED.edad,"
            f"\n  fecha_nacimiento = EXCLUDED.fecha_nacimiento,"
            f"\n  fecha_nacimiento_raw = EXCLUDED.fecha_nacimiento_raw,"
            f"\n  rut_raw = EXCLUDED.rut_raw,"
            f"\n  rut_normalizado = EXCLUDED.rut_normalizado,"
            f"\n  prevision = EXCLUDED.prevision,"
            f"\n  servicio_origen = EXCLUDED.servicio_origen,"
            f"\n  usuario_o2 = EXCLUDED.usuario_o2,"
            f"\n  requerimiento_hodom_o2 = EXCLUDED.requerimiento_hodom_o2,"
            f"\n  categorizacion = EXCLUDED.categorizacion,"
            f"\n  diagnostico_egreso = EXCLUDED.diagnostico_egreso,"
            f"\n  domicilio = EXCLUDED.domicilio,"
            f"\n  comuna = EXCLUDED.comuna,"
            f"\n  cesfam = EXCLUDED.cesfam,"
            f"\n  nro_contacto = EXCLUDED.nro_contacto,"
            f"\n  nacionalidad = EXCLUDED.nacionalidad,"
            f"\n  requiere_enfermeria = EXCLUDED.requiere_enfermeria,"
            f"\n  requiere_kinesiologia = EXCLUDED.requiere_kinesiologia,"
            f"\n  requiere_fonoaudiologia = EXCLUDED.requiere_fonoaudiologia,"
            f"\n  raw_record = EXCLUDED.raw_record,"
            f"\n  calidad_flags = EXCLUDED.calidad_flags,"
            f"\n  source_hash = EXCLUDED.source_hash,"
            f"\n  updated_at = now();\n"
        )

    lines.append("COMMIT;\n")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")


def main():
    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    print(f"Sheets: {wb.sheetnames}")

    all_rows = []
    stats = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = parse_sheet(ws, name)
        all_rows.extend(rows)
        stats[name] = {
            "rows": len(rows),
            "with_rut": sum(1 for r in rows if r["rut_normalizado"]),
            "with_fecha_ing": sum(1 for r in rows if r["fecha_ingreso"]),
            "with_fecha_egr": sum(1 for r in rows if r["fecha_egreso"]),
            "flags": sum(1 for r in rows if r["calidad_flags"]),
        }
        print(f"  {name}: {len(rows)} rows, "
              f"{stats[name]['with_rut']} with RUT, "
              f"{stats[name]['flags']} with quality flags")

    print(f"\nTotal: {len(all_rows)} rows")

    # Quality summary
    all_flags = []
    for r in all_rows:
        all_flags.extend(r["calidad_flags"])
    from collections import Counter
    flag_counts = Counter(all_flags)
    if flag_counts:
        print("\nQuality flags:")
        for flag, count in flag_counts.most_common():
            print(f"  {flag}: {count}")

    # Dedup check
    id_counts = Counter(r["ingreso_id"] for r in all_rows)
    dup_ids = {k: v for k, v in id_counts.items() if v > 1}
    if dup_ids:
        print(f"\nWARNING: {len(dup_ids)} duplicate ingreso_ids detected!")

    # Generate SQL
    generate_sql(all_rows, OUT_SQL)
    print(f"\nSQL written to: {OUT_SQL}")
    print(f"Size: {OUT_SQL.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
