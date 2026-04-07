#!/usr/bin/env python3
"""
CORR-15: Entrega de turno kinesiología → clinical.nota_evolucion

Fuente: documentacion-legacy/drive-hodom/ENTREGA KINE/Ent. Turno Hodom KINE.xlsx
  112 hojas diarias (dic 2025 → abr 2026), ~987 registros paciente-día.

Cada fila: paciente, RUT, cobertura (CONTROL/INGRESO/ALTA), diagnóstico,
observaciones de tratamiento, hora atención, registro FC (KTM/KTR/ED).

Target: clinical.nota_evolucion (tipo='kinesiologia')
  notas_clinicas = "Dx: {diagnóstico}. {observaciones}"
  plan_enfermeria = cobertura + registro FC
"""

from __future__ import annotations

import hashlib
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl
import psycopg


def _normalize_rut(rut: str) -> str:
    """Remove dots, keep hyphen, uppercase K."""
    return rut.replace('.', '').replace(' ', '').upper().strip()


def _parse_date_from_sheet(ws) -> date | None:
    """Extract date from sheet header (row 1, col F or G)."""
    for col in [6, 7]:
        val = ws.cell(1, col).value
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
    return None


def _parse_sheet(ws, sheet_name: str) -> list[dict]:
    """Parse one daily sheet into patient records."""
    fecha = _parse_date_from_sheet(ws)
    if not fecha:
        # Try to infer from sheet name (DD-MM format, year from context)
        m = re.match(r'(\d{1,2})\s*-\s*(\d{1,2})', sheet_name.strip())
        if m:
            day, month = int(m.group(1)), int(m.group(2))
            year = 2026 if month <= 4 else 2025
            try:
                fecha = date(year, month, day)
            except ValueError:
                pass

    if not fecha:
        return []

    records = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        vals = [str(v).strip() if v else '' for v in row]
        if len(vals) < 5:
            continue

        nombre = vals[0]
        rut_raw = vals[1]

        # Skip non-patient rows
        if not nombre or not re.search(r'\d+[\.\-]', rut_raw):
            continue
        if 'KLGO' in nombre.upper() or 'TIMBRE' in nombre.upper() or 'NOMBRE' in nombre.upper():
            continue

        rut = _normalize_rut(rut_raw)
        cobertura = vals[2] if len(vals) > 2 else ''
        diagnostico = vals[3] if len(vals) > 3 else ''
        observaciones = vals[4] if len(vals) > 4 else ''
        hora = vals[5] if len(vals) > 5 else ''
        registro_fc = vals[6] if len(vals) > 6 else ''

        # Build clinical note
        parts = []
        if diagnostico:
            parts.append(f"Dx: {diagnostico}")
        if observaciones:
            parts.append(observaciones)
        notas = '. '.join(parts) if parts else None

        # Build plan
        plan_parts = []
        if cobertura:
            plan_parts.append(f"Cobertura: {cobertura}")
        if registro_fc:
            plan_parts.append(f"Prestación: {registro_fc}")
        plan = ' | '.join(plan_parts) if plan_parts else None

        # Clean hora (handle "17:40/15:00" format → take first)
        if hora:
            hora = hora.split('/')[0].strip()
            if isinstance(hora, str) and not re.match(r'\d{1,2}:\d{2}', hora):
                hora = ''

        records.append({
            'nombre': nombre,
            'rut': rut,
            'fecha': fecha,
            'hora': hora or None,
            'notas_clinicas': notas,
            'plan': plan,
            'cobertura': cobertura,
        })

    return records


def load_kine(xlsx_path: str, db_url: str, dry_run: bool = False):
    wb = openpyxl.load_workbook(xlsx_path)
    conn = psycopg.connect(db_url, autocommit=False)

    # Build RUT → patient_id
    rut_index = {}
    for pid, rut in conn.execute(
        "SELECT patient_id, rut FROM clinical.paciente WHERE rut IS NOT NULL"
    ).fetchall():
        rut_index[rut.replace('.', '').replace(' ', '').upper().strip()] = pid

    # patient_id → latest stay
    stay_lookup = {}
    for sid, pid, fi, fe in conn.execute(
        "SELECT stay_id, patient_id, fecha_ingreso, fecha_egreso "
        "FROM clinical.estadia ORDER BY fecha_ingreso DESC"
    ).fetchall():
        if pid not in stay_lookup:
            stay_lookup[pid] = sid

    # Parse all sheets
    all_records = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        records = _parse_sheet(ws, sname)
        all_records.extend(records)

    # Disable triggers for bulk load
    if not dry_run:
        conn.execute("ALTER TABLE clinical.nota_evolucion DISABLE TRIGGER trg_nota_evolucion_pe1")
        conn.execute("ALTER TABLE clinical.nota_evolucion DISABLE TRIGGER trg_nota_evolucion_stay_coherence")
        # Delete previous CORR-15 entries for idempotency
        conn.execute(
            "DELETE FROM clinical.nota_evolucion WHERE nota_id LIKE 'nk_%'"
        )
        conn.execute(
            "DELETE FROM migration.provenance WHERE phase = 'CORR-15'"
        )

    inserted = 0
    unmatched = []
    seen = set()

    for rec in all_records:
        patient_id = rut_index.get(rec['rut'])
        if not patient_id:
            unmatched.append(f"{rec['nombre']} ({rec['rut']})")
            continue

        stay_id = stay_lookup.get(patient_id)
        if not stay_id:
            continue

        # Dedup key
        key = f"{patient_id}|{rec['fecha']}|{rec['hora'] or ''}"
        if key in seen:
            continue
        seen.add(key)

        raw = f"nk|{patient_id}|{rec['fecha']}|{rec['hora'] or ''}"
        nota_id = "nk_" + hashlib.sha256(raw.encode()).hexdigest()[:12]

        if not dry_run:
            conn.execute(
                """
                INSERT INTO clinical.nota_evolucion
                    (nota_id, stay_id, patient_id, tipo, fecha, hora, notas_clinicas, plan_enfermeria)
                VALUES (%s, %s, %s, 'kinesiologia', %s, %s, %s, %s)
                ON CONFLICT (nota_id) DO NOTHING
                """,
                (nota_id, stay_id, patient_id, rec['fecha'], rec['hora'],
                 rec['notas_clinicas'], rec['plan']),
            )
            conn.execute(
                """
                INSERT INTO migration.provenance
                    (target_table, target_pk, source_type, source_file, source_key, phase)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT DO NOTHING
                """,
                ("clinical.nota_evolucion", nota_id, "legacy",
                 "ENTREGA KINE/Ent. Turno Hodom KINE.xlsx",
                 rec['rut'], "CORR-15"),
            )
        inserted += 1

    if not dry_run:
        conn.execute("ALTER TABLE clinical.nota_evolucion ENABLE TRIGGER trg_nota_evolucion_pe1")
        conn.execute("ALTER TABLE clinical.nota_evolucion ENABLE TRIGGER trg_nota_evolucion_stay_coherence")
        conn.commit()

    conn.close()

    unique_unmatched = sorted(set(unmatched))
    return {
        'parsed': len(all_records),
        'inserted': inserted,
        'deduped': len(all_records) - inserted - len(unmatched),
        'unmatched': unique_unmatched,
    }


def main():
    import argparse
    parser = argparse.ArgumentParser(description="CORR-15: Entrega kine → PG")
    parser.add_argument("--db-url", default="postgresql://hodom:hodom@localhost:5555/hodom")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--xlsx",
                        default="documentacion-legacy/drive-hodom/ENTREGA KINE/Ent. Turno Hodom KINE.xlsx")
    args = parser.parse_args()

    result = load_kine(args.xlsx, args.db_url, args.dry_run)

    print(f"Registros parseados: {result['parsed']}")
    print(f"Insertados:          {result['inserted']}")
    print(f"Deduplicados:        {result['deduped']}")
    if result['unmatched']:
        print(f"\nSin match ({len(result['unmatched'])}):")
        for u in result['unmatched'][:15]:
            print(f"  - {u}")
        if len(result['unmatched']) > 15:
            print(f"  ... y {len(result['unmatched']) - 15} más")

    if not args.dry_run:
        conn = psycopg.connect(args.db_url)
        stats = conn.execute("""
            SELECT tipo, count(*) FROM clinical.nota_evolucion GROUP BY tipo ORDER BY count(*) DESC
        """).fetchall()
        conn.close()
        print(f"\n--- Notas por tipo ---")
        for tipo, n in stats:
            print(f"  {tipo}: {n}")


if __name__ == "__main__":
    main()
