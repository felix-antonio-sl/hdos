import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from scripts.hodom_annual_metrics import (
    OUT_DIR,
    aggregate_rem_years,
    build_unique_user_rows,
    extract_hodom_daily_admissions,
    extract_nominal_admissions,
    extract_rem_a21_metrics,
)


class HodomAnnualMetricsTests(unittest.TestCase):
    def test_outputs_are_written_under_normative_specs_directory(self):
        self.assertEqual(OUT_DIR, Path("/home/felix/projects/hdos/docs/specs/metricas-hodom"))

    def test_extract_rem_a21_metrics_handles_2025_profession_labels(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rem_2025.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Hoja1"
            ws["A1"] = "SECCIÓN C: HOSPITALIZACIÓN DOMICILIARIA"
            ws["A6"] = "Ingresos"
            ws["C6"] = 56
            ws["A7"] = "Personas Atendidas"
            ws["C7"] = 61
            ws["A8"] = "Días Personas Atendidas"
            ws["C8"] = 404
            ws["A15"] = "Médico"
            ws["B15"] = 176
            ws["A16"] = "Enfermera"
            ws["B16"] = 250
            ws["A17"] = "Técnico en Enfermería"
            ws["B17"] = 44
            ws["A19"] = "Kinesiólogo"
            ws["B19"] = 241
            ws["A21"] = "Fonoaudiólogo/a"
            ws["B21"] = 104
            ws["A22"] = "Trabajador/a Social"
            ws["B22"] = 5
            wb.save(path)

            metrics = extract_rem_a21_metrics(path)

        self.assertEqual(metrics["personas"]["ingresos"], 56)
        self.assertEqual(metrics["personas"]["personas_atendidas"], 61)
        self.assertEqual(metrics["visitas"]["tecnico_paramedico"], 44)
        self.assertEqual(metrics["visitas"]["fonoaudiologo"], 104)
        self.assertEqual(metrics["visitas"]["trabajador_social"], 5)
        self.assertEqual(metrics["total_visitas"], 820)

    def test_aggregate_rem_years_sums_annual_visits_and_users(self):
        rows = [
            {
                "period": "2025-01",
                "metrics": {
                    "personas": {"ingresos": 10, "personas_atendidas": 12, "dias_persona": 80},
                    "visitas": {"medico": 5, "enfermera": 20},
                    "total_visitas": 25,
                },
            },
            {
                "period": "2025-02",
                "metrics": {
                    "personas": {"ingresos": 11, "personas_atendidas": 13, "dias_persona": 90},
                    "visitas": {"medico": 6, "enfermera": 21},
                    "total_visitas": 27,
                },
            },
        ]

        annual = aggregate_rem_years(rows)

        self.assertEqual(annual[2025]["atenciones_rem_visitas"], 52)
        self.assertEqual(annual[2025]["usuarios_rem_personas_atendidas_suma_mensual"], 25)
        self.assertEqual(annual[2025]["ingresos_rem"], 21)
        self.assertEqual(annual[2025]["visitas_by_profession"]["enfermera"], 41)

    def test_extract_nominal_admissions_counts_rows_and_unique_rut(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "ingresos.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "INGRESOS"
            ws.append(["ESTADO", "FECHA DE INGRESO", "RUT"])
            ws.append(["EGRESADO", datetime(2024, 11, 1), "12.345.678-9"])
            ws.append(["EGRESADO", datetime(2024, 11, 3), "12345678-9"])
            ws.append(["EGRESADO", datetime(2024, 12, 1), "9.876.543-K"])
            ws.append(["EGRESADO", datetime(2025, 1, 1), "1-9"])
            wb.save(path)

            metrics = extract_nominal_admissions(path, "INGRESOS", 2024)

        self.assertEqual(metrics["admissions"], 3)
        self.assertEqual(metrics["unique_users"], 2)
        self.assertEqual(metrics["missing_user_id"], 0)

    def test_extract_hodom_daily_admissions_uses_total_ingresos_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "hodom.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "ENERO2025"
            ws.append(["Día del Mes", None, "INGRESOS", None, None, None, None, None, None, None])
            ws.append([None, "Existenc. Día anterior", "Urgencia", "Hospitalizacion", "Ambulatorio", "Ley de Urgencia", "UGCC", "Total", "Por traslados", "Total Ingresos"])
            ws.append([None])
            ws.append([None])
            ws.append([1, 17, None, None, None, None, None, 0, None, 2])
            ws.append([2, 19, 1, None, None, None, None, 1, None, 1])
            wb.save(path)

            metrics = extract_hodom_daily_admissions(path, 2025)

        self.assertEqual(metrics["monthly"]["ENERO2025"], 3)
        self.assertEqual(metrics["annual_total"], 3)

    def test_build_unique_user_rows_includes_partial_2025_nominal_source(self):
        rows = build_unique_user_rows(
            nominal_2024={"admissions": 858, "unique_users": 707, "missing_user_id": 1},
            nominal_2025={
                "admissions": 500,
                "unique_users": 445,
                "missing_user_id": 2,
                "monthly_admissions": {"2025-01": 52, "2025-10": 15},
            },
            database={"available": True, "rows": {"2025": {"usuarios_ingresados": 559, "estadia_ingresos": 634}}},
        )

        self.assertEqual(rows[1]["year"], 2024)
        self.assertEqual(rows[1]["unique_users"], 707)
        self.assertEqual(rows[2]["year"], 2025)
        self.assertEqual(rows[2]["unique_users"], 559)
        self.assertIn("nominal parcial", rows[2]["note"])
        self.assertIn("445", rows[2]["note"])


if __name__ == "__main__":
    unittest.main()
