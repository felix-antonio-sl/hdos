#!/usr/bin/env python3
"""Build annual HODOM attention/user metrics from Drive sources and local DB."""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / ".tmp" / "drive-refresh" / "raw"
OUT_DIR = ROOT / "docs" / "specs" / "metricas-hodom"
REPORT_STEM = "metricas-anuales-atenciones-usuarios-2026-05-25"
JSON_OUT = OUT_DIR / f"{REPORT_STEM}.json"
MD_OUT = OUT_DIR / f"{REPORT_STEM}.md"
CSV_OUT = OUT_DIR / f"{REPORT_STEM}.csv"

DRIVE_ROOT_URL = "https://drive.google.com/drive/folders/1P6ur0dHwmECADUluzwfVvA5g9uTO9oVe?usp=sharing"
DB_URL = "postgresql://hodom:hodom@localhost:5555/hodom"

MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "SETIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

REM_SOURCES: list[dict[str, Any]] = [
    {"period": "2023-07", "id": "1YqqaWrx5ktmOozPsPgow-Vj_n277pT8r", "path": "2023/REM/REM JULIO.xlsx"},
    {"period": "2023-08", "id": "16EhYHzj79ejkkcbv-U-2AkpqWWyyLozL", "path": "2023/REM/REM AGOSTO .xlsx"},
    {"period": "2023-09", "id": "1KLqIZst4Io8XPpCbjbMZoQ3CQQjWD_r2", "path": "2023/REM/REM SEPTIEMBRE.xlsx"},
    {"period": "2023-10", "id": "1zJkabCxZf-tmcVh3QjCg8a8pVm3Enrbt", "path": "2023/REM/REM OCTUBRE.xlsx"},
    {"period": "2023-11", "id": "1wx-I3In7JF4zsjst_e3lLi7jQ-ttQxno", "path": "2023/REM/REM NOVIEMBRE.xlsx"},
    {"period": "2024-01", "id": "1mzhxOD6oZNKOVZhmtyOAUaG9oCf1EA_3", "path": "2024/REM/REM ENERO.xlsx"},
    {"period": "2024-02", "id": "1nRSfmUDKCDAL9i1itUpuLaZTGszTtJWD", "path": "2024/REM/REM FEBRERO.xlsx"},
    {"period": "2024-03", "id": "1bX7VftfBpK0sKFvL1_-TYj5zKsCtX-l6", "path": "2024/REM/REM MARZO.xlsx"},
    {"period": "2024-04", "id": "1TJprQcKtmepC7T3dinRhNhuLb8sZSSVk", "path": "2024/REM/REM ABRIL 24.xlsx"},
    {"period": "2024-05", "id": "1o89epL0PR0uARpLos8La4djWu6XemGCE", "path": "2024/REM/REM MAYO.xlsx"},
    {"period": "2024-06", "id": "1O0vZ-rS6zmXYc3hiO2uJctAuSAGn6qHy", "path": "2024/REM/REM JUNIO 24.xlsx"},
    {"period": "2024-07", "id": "1Y4Z2qhJUH5ac_2TBwyMK3gJEBQQqtU8h", "path": "2024/REM/REM JULIO .xlsx"},
    {"period": "2024-08", "id": "1gJmNAnkjixRWd9iUS2hMzrf1PE7saon5", "path": "2024/REM/REM AGOSTO .xlsx"},
    {"period": "2024-09", "id": "1lcSUJDJZY7smocpthV4U4C_gDoRlY7Jr", "path": "2024/REM/REM SEPTIEMBRE.xlsx"},
    {"period": "2024-10", "id": "1tUk3kjN3wMZUrcrfuHOxSS2KeXmTsc8E", "path": "2024/REM/REM OCTUBRE 24.xlsx"},
    {"period": "2024-11", "id": "1QvpN7L-sDHzv832n6wIeG9BF-DZ-Xw11", "path": "2024/REM/REM NOVIEMBRE 2024.xlsx"},
    {"period": "2024-12", "id": "16a0d9oKZY1CLbNMrI31hRsNLtQB3rspS", "path": "2024/REM/REM DICIEMBRE.xlsx"},
    {"period": "2025-01", "id": "1ZDjH03MGi1-6IXBzvnoMUh4hNwZTNkU2", "path": "2025/REM ENERO 2025.xlsx"},
    {"period": "2025-02", "id": "17dnhkUd_livEcTq20R5d4_vVoSK7YdLz", "path": "2025/REM FEBRERO 2025.xlsx"},
    {"period": "2025-03", "id": "1o4Gzm6gIRIUB0NPk9qpkk2npgnGC-BOm", "path": "2025/REM MARZO 25.xlsx"},
    {"period": "2025-04", "id": "1GjYTueXURnPif_ROkNqSbwJwyZwfVqLi", "path": "2025/REM ABRIL 25.xlsx"},
    {"period": "2025-05", "id": "10LPwBU_IBAzbLy3WqVkiInjdcgoaomQN", "path": "2025/REM MAYO 25.xlsx"},
    {"period": "2025-06", "id": "1G8Irz-5jRvYnITkwqi6bRRfItYH3mk8m", "path": "2025/REM JUNIO 25.xlsx"},
    {"period": "2025-07", "id": "1C4aQqgQLwEUDQeIOWETki5oe6Kg9gSmD", "path": "2025/REM JULIO 25.xlsx"},
    {"period": "2025-08", "id": "1Sn8gTpfkgYlt9cHaqIaOfZQO65fPs77N", "path": "2025/REM AGOSTO 25.xlsx"},
    {"period": "2025-09", "id": "1-AALBIe9XfeIv74dKJfILmcwgsPHR9vY", "path": "2025/REM SEPTIEMBRE 2025.xlsx"},
    {"period": "2025-10", "id": "1egEOppOHucpPKbjJ-JwmEm4nzuZN0r62", "path": "2025/REM OCTUBRE 2025.xlsx"},
    {"period": "2025-11", "id": "1Cllu661Xrnv6HOu3-XC1o0nYELk-GjKI", "path": "2025/REM NOVIEMBRE 2025.xlsx"},
    {"period": "2025-12", "id": "1o33DqlEBKjxuHvaUvtkuyPASKG_k_6Lj", "path": "2025/REM DICIEMBRE 2025.xlsx"},
]

NOMINAL_2024_SOURCES = [
    {
        "id": "1g-ojvSZh2uuYHwKz7ch7So9p0V2yLbtp",
        "path": "2024/INGRESOS ENE-OCT 2024.xlsx",
        "sheet": "INGRESOS 2024",
    },
    {
        "id": "15s5FU9UKBjgTQxz277_M7zSq9KGavLXN",
        "path": "2024/PLANILLA NOV DIC 2024.xlsx",
        "sheet": "INGRESOS",
    },
]

HODOM_2025_SOURCE = {
    "id": "1Sr27xmW0V7eqBZd4tjAJ2AgcID0VSNze",
    "path": "HOMOM 2025.xlsx",
}

NOMINAL_2025_SOURCE = {
    "id": "1EmbYGViWjZKe_uQoEi9OiGhgXiMdzFtO",
    "path": "2025/INGRESOS 2025 (3).xlsx",
    "sheet": "INGRESOS",
}

DOC_2023_SOURCE = {
    "id": "1qxXDDohR2_tIFn4bKRfvevkYyChSVsdP",
    "path": "2023/PACIENTES 2023.docx",
}

PROFESSION_ALIASES = {
    "MEDICO": "medico",
    "ENFERMERA": "enfermera",
    "TECNICO PARAMEDICO": "tecnico_paramedico",
    "TECNICO EN ENFERMERIA": "tecnico_paramedico",
    "KINESIOLOGO": "kinesiologo",
    "FONOAUDIOLOGO": "fonoaudiologo",
    "FONOAUDIOLOGO/A": "fonoaudiologo",
    "TRABAJADOR SOCIAL": "trabajador_social",
    "TRABAJADOR/A SOCIAL": "trabajador_social",
    "TERAPEUTA OCUPACIONAL": "terapeuta_ocupacional",
    "MATRONA": "matrona",
    "PSICOLOGO": "psicologo",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.upper().split())


def numeric(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def safe_name(path: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", path)


def local_path(source: dict[str, Any]) -> Path:
    name = safe_name(source["path"])
    if not name.lower().endswith((".xlsx", ".docx")):
        name += ".xlsx"
    return RAW_DIR / name


def download_source(source: dict[str, Any], force: bool = False) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    target = local_path(source)
    if target.exists() and target.stat().st_size > 0 and not force:
        return target
    subprocess.run(
        ["curl", "-L", "-s", "-o", str(target), f"https://drive.google.com/uc?export=download&id={source['id']}"],
        check=True,
    )
    return target


def choose_rem_sheet(workbook: Any) -> Any:
    for sheet in workbook.worksheets:
        values = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 60), values_only=True):
            values.extend(row[:8])
        if "HOSPITALIZACION DOMICILIARIA" in normalize_text(" ".join(str(value) for value in values if value is not None)):
            return sheet
    return workbook.worksheets[0]


def extract_rem_a21_metrics(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = choose_rem_sheet(workbook)
    metrics: dict[str, Any] = {"personas": {}, "visitas": {}, "total_visitas": 0}
    person_labels = {
        "INGRESOS": "ingresos",
        "PERSONAS ATENDIDAS": "personas_atendidas",
        "DIAS PERSONAS ATENDIDAS": "dias_persona",
        "ALTAS": "altas",
        "REINGRESOS A HOSPITALIZACION TRADICIONAL": "reingresos_hospitalizacion",
    }

    for row in sheet.iter_rows(values_only=True):
        values = list(row)
        label = normalize_text(values[0] if values else None)
        for prefix, key in person_labels.items():
            if label.startswith(prefix):
                metrics["personas"][key] = numeric(values[2] if len(values) > 2 else None)
        if label.startswith("FALLECIDOS"):
            metrics["personas"]["fallecidos_esperados"] = numeric(values[2] if len(values) > 2 else None)
        if label in PROFESSION_ALIASES:
            metrics["visitas"][PROFESSION_ALIASES[label]] = numeric(values[1] if len(values) > 1 else None)

    metrics["total_visitas"] = sum(value for value in metrics["visitas"].values() if isinstance(value, (int, float)))
    return metrics


def aggregate_rem_years(rows: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
    annual: dict[int, dict[str, Any]] = {}
    for row in rows:
        period = row["period"]
        year = int(period[:4])
        metrics = row["metrics"]
        bucket = annual.setdefault(
            year,
            {
                "periods": [],
                "atenciones_rem_visitas": 0,
                "usuarios_rem_personas_atendidas_suma_mensual": 0,
                "ingresos_rem": 0,
                "dias_persona": 0,
                "visitas_by_profession": defaultdict(int),
            },
        )
        bucket["periods"].append(period)
        bucket["atenciones_rem_visitas"] += metrics.get("total_visitas") or 0
        bucket["usuarios_rem_personas_atendidas_suma_mensual"] += metrics.get("personas", {}).get("personas_atendidas") or 0
        bucket["ingresos_rem"] += metrics.get("personas", {}).get("ingresos") or 0
        bucket["dias_persona"] += metrics.get("personas", {}).get("dias_persona") or 0
        for profession, value in metrics.get("visitas", {}).items():
            bucket["visitas_by_profession"][profession] += value or 0

    for bucket in annual.values():
        bucket["periods"] = sorted(bucket["periods"])
        bucket["visitas_by_profession"] = dict(sorted(bucket["visitas_by_profession"].items()))
    return annual


def normalize_rut(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).upper().replace(".", "").replace(" ", "").strip()
    return text or None


def find_header_index(headers: list[Any], candidates: list[str]) -> int | None:
    normalized = [normalize_text(header) for header in headers]
    for candidate in candidates:
        wanted = normalize_text(candidate)
        for index, header in enumerate(normalized):
            if wanted in header:
                return index
    return None


def extract_nominal_admissions(path: Path, sheet_name: str, year: int) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    fecha_index = find_header_index(headers, ["FECHA DE INGRESO"])
    rut_index = find_header_index(headers, ["RUT"])
    if fecha_index is None:
        raise ValueError(f"Fecha de ingreso column not found in {path.name}:{sheet_name}")

    admissions = 0
    missing_user_id = 0
    users: set[str] = set()
    monthly = Counter()
    for row in rows:
        values = list(row)
        fecha = values[fecha_index] if len(values) > fecha_index else None
        if not isinstance(fecha, datetime) or fecha.year != year:
            continue
        admissions += 1
        monthly[f"{year}-{fecha.month:02d}"] += 1
        rut = normalize_rut(values[rut_index] if rut_index is not None and len(values) > rut_index else None)
        if rut:
            users.add(rut)
        else:
            missing_user_id += 1

    return {
        "source_file": str(path),
        "sheet": sheet_name,
        "year": year,
        "admissions": admissions,
        "unique_users": len(users),
        "missing_user_id": missing_user_id,
        "monthly_admissions": dict(sorted(monthly.items())),
        "_user_ids": users,
    }


def merge_nominal_admissions(items: list[dict[str, Any]], year: int) -> dict[str, Any]:
    users: set[str] = set()
    monthly = Counter()
    admissions = 0
    missing = 0
    for item in items:
        admissions += item["admissions"]
        missing += item["missing_user_id"]
        users.update(item.get("_user_ids", set()))
        monthly.update(item["monthly_admissions"])
    return {
        "year": year,
        "admissions": admissions,
        "unique_users": len(users),
        "missing_user_id": missing,
        "monthly_admissions": dict(sorted(monthly.items())),
        "source_count": len(items),
    }


def extract_hodom_daily_admissions(path: Path, year: int) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    monthly: dict[str, int | float] = {}
    for sheet in workbook.worksheets:
        if str(year) not in normalize_text(sheet.title):
            continue
        header_row = None
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if normalize_text(row[0] if row else None).startswith("DIA DEL MES"):
                header_row = index
                break
        if header_row is None:
            continue
        labels = list(sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True))[0]
        total_index = None
        for index, value in enumerate(labels):
            if normalize_text(value) == "TOTAL INGRESOS":
                total_index = index
                break
        if total_index is None:
            total_index = 9
        total = 0
        for row in sheet.iter_rows(min_row=header_row + 4, values_only=True):
            values = list(row)
            day = numeric(values[0] if values else None)
            if not isinstance(day, (int, float)) or day < 1 or day > 31:
                continue
            total += numeric(values[total_index] if len(values) > total_index else None) or 0
        monthly[sheet.title] = total
    return {"annual_total": sum(monthly.values()), "monthly": monthly}


def extract_2023_certificate(path: Path) -> dict[str, Any]:
    text = ""
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("word/document.xml"))
        text = "\n".join(node.text or "" for node in root.iter() if node.tag.endswith("}t"))
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    monthly: dict[str, int] = {}
    for index, line in enumerate(lines[:-1]):
        month = MONTHS.get(normalize_text(line))
        value = numeric(lines[index + 1]) if month else None
        if month and isinstance(value, int):
            monthly[f"2023-{month:02d}"] = value
    total = None
    for index, line in enumerate(lines[:-1]):
        if normalize_text(line) == "TOTAL":
            total = numeric(lines[index + 1])
            break
    return {"periods": dict(sorted(monthly.items())), "total_ingresos_certificados": total}


def query_database_metrics() -> dict[str, Any]:
    sql = """
WITH years(anio) AS (VALUES (2023), (2024), (2025))
SELECT
    y.anio,
    COALESCE(v.visitas_rows, 0) AS visitas_rows,
    COALESCE(v.usuarios_con_visita, 0) AS usuarios_con_visita,
    COALESCE(v.visitas_realizadas, 0) AS visitas_realizadas,
    COALESCE(v.visitas_rem_reportable, 0) AS visitas_rem_reportable,
    COALESCE(e.estadia_ingresos, 0) AS estadia_ingresos,
    COALESCE(e.usuarios_ingresados, 0) AS usuarios_ingresados
FROM years y
LEFT JOIN LATERAL (
    SELECT
        count(*) AS visitas_rows,
        count(DISTINCT patient_id) AS usuarios_con_visita,
        count(*) FILTER (WHERE estado IN ('COMPLETA','DOCUMENTADA','VERIFICADA','REPORTADA_REM')) AS visitas_realizadas,
        count(*) FILTER (WHERE rem_reportable) AS visitas_rem_reportable
    FROM operational.visita
    WHERE fecha >= make_date(y.anio, 1, 1)
      AND fecha < make_date(y.anio + 1, 1, 1)
) v ON TRUE
LEFT JOIN LATERAL (
    SELECT
        count(*) AS estadia_ingresos,
        count(DISTINCT patient_id) AS usuarios_ingresados
    FROM clinical.estadia
    WHERE fecha_ingreso >= make_date(y.anio, 1, 1)
      AND fecha_ingreso < make_date(y.anio + 1, 1, 1)
      AND estado <> 'pendiente_evaluacion'
) e ON TRUE
ORDER BY y.anio;
"""
    env = {**os.environ, "PGPASSWORD": "hodom"}
    try:
        completed = subprocess.run(
            ["psql", DB_URL, "-At", "-F", "\t", "-v", "ON_ERROR_STOP=1", "-c", sql],
            check=True,
            capture_output=True,
            text=True,
            env=env,
        )
    except (subprocess.CalledProcessError, FileNotFoundError) as exc:
        return {"available": False, "error": str(exc)}

    rows: dict[int, dict[str, int]] = {}
    for line in completed.stdout.splitlines():
        anio, visitas, usuarios_visita, realizadas, rem_reportable, ingresos, usuarios = line.split("\t")
        rows[int(anio)] = {
            "visitas_rows": int(visitas),
            "usuarios_con_visita": int(usuarios_visita),
            "visitas_realizadas": int(realizadas),
            "visitas_rem_reportable": int(rem_reportable),
            "estadia_ingresos": int(ingresos),
            "usuarios_ingresados": int(usuarios),
        }
    return {"available": True, "rows": rows}


def build_quality_flags(
    annual_rem: dict[int, dict[str, Any]],
    monthly_rem: list[dict[str, Any]],
    db_metrics: dict[str, Any],
    nominal_2024: dict[str, Any],
    nominal_2025: dict[str, Any],
    hodom_2025: dict[str, Any],
) -> list[dict[str, Any]]:
    flags: list[dict[str, Any]] = []
    for year, bucket in sorted(annual_rem.items()):
        expected = {f"{year}-{month:02d}" for month in range(1, 13)}
        present = set(bucket["periods"])
        if expected - present:
            flags.append(
                {
                    "code": "rem_coverage_gap",
                    "severity": "review",
                    "year": year,
                    "missing_periods": sorted(expected - present),
                }
            )
    for row in monthly_rem:
        period = row["period"]
        personas = row["metrics"]["personas"]
        ingresos = personas.get("ingresos")
        atendidas = personas.get("personas_atendidas")
        dias = personas.get("dias_persona")
        if ingresos and atendidas and atendidas > max(ingresos * 3, ingresos + 100):
            flags.append(
                {
                    "code": "rem_personas_atendidas_outlier_high",
                    "severity": "review",
                    "period": period,
                    "ingresos": ingresos,
                    "personas_atendidas": atendidas,
                }
            )
        if ingresos and atendidas is not None and atendidas < ingresos * 0.6:
            flags.append(
                {
                    "code": "rem_personas_atendidas_outlier_low",
                    "severity": "review",
                    "period": period,
                    "ingresos": ingresos,
                    "personas_atendidas": atendidas,
                }
            )
        if dias and atendidas and atendidas > dias:
            flags.append(
                {
                    "code": "rem_personas_atendidas_exceeds_dias",
                    "severity": "review",
                    "period": period,
                    "personas_atendidas": atendidas,
                    "dias_persona": dias,
                }
            )

    rem_2024_ingresos = annual_rem.get(2024, {}).get("ingresos_rem")
    if rem_2024_ingresos is not None and abs(nominal_2024["admissions"] - rem_2024_ingresos) > 0:
        flags.append(
            {
                "code": "source_discrepancy_2024_ingresos",
                "severity": "review",
                "rem_ingresos": rem_2024_ingresos,
                "nominal_admissions": nominal_2024["admissions"],
            }
        )
    rem_2025_ingresos = annual_rem.get(2025, {}).get("ingresos_rem")
    if rem_2025_ingresos is not None and abs(hodom_2025["annual_total"] - rem_2025_ingresos) > 0:
        flags.append(
            {
                "code": "source_discrepancy_2025_ingresos",
                "severity": "review",
                "rem_ingresos": rem_2025_ingresos,
                "hodom_daily_admissions": hodom_2025["annual_total"],
            }
        )
    nominal_2025_periods = set(nominal_2025.get("monthly_admissions", {}))
    missing_nominal_2025 = sorted({f"2025-{month:02d}" for month in range(1, 13)} - nominal_2025_periods)
    if missing_nominal_2025:
        flags.append(
            {
                "code": "nominal_2025_coverage_gap",
                "severity": "review",
                "missing_periods": missing_nominal_2025,
                "nominal_admissions": nominal_2025["admissions"],
            }
        )
    if rem_2025_ingresos is not None and abs(nominal_2025["admissions"] - rem_2025_ingresos) > 0:
        flags.append(
            {
                "code": "source_discrepancy_2025_nominal_vs_rem",
                "severity": "review",
                "rem_ingresos": rem_2025_ingresos,
                "nominal_admissions": nominal_2025["admissions"],
            }
        )
    if db_metrics.get("available"):
        row_2025 = db_metrics["rows"].get(2025) or db_metrics["rows"].get("2025", {})
        if row_2025.get("visitas_rows") and not row_2025.get("visitas_realizadas"):
            flags.append(
                {
                    "code": "db_visits_not_realized_for_rem",
                    "severity": "info",
                    "year": 2025,
                    "visitas_rows": row_2025["visitas_rows"],
                    "visitas_realizadas": row_2025["visitas_realizadas"],
                }
            )
    return flags


def public_nominal(item: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in item.items() if key != "_user_ids"}


def build_unique_user_rows(nominal_2024: dict[str, Any], nominal_2025: dict[str, Any], database: dict[str, Any]) -> list[dict[str, Any]]:
    rows = [
        {
            "year": 2023,
            "unique_users": 340,
            "source": "Certificado PACIENTES 2023.docx",
            "note": "Ingresos segundo semestre; no entrega usuarios unicos anuales.",
        },
        {
            "year": 2024,
            "unique_users": nominal_2024["unique_users"],
            "source": "Planillas nominales 2024 por RUT",
            "note": f"{nominal_2024['admissions']} ingresos; {nominal_2024['missing_user_id']} sin RUT.",
        },
    ]
    if database.get("available"):
        db_2025 = database["rows"].get("2025") or database["rows"].get(2025, {})
        rows.append(
            {
                "year": 2025,
                "unique_users": db_2025.get("usuarios_ingresados"),
                "source": "Base local clinical.estadia",
                "note": (
                    f"{db_2025.get('estadia_ingresos')} estadias ingresadas. "
                    f"Planilla nominal parcial: {nominal_2025['admissions']} ingresos, "
                    f"{nominal_2025['unique_users']} RUT unicos, {nominal_2025['missing_user_id']} sin RUT."
                ),
            }
        )
    else:
        rows.append(
            {
                "year": 2025,
                "unique_users": nominal_2025["unique_users"],
                "source": "Planilla nominal 2025 parcial por RUT",
                "note": f"{nominal_2025['admissions']} ingresos; {nominal_2025['missing_user_id']} sin RUT.",
            }
        )
    return rows


def write_outputs(payload: dict[str, Any]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    with CSV_OUT.open("w", newline="", encoding="utf-8") as handle:
        columns = [
            "anio",
            "periodos_rem",
            "atenciones_rem_visitas",
            "usuarios_rem_personas_atendidas_suma_mensual",
            "ingresos_rem",
            "dias_persona",
            "usuarios_unicos_nominales_o_db",
            "fuente_usuarios_unicos",
        ]
        writer = csv.DictWriter(handle, fieldnames=columns, lineterminator="\n")
        writer.writeheader()
        unique_rows = {row["year"]: row for row in payload["unique_user_rows"]}
        for year in (2023, 2024, 2025):
            rem = payload["annual_rem"].get(str(year), {})
            unique_row = unique_rows.get(year, {})
            writer.writerow(
                {
                    "anio": year,
                    "periodos_rem": len(rem.get("periods", [])),
                    "atenciones_rem_visitas": rem.get("atenciones_rem_visitas"),
                    "usuarios_rem_personas_atendidas_suma_mensual": rem.get("usuarios_rem_personas_atendidas_suma_mensual"),
                    "ingresos_rem": rem.get("ingresos_rem"),
                    "dias_persona": rem.get("dias_persona"),
                    "usuarios_unicos_nominales_o_db": unique_row.get("unique_users"),
                    "fuente_usuarios_unicos": unique_row.get("source"),
                }
            )

    lines = [
        "# Especificacion de metricas anuales HODOM: atenciones y usuarios",
        "",
        f"Fuente Drive: {DRIVE_ROOT_URL}",
        f"Generado: {payload['generated_at']}",
        "",
        "## Resultado principal desde REM",
        "",
        "| Año | Periodos REM | Atenciones REM (visitas) | Usuarios REM (personas atendidas, suma mensual) | Ingresos REM | Días persona |",
        "| --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for year in (2023, 2024, 2025):
        rem = payload["annual_rem"].get(str(year), {})
        lines.append(
            f"| {year} | {len(rem.get('periods', []))} | {rem.get('atenciones_rem_visitas', '')} | "
            f"{rem.get('usuarios_rem_personas_atendidas_suma_mensual', '')} | {rem.get('ingresos_rem', '')} | {rem.get('dias_persona', '')} |"
        )

    lines.extend(
        [
            "",
            "## Usuarios unicos disponibles",
            "",
            "| Año | Usuarios/ingresos con identificador unico | Fuente | Nota |",
            "| --- | ---: | --- | --- |",
        ]
    )
    for row in payload["unique_user_rows"]:
        lines.append(f"| {row['year']} | {row.get('unique_users', '')} | {row['source']} | {row['note']} |")

    lines.extend(
        [
            "",
            "## Contraste 2025",
            "",
            f"- HODOM 2025 resumen diario: {payload['hodom_2025']['annual_total']} ingresos.",
            f"- Planilla nominal 2025 `INGRESOS 2025 (3).xlsx`: {payload['nominal_2025']['admissions']} ingresos con fecha 2025 y {payload['nominal_2025']['unique_users']} RUT unicos.",
        ]
    )
    if payload["database"].get("available"):
        db_2025 = payload["database"]["rows"]["2025"]
        lines.extend(
            [
                f"- Base local 2025: {db_2025['visitas_rows']} filas en `operational.visita`, {db_2025['usuarios_con_visita']} usuarios con visita registrada.",
                f"- Base local 2025: {db_2025['visitas_realizadas']} visitas en estados realizadas/REM; por eso no se usa como fuente principal de atenciones efectivas.",
            ]
        )

    lines.extend(["", "## Banderas de calidad", ""])
    for flag in payload["quality_flags"]:
        context = "; ".join(f"{key}={value}" for key, value in flag.items() if key not in {"code", "severity"})
        lines.append(f"- [{flag['severity']}] {flag['code']} · {context}")
    lines.extend(
        [
            "",
            "## Criterios",
            "",
            "- `Atenciones REM` suma las visitas de la seccion C.2 por profesion.",
            "- `Usuarios REM` es suma mensual de `Personas Atendidas`; no equivale a usuarios unicos anuales si una persona aparece en mas de un mes.",
            "- Esta especificacion es la fuente normativa versionada para consultas agregadas previas a migracion.",
            "- Las fuentes nominales se usan solo para conteo agregado; no se exportan RUT, nombres ni filas identificables.",
        ]
    )
    MD_OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(force_download: bool = False, include_database: bool = True) -> dict[str, Any]:
    rem_rows = []
    for source in REM_SOURCES:
        path = download_source(source, force=force_download)
        rem_rows.append({"period": source["period"], "source": source["path"], "metrics": extract_rem_a21_metrics(path)})

    annual_rem = aggregate_rem_years(rem_rows)
    nominal_parts = []
    for source in NOMINAL_2024_SOURCES:
        path = download_source(source, force=force_download)
        nominal_parts.append(extract_nominal_admissions(path, source["sheet"], 2024))
    nominal_2024 = merge_nominal_admissions(nominal_parts, 2024)

    nominal_2025 = extract_nominal_admissions(
        download_source(NOMINAL_2025_SOURCE, force=force_download),
        NOMINAL_2025_SOURCE["sheet"],
        2025,
    )
    hodom_2025 = extract_hodom_daily_admissions(download_source(HODOM_2025_SOURCE, force=force_download), 2025)
    certificate_2023 = extract_2023_certificate(download_source(DOC_2023_SOURCE, force=force_download))
    db_metrics = query_database_metrics() if include_database else {"available": False, "error": "database query disabled"}
    if db_metrics.get("available"):
        db_metrics["rows"] = {str(year): values for year, values in db_metrics["rows"].items()}

    annual_json = {str(year): values for year, values in annual_rem.items()}
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "drive_root_url": DRIVE_ROOT_URL,
        "annual_rem": annual_json,
        "monthly_rem": rem_rows,
        "nominal_2024": public_nominal(nominal_2024),
        "nominal_2024_parts": [public_nominal(item) for item in nominal_parts],
        "nominal_2025": public_nominal(nominal_2025),
        "hodom_2025": hodom_2025,
        "certificate_2023": certificate_2023,
        "database": db_metrics,
        "unique_user_rows": build_unique_user_rows(nominal_2024, nominal_2025, db_metrics),
        "quality_flags": build_quality_flags(annual_rem, rem_rows, db_metrics, nominal_2024, nominal_2025, hodom_2025),
        "outputs": {"json": str(JSON_OUT), "markdown": str(MD_OUT), "csv": str(CSV_OUT)},
    }
    write_outputs(payload)
    return payload


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--force-download", action="store_true")
    parser.add_argument("--no-database", action="store_true")
    args = parser.parse_args()
    payload = run(force_download=args.force_download, include_database=not args.no_database)
    print(json.dumps(payload["annual_rem"], ensure_ascii=False, indent=2))
    print(f"Wrote {MD_OUT}")


if __name__ == "__main__":
    main()
