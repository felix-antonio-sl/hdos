from __future__ import annotations

import csv
import re
import subprocess
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DOWNLOADS_DIR = Path.home() / "Downloads"
CANONICAL_STAYS = BASE_DIR / "output" / "spreadsheet" / "canonical" / "hospitalization_stay.csv"
CANONICAL_PATIENTS = BASE_DIR / "output" / "spreadsheet" / "canonical" / "patient_master.csv"
FORMS_2026 = DOWNLOADS_DIR / "FORMULARIO 2026 RESP (respuestas) (2).xlsx"
KINE_WORKBOOK = DOWNLOADS_DIR / "Ent. Turno Hodom KINE (1).xlsx"
TURNO_DIR = DOWNLOADS_DIR / "drive-download-20260401T185340Z-1-001"
OUTPUT_DIR = BASE_DIR / "output" / "clinical"
OUTPUT_MD = OUTPUT_DIR / "actualizacion_clinica_integral_hodom_2026-04-01.md"
OUTPUT_CSV = OUTPUT_DIR / "actualizacion_clinica_integral_hodom_2026-04-01.csv"


MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def normalize(text: str) -> str:
    text = unicodedata.normalize("NFKD", str(text or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_rut(value: str) -> str:
    value = normalize(value).replace(" ", "")
    if not value:
        return ""
    if "-" not in value and len(value) > 1:
        return f"{value[:-1]}-{value[-1]}"
    return value


def parse_iso_date(value: str) -> date | None:
    value = (value or "").strip()
    if not value:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def format_date(value: str) -> str:
    parsed = parse_iso_date(value)
    return parsed.strftime("%d-%m-%Y") if parsed else ""


def sex_label(value: str) -> str:
    token = (value or "").strip().upper()
    if token == "F":
        return "femenina"
    if token == "M":
        return "masculino"
    return "paciente"


def parse_turno_doc_date(path: Path) -> date | None:
    name = normalize(path.stem)
    match = re.search(r"ENTREGA TURNO (\d{1,2}) ?(\d{1,2})? ([A-Z]+) (20\d{2})", name)
    if not match:
        return None
    day = int(match.group(1))
    month_name = match.group(3)
    year = int(match.group(4))
    month = MONTHS.get(month_name)
    if not month:
        return None
    return date(year, month, day)


def parse_kine_sheet_date(name: str) -> date | None:
    normalized = normalize(name)
    match = re.search(r"(\d{1,2}) (\d{1,2})(?: (\d{2,4}))?", normalized)
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        year_token = match.group(3)
        if year_token:
            year = int(year_token)
            if year < 100:
                year += 2000
        else:
            year = 2026 if month <= 3 else 2025
        return date(year, month, day)
    return None


def looks_like_name_line(line: str) -> bool:
    normalized = normalize(line)
    if not normalized:
        return False
    if any(token in normalized for token in ["NOMBRE PACIENTE", "EDAD DIAGNOSTICO", "OBSERVACIONES", "PENDIENTES", "TTO EV", "INVASIVOS O2", "CS CA", "RHB"]):
        return False
    if re.search(r"\d", normalized):
        return False
    words = normalized.split()
    if len(words) < 2 or len(words) > 6:
        return False
    return True


@dataclass
class ClinicalCase:
    stay_id: str
    patient_id: str
    nombre: str
    rut: str
    sexo: str
    edad: str
    fecha_ingreso: str
    diagnostico: str
    servicio: str
    prevision: str
    establecimiento: str
    comuna: str
    localidad: str
    gestora: str
    usuario_o2: str
    source_episode_ids: str


def load_current_cases() -> list[ClinicalCase]:
    stays = read_csv(CANONICAL_STAYS)
    cases: list[ClinicalCase] = []
    for row in stays:
        if row.get("estado", "").upper() == "EGRESADO" and row.get("fecha_egreso"):
            continue
        cases.append(
            ClinicalCase(
                stay_id=row["stay_id"],
                patient_id=row["patient_id"],
                nombre=row["nombre_completo"],
                rut=row["rut"],
                sexo=row["sexo_resuelto"],
                edad=row["edad_reportada"],
                fecha_ingreso=row["fecha_ingreso"],
                diagnostico=row["diagnostico_principal"],
                servicio=row["servicio_origen"],
                prevision=row["prevision"],
                establecimiento=row["establecimiento"],
                comuna=row["comuna"],
                localidad=row["localidad"],
                gestora=row["gestora"],
                usuario_o2=row["usuario_o2"],
                source_episode_ids=row["source_episode_ids"],
            )
        )
    return cases


def load_form_index() -> dict[tuple[str, str], dict[str, str]]:
    wb = load_workbook(FORMS_2026, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    best: dict[tuple[str, str], dict[str, str]] = {}
    for row in rows[1:]:
        vals = ["" if value is None else str(value) for value in row]
        if len(vals) < 16:
            continue
        rut = normalize_rut(vals[3])
        name = normalize(f"{vals[1]} {vals[2]}")
        ts = vals[0]
        key = (rut, name)
        current = {
            "timestamp": ts,
            "servicio_origen_solicitud": vals[8],
            "diagnostico_form": vals[9],
            "direccion": vals[10],
            "cesfam": vals[11],
            "prestacion_solicitada": vals[13],
            "gestora": vals[15],
            "usuario_o2": vals[16] if len(vals) > 16 else "",
        }
        if key not in best or current["timestamp"] > best[key]["timestamp"]:
            best[key] = current
    return best


def load_kine_index() -> dict[tuple[str, str], dict[str, str]]:
    wb = load_workbook(KINE_WORKBOOK, data_only=True, read_only=True)
    best: dict[tuple[str, str], dict[str, str]] = {}
    for sheet_name in wb.sheetnames:
        sheet_date = parse_kine_sheet_date(sheet_name)
        if sheet_date is None:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=6, values_only=True):
            vals = ["" if value is None else str(value) for value in row]
            if not vals or not vals[0]:
                continue
            name = normalize(vals[0])
            rut = normalize_rut(vals[1] if len(vals) > 1 else "")
            key = (rut, name)
            current = {
                "sheet_date": sheet_date.isoformat(),
                "cobertura": vals[2] if len(vals) > 2 else "",
                "diagnostico_kine": vals[3] if len(vals) > 3 else "",
                "observaciones_kine": vals[4] if len(vals) > 4 else "",
                "hora_atencion": vals[5] if len(vals) > 5 else "",
                "registro_fc": vals[6] if len(vals) > 6 else "",
            }
            if key not in best or current["sheet_date"] > best[key]["sheet_date"]:
                best[key] = current
    return best


def load_turno_index(cases: list[ClinicalCase]) -> dict[str, dict[str, str]]:
    active_names = {normalize(case.nombre) for case in cases}
    best: dict[str, dict[str, str]] = {}
    for path in sorted(TURNO_DIR.glob("*.docx")):
        doc_date = parse_turno_doc_date(path)
        if doc_date is None or doc_date.year != 2026:
            continue
        try:
            proc = subprocess.run(
                ["textutil", "-convert", "txt", "-stdout", str(path)],
                check=True,
                capture_output=True,
                text=True,
            )
        except subprocess.CalledProcessError:
            continue
        raw_lines = [line.strip() for line in proc.stdout.splitlines() if line.strip()]
        normalized_lines = [normalize(line) for line in raw_lines]
        for idx, norm_line in enumerate(normalized_lines):
            if norm_line not in active_names:
                continue
            key = norm_line
            snippet_lines: list[str] = []
            for probe in range(idx, min(idx + 12, len(raw_lines))):
                candidate = raw_lines[probe].strip()
                if probe > idx and looks_like_name_line(candidate):
                    break
                snippet_lines.append(candidate)
            current = {
                "doc_date": doc_date.isoformat(),
                "source_file": path.name,
                "snippet": " | ".join(snippet_lines),
            }
            if key not in best or current["doc_date"] > best[key]["doc_date"]:
                best[key] = current
    return best


def choose_form(case: ClinicalCase, forms: dict[tuple[str, str], dict[str, str]]) -> dict[str, str]:
    rut_key = normalize_rut(case.rut)
    name_key = normalize(case.nombre)
    if rut_key:
        for (form_rut, _), payload in forms.items():
            if form_rut == rut_key:
                return payload
    return forms.get((rut_key, name_key), {})


def choose_kine(case: ClinicalCase, kine: dict[tuple[str, str], dict[str, str]]) -> dict[str, str]:
    rut_key = normalize_rut(case.rut)
    name_key = normalize(case.nombre)
    if rut_key:
        by_rut = [payload for (k_rut, _), payload in kine.items() if k_rut == rut_key]
        if by_rut:
            return sorted(by_rut, key=lambda payload: payload.get("sheet_date", ""), reverse=True)[0]
    return kine.get((rut_key, name_key), {})


def choose_turno(case: ClinicalCase, turno: dict[str, dict[str, str]]) -> dict[str, str]:
    name_key = normalize(case.nombre)
    if name_key in turno:
        return turno[name_key]
    for turno_name, payload in turno.items():
        if turno_name in name_key or name_key in turno_name:
            return payload
    return {}


def is_current_case(case: ClinicalCase, form: dict[str, str], kine: dict[str, str], turno: dict[str, str]) -> bool:
    ingreso = parse_iso_date(case.fecha_ingreso)
    if ingreso and ingreso >= date(2026, 1, 1):
        return True
    return bool(form or kine or turno)


def sentence_case_intro(case: ClinicalCase, form: dict[str, str]) -> str:
    sexo = sex_label(case.sexo)
    edad = case.edad or "s/d"
    diagnostico = case.diagnostico or form.get("diagnostico_form", "sin diagnóstico consignado")
    servicio = case.servicio or form.get("servicio_origen_solicitud", "")
    ingreso = format_date(case.fecha_ingreso)
    parts = [f"Paciente {sexo} de {edad} años"]
    if diagnostico:
        parts.append(f"en seguimiento por {diagnostico.lower()}")
    if ingreso:
        parts.append(f"con ingreso HODOM vigente desde {ingreso}")
    if servicio:
        parts.append(f"derivado desde {servicio}")
    return ", ".join(parts) + "."


def build_current_status(case: ClinicalCase, form: dict[str, str], kine: dict[str, str], turno: dict[str, str]) -> str:
    fragments: list[str] = []
    if form.get("prestacion_solicitada"):
        fragments.append(f"Prestaciones solicitadas al ingreso: {form['prestacion_solicitada']}.")
    if kine.get("observaciones_kine"):
        fragments.append(
            f"Último registro KINE {format_date(kine.get('sheet_date', '')) or kine.get('sheet_date', '')}: "
            f"{kine['observaciones_kine']}."
        )
    if turno.get("snippet"):
        snippet = turno["snippet"]
        fragments.append(f"Última entrega de turno ({turno['source_file']}): {snippet}.")
    if not fragments:
        fragments.append(
            "Sin registro narrativo adicional en los insumos operativos revisados; mantener evaluación clínica directa en visita."
        )
    return " ".join(fragments)


def extract_plan_items(case: ClinicalCase, form: dict[str, str], kine: dict[str, str], turno: dict[str, str]) -> list[str]:
    items: list[str] = []
    text_sources = " | ".join(
        value
        for value in [
            form.get("prestacion_solicitada", ""),
            kine.get("observaciones_kine", ""),
            turno.get("snippet", ""),
        ]
        if value
    )
    upper = normalize(text_sources)
    if "CURACION" in upper:
        items.append("Mantener curaciones según indicación vigente y reevaluar respuesta local.")
    if "KTR" in upper or "RESPIRATORIA" in upper:
        items.append("Mantener kinesioterapia respiratoria y vigilancia de tolerancia ventilatoria.")
    if "KTM" in upper or "MARCHA" in upper or "REHABILITACION" in upper or "REHABILITACION" in upper:
        items.append("Mantener rehabilitación motora con progresión según tolerancia funcional.")
    if "FONO" in upper:
        items.append("Continuar apoyo fonoaudiológico según indicación clínica.")
    if "TTO EV" in upper or "ATB" in upper or "ERTA" in upper:
        items.append("Continuar terapia endovenosa indicada, con control de respuesta clínica y vigilancia de accesos.")
    if "EXAM" in upper or "CONTROL" in upper:
        items.append("Revisar resultados de exámenes y controles pendientes en próxima visita médica.")
    if "HORA" in upper:
        items.append("Verificar cumplimiento de interconsultas, controles y horas ya gestionadas.")
    if "O2" in upper or "OXIG" in upper:
        items.append("Monitorizar requerimiento de oxígeno y signos de descompensación respiratoria.")
    if not items:
        items.append("Mantener seguimiento clínico evolutivo y reevaluar necesidad de prestaciones en próxima visita.")
    # de-dup preserve order
    deduped: list[str] = []
    for item in items:
        if item not in deduped:
            deduped.append(item)
    return deduped


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    cases = load_current_cases()
    form_index = load_form_index()
    kine_index = load_kine_index()
    turno_index = load_turno_index(cases)

    selected: list[dict[str, str]] = []
    for case in cases:
        form = choose_form(case, form_index)
        kine = choose_kine(case, kine_index)
        turno = choose_turno(case, turno_index)
        if not is_current_case(case, form, kine, turno):
            continue
        plan_items = extract_plan_items(case, form, kine, turno)
        record = {
            "stay_id": case.stay_id,
            "nombre_completo": case.nombre,
            "rut": case.rut,
            "fecha_ingreso": case.fecha_ingreso,
            "diagnostico_principal": case.diagnostico,
            "actualizacion": sentence_case_intro(case, form) + " " + build_current_status(case, form, kine, turno),
            "plan": " ".join(f"- {item}" for item in plan_items),
            "gestora": form.get("gestora", case.gestora),
            "prestacion_solicitada": form.get("prestacion_solicitada", ""),
            "ultimo_turno": turno.get("source_file", ""),
            "ultimo_kine": kine.get("sheet_date", ""),
        }
        selected.append(record)

    selected.sort(key=lambda row: (row["fecha_ingreso"] or "9999-99-99", row["nombre_completo"]))

    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(selected[0].keys()) if selected else [])
        if selected:
            writer.writeheader()
            writer.writerows(selected)

    lines = ["# Actualización clínica integral HODOM", "", f"Casos vigentes incluidos: {len(selected)}", ""]
    for row in selected:
        lines.append(f"## {row['nombre_completo']} ({row['rut'] or 'sin RUT'})")
        lines.append("")
        lines.append(row["actualizacion"])
        lines.append("")
        lines.append("Plan actual:")
        for part in row["plan"].split(" - "):
            part = part.strip()
            if not part:
                continue
            if not part.startswith("-"):
                part = f"- {part}"
            lines.append(part)
        if row["prestacion_solicitada"]:
            lines.append("")
            lines.append(f"Prestaciones solicitadas: {row['prestacion_solicitada']}")
        if row["ultimo_turno"] or row["ultimo_kine"]:
            lines.append("")
            lines.append(
                f"Fuente operativa más reciente: turno={row['ultimo_turno'] or 's/d'}; kine={row['ultimo_kine'] or 's/d'}."
            )
        lines.append("")

    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Generated {len(selected)} updates")
    print(OUTPUT_MD)
    print(OUTPUT_CSV)


if __name__ == "__main__":
    main()
