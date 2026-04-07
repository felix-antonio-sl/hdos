# scripts/migrate_to_pg/functors/f10_kpi_diario.py
"""
F₁₀: KPI Diario — legacy/Consolidado atenciones diarias.xlsx → reporting.kpi_diario

Source:  legacy XLSX (88 days, trust=0.4)
Target:  reporting.kpi_diario

Pre-requisite: territorial.zona must have at least one zone.
This functor creates a default zone 'zone_hodom_hsc' if none exists.

The XLSX has daily visit counts by profession (enfermero, kinesiologo, fonoaudiologo,
medico, tecnico). We map the sum to visitas_realizadas. The table also computes
pacientes_activos from clinical.estadia for each date.

Functor Information Loss:
  - Per-profession breakdown is collapsed into a single visitas_realizadas count.
  - Provenance preserves the per-profession values as field-level records.
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path

import psycopg

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

try:
    from ..framework.category import Functor, PathEquation
    from ..framework.provenance import NaturalTransformation
    from ..framework.hash_ids import make_id
except ImportError:
    from framework.category import Functor, PathEquation  # type: ignore[no-redef]
    from framework.provenance import NaturalTransformation  # type: ignore[no-redef]
    from framework.hash_ids import make_id  # type: ignore[no-redef]


ZONE_ID = "zone_hodom_hsc"
ZONE_NAME = "HODOM Hospital San Carlos"

XLSX_RELPATH = (
    "DATOS ESTADÍSTICOS /Consolidado atenciones diarias.xlsx"
)

PROFESION_COLS = ["enfermero", "kinesiologo", "fonoaudiologo", "medico", "tecnico"]


class F10KpiDiario(Functor):
    name = "F10_kpi_diario"
    depends_on = ["F0_bootstrap", "F3_estadias"]

    def objects(self, conn: psycopg.Connection, sources) -> int:
        if openpyxl is None:
            raise RuntimeError("openpyxl required for F10 — pip install openpyxl")

        eta = NaturalTransformation()

        # Idempotency
        conn.execute("DELETE FROM reporting.kpi_diario WHERE TRUE")

        # Ensure default zone exists
        conn.execute(
            """
            INSERT INTO territorial.zona (zone_id, nombre, tipo)
            VALUES (%s, %s, 'URBANO')
            ON CONFLICT (zone_id) DO NOTHING
            """,
            (ZONE_ID, ZONE_NAME),
        )

        # Read XLSX
        xlsx_path = sources.legacy_dir / XLSX_RELPATH
        if not xlsx_path.exists():
            return 0

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        header = [str(h).strip().lower() if h else "" for h in rows[0]]

        n = 0
        for row in rows[1:]:
            fecha_raw = row[0]
            if not fecha_raw:
                continue

            if isinstance(fecha_raw, datetime):
                fecha = fecha_raw.date()
            else:
                continue

            # Parse profession counts
            vals = {}
            total_visitas = 0
            for i, col in enumerate(PROFESION_COLS):
                idx = i + 1  # columns after fecha
                v = row[idx] if idx < len(row) else None
                if v is not None:
                    try:
                        count = int(float(v))
                    except (ValueError, TypeError):
                        count = 0
                else:
                    count = 0
                vals[col] = count
                total_visitas += count

            if total_visitas == 0:
                continue

            # Compute pacientes_activos for this date from clinical.estadia
            active_row = conn.execute(
                """
                SELECT COUNT(*) FROM clinical.estadia
                WHERE fecha_ingreso <= %s
                  AND (fecha_egreso IS NULL OR fecha_egreso >= %s)
                """,
                (fecha, fecha),
            ).fetchone()
            pacientes_activos = active_row[0] if active_row else 0

            conn.execute(
                """
                INSERT INTO reporting.kpi_diario
                    (fecha, zone_id, pacientes_activos, visitas_realizadas)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (fecha, zone_id) DO UPDATE SET
                    pacientes_activos = EXCLUDED.pacientes_activos,
                    visitas_realizadas = EXCLUDED.visitas_realizadas
                """,
                (fecha, ZONE_ID, pacientes_activos, total_visitas),
            )

            # Provenance: per-profession breakdown
            for prof, count in vals.items():
                if count > 0:
                    eta.record(
                        conn,
                        target_table="reporting.kpi_diario",
                        target_pk=f"{fecha}|{ZONE_ID}",
                        source_type="legacy",
                        source_file="Consolidado atenciones diarias.xlsx",
                        source_key=f"{fecha}|{prof}={count}",
                        phase="F10",
                        field_name=f"visitas_{prof}",
                    )

            n += 1

        return n

    def path_equations(self) -> list[PathEquation]:
        return [
            PathEquation(
                name="PE-F10-FK-ZONE",
                sql="""SELECT fecha FROM reporting.kpi_diario k
                    LEFT JOIN territorial.zona z ON z.zone_id = k.zone_id
                    WHERE z.zone_id IS NULL""",
                expected="empty",
            ),
            PathEquation(
                name="PE-F10-POSITIVE",
                sql="""SELECT fecha FROM reporting.kpi_diario
                    WHERE visitas_realizadas < 0 OR pacientes_activos < 0""",
                expected="empty",
            ),
            PathEquation(
                name="PE-F10-DATE-RANGE",
                sql="""SELECT COUNT(*) FROM reporting.kpi_diario
                    WHERE fecha < '2025-01-01' OR fecha > '2027-01-01'""",
                expected=0,
            ),
        ]
