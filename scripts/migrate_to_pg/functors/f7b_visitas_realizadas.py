# scripts/migrate_to_pg/functors/f7b_visitas_realizadas.py
"""
F₇b: Visitas Realizadas — legacy RUTAS XLSX → operational.visita (estado=COMPLETA)

Source:  RUTAS/RUTAS 2025/ (12 files) + RUTAS/2026/ (4 files)
Target:  operational.visita

Unlike F₇ (programación → PROGRAMADA), this functor loads actual visits
from daily route sheets with:
  - fecha, hora (real schedule)
  - paciente (nombre → fuzzy match to clinical.paciente)
  - prestación (KTM, TTO EV, FONO, CA, CS, etc.)
  - profesionales asignados (columns: MEDICO, FONO, KINE, ENFERMERO, TENS)
  - dirección del paciente

The functor is a coproduct with F₇:
  visita_total = F₇(PROGRAMADA) ⊕ F₇b(COMPLETA)

Deduplication: if a (patient_id, fecha, prestación) already exists from F₇,
F₇b updates it to COMPLETA and adds provider/hora. Otherwise inserts new.

Categorical note: F₇b acts as a left Kan extension of the route data
along the inclusion functor RUTAS ↪ operational.visita, enriching
existing programmed visits with execution data.
"""

from __future__ import annotations
import re
import unicodedata
import glob
from datetime import datetime, date
from pathlib import Path

import psycopg

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

try:
    from ..framework.category import Functor, PathEquation
    from ..framework.provenance import NaturalTransformation
    from ..framework.hash_ids import make_id
except ImportError:
    from framework.category import Functor, PathEquation  # type: ignore[no-redef]
    from framework.provenance import NaturalTransformation  # type: ignore[no-redef]
    from framework.hash_ids import make_id  # type: ignore[no-redef]


def _normalize_name(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().upper()


def _classify_prestacion(raw: str) -> str:
    """Reuse the same normalization as F₇/CORR-06."""
    u = raw.upper().strip()
    if not u or u in ("$", "-", "?", "."):
        return ""

    has_ktm = bool(re.search(r"KTM|KT\s*MOTORA", u))
    has_ktr = bool(re.search(r"KTR|KT\s*RESP", u))
    has_fono = "FONO" in u
    has_ca = bool(re.search(r"\bCA\b|CURACION|CURACIÓN", u))
    has_cs = bool(re.search(r"\bCS\b|CONTROL.*SIGNOS", u))
    has_tto = bool(re.search(r"TTO|CEF|ATB|ERTA|HIDRO|CLOX", u))
    has_npt = bool(re.search(r"NPT|NTP", u))
    has_exam = bool(re.search(r"EXAM", u))
    has_ing = bool(re.search(r"ING\w*\s*ENF|INGRESO", u))
    has_vm_ing = "VM INGRESO" in u or "VM ING" in u or "VM MINGRESO" in u
    has_vm_egr = "VM EGRESO" in u or "VM EGR" in u
    has_alta = "ALTA" in u

    if has_alta and ("HODOM" in u or has_vm_egr):
        return "ALTA_HODOM"
    if has_alta and (has_ktm or has_ktr):
        return "ALTA_KINE"
    if has_alta and has_fono:
        return "ALTA_FONO"

    parts = []
    if has_ing: parts.append("ING_ENF")
    if has_tto: parts.append("TTO_EV")
    if has_npt: parts.append("NPT")
    if has_ktm: parts.append("KTM")
    if has_ktr: parts.append("KTR")
    if has_fono: parts.append("FONO")
    if has_ca: parts.append("CA")
    if has_cs: parts.append("CS")
    if has_exam: parts.append("EXAM")
    if has_vm_ing: parts.append("VM_ING")
    if has_vm_egr: parts.append("VM_EGR")

    return "+".join(parts) if parts else "OTRO"


# Column indices in route sheets
COL_CONDUCTOR = 0
COL_HORA = 1
COL_MEDICO = 2
COL_FONO = 3
COL_KINE = 4
COL_ENFERMERO = 5
COL_TENS = 6
COL_PACIENTE = 7
COL_PRESTACION = 8
COL_DIRECCION = 9
COL_TELEFONO = 10

PROFESION_COLS = {
    COL_MEDICO: "MEDICO",
    COL_FONO: "FONOAUDIOLOGIA",
    COL_KINE: "KINESIOLOGIA",
    COL_ENFERMERO: "ENFERMERIA",
    COL_TENS: "TENS",
}

ROUTE_DIRS = [
    "RUTAS/RUTAS 2025",
    "RUTAS/ENERO 2026.xlsx",
    "RUTAS/FEBRERO 2026.xlsx",
    "RUTAS/MARZO 2026.xlsx",
    "RUTAS/ABRIL 2026.xlsx",
]


class F7bVisitasRealizadas(Functor):
    name = "F7b_visitas_realizadas"
    depends_on = ["F0_bootstrap", "F2_pacientes", "F3_estadias", "F6_profesionales"]

    def objects(self, conn: psycopg.Connection, sources) -> int:
        if openpyxl is None:
            raise RuntimeError("openpyxl required for F7b")

        eta = NaturalTransformation()

        # Build patient name index
        patient_index: dict[str, str] = {}
        for pid, nombre in conn.execute(
            "SELECT patient_id, nombre_completo FROM clinical.paciente"
        ).fetchall():
            norm = _normalize_name(nombre)
            patient_index[norm] = pid
            parts = norm.split()
            if len(parts) >= 2:
                patient_index[" ".join(parts[-2:])] = pid
                if len(parts) >= 3:
                    patient_index[f"{parts[0]} {parts[-2]} {parts[-1]}"] = pid

        # Build provider name index
        provider_index: dict[str, str] = {}
        for prov_id, nombre in conn.execute(
            "SELECT provider_id, nombre FROM operational.profesional"
        ).fetchall():
            norm = _normalize_name(nombre)
            provider_index[norm] = prov_id
            # Also first name only
            parts = norm.split()
            if parts:
                provider_index[parts[0]] = prov_id

        # Build stay lookup
        stay_index: dict[str, list[tuple[str, date, date | None]]] = {}
        for stay_id, patient_id, fi, fe in conn.execute(
            "SELECT stay_id, patient_id, fecha_ingreso, fecha_egreso FROM clinical.estadia ORDER BY fecha_ingreso"
        ).fetchall():
            stay_index.setdefault(patient_id, []).append((stay_id, fi, fe))

        # Build existing visit index for dedup/upgrade
        existing_visits: dict[tuple[str, date, str], str] = {}
        for vid, pid, fecha, prest in conn.execute(
            "SELECT visit_id, patient_id, fecha, rem_prestacion FROM operational.visita"
        ).fetchall():
            existing_visits[(pid, fecha, prest or "")] = vid

        # Disable state-machine trigger for bulk load
        conn.execute("ALTER TABLE operational.visita DISABLE TRIGGER ALL")

        n = 0
        xlsx_files = []
        for entry in ROUTE_DIRS:
            full = sources.legacy_dir / entry
            if full.is_dir():
                xlsx_files.extend(sorted(full.glob("*.xlsx")))
            elif full.exists():
                xlsx_files.append(full)

        for xlsx_path in xlsx_files:
            try:
                wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            except Exception:
                continue

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 3:
                    continue

                # Row 0: date
                fecha_val = rows[0][0] if rows[0] else None
                if not isinstance(fecha_val, datetime):
                    continue
                fecha = fecha_val.date()
                if fecha.year < 2025:
                    continue

                # Rows 2+: visit data
                for row in rows[2:]:
                    if not row or len(row) <= COL_PRESTACION:
                        continue

                    paciente_raw = str(row[COL_PACIENTE]).strip() if row[COL_PACIENTE] else ""
                    if not paciente_raw or len(paciente_raw) < 4:
                        continue

                    prestacion_raw = str(row[COL_PRESTACION]).strip() if row[COL_PRESTACION] else ""
                    if not prestacion_raw:
                        continue

                    # Match patient
                    norm_pac = _normalize_name(paciente_raw)
                    patient_id = patient_index.get(norm_pac)
                    if not patient_id:
                        for idx_name, pid in patient_index.items():
                            if norm_pac in idx_name or idx_name in norm_pac:
                                patient_id = pid
                                break
                    if not patient_id:
                        continue

                    # Find stay
                    stays = stay_index.get(patient_id, [])
                    stay_id = None
                    for sid, fi, fe in stays:
                        if fi <= fecha and (fe is None or fe >= fecha):
                            stay_id = sid
                            break
                    if not stay_id:
                        continue

                    prestacion = _classify_prestacion(prestacion_raw)
                    if not prestacion:
                        continue

                    # Hour
                    hora_raw = row[COL_HORA] if len(row) > COL_HORA else None
                    hora = None
                    if hora_raw:
                        if isinstance(hora_raw, datetime):
                            hora = hora_raw.strftime("%H:%M")
                        else:
                            h = str(hora_raw).strip()
                            if re.match(r"\d{1,2}:\d{2}", h):
                                hora = h

                    # Resolve providers from columns
                    providers = []
                    for col_idx, profesion in PROFESION_COLS.items():
                        if col_idx < len(row) and row[col_idx]:
                            prov_name = _normalize_name(str(row[col_idx]))
                            if prov_name and len(prov_name) >= 2:
                                prov_id = provider_index.get(prov_name)
                                if prov_id:
                                    providers.append(prov_id)

                    primary_provider = providers[0] if providers else None

                    # Check if exists from F7 (upgrade PROGRAMADA → COMPLETA)
                    existing_key = (patient_id, fecha, prestacion)
                    existing_vid = existing_visits.get(existing_key)

                    if existing_vid:
                        conn.execute(
                            """UPDATE operational.visita
                               SET estado = 'COMPLETA', hora_plan_inicio = %s,
                                   provider_id = COALESCE(provider_id, %s),
                                   updated_at = NOW()
                               WHERE visit_id = %s AND estado = 'PROGRAMADA'""",
                            (hora, primary_provider, existing_vid),
                        )
                    else:
                        visit_id = make_id("vr", f"{patient_id}|{fecha}|{prestacion}")
                        if visit_id in {v for v in existing_visits.values()}:
                            continue

                        conn.execute(
                            """INSERT INTO operational.visita
                                (visit_id, stay_id, patient_id, provider_id,
                                 fecha, hora_plan_inicio, estado, rem_prestacion)
                               VALUES (%s, %s, %s, %s, %s, %s, 'COMPLETA', %s)
                               ON CONFLICT (visit_id) DO NOTHING""",
                            (visit_id, stay_id, patient_id, primary_provider,
                             fecha, hora, prestacion),
                        )

                    eta.record(
                        conn,
                        target_table="operational.visita",
                        target_pk=existing_vid or visit_id,
                        source_type="legacy",
                        source_file=xlsx_path.name,
                        source_key=f"{fecha}|{norm_pac}|{sheet_name}",
                        phase="F7b",
                    )
                    n += 1

            wb.close()

        # Re-enable triggers
        conn.execute("ALTER TABLE operational.visita ENABLE TRIGGER ALL")

        return n

    def path_equations(self) -> list[PathEquation]:
        return [
            PathEquation(
                name="PE-F7b-FK-STAY",
                sql="""SELECT visit_id FROM operational.visita
                    WHERE stay_id NOT IN (SELECT stay_id FROM clinical.estadia)""",
                expected="empty",
            ),
            PathEquation(
                name="PE-F7b-FK-PATIENT",
                sql="""SELECT visit_id FROM operational.visita
                    WHERE patient_id NOT IN (SELECT patient_id FROM clinical.paciente)""",
                expected="empty",
            ),
            PathEquation(
                name="PE-F7b-COMPLETAS",
                sql="SELECT COUNT(*) FROM operational.visita WHERE estado = 'COMPLETA'",
                expected=0,  # will be > 0
                severity="warning",
            ),
        ]
