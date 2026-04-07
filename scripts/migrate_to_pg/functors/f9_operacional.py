# scripts/migrate_to_pg/functors/f9_operacional.py
"""
F₉: Operacional — legacy XLSX → operational.registro_llamada + operational.canasta_valorizada

Sources:
  - LLAMADAS/REGISTRO LLAMADAS.xlsx (7 sheets, ~397 llamadas)
  - CANASTA/CANASTA HODOM.xlsx (catálogo MAI, no patient-level)

For llamadas: each row has fecha, hora, duración, teléfono, motivo, usuario HODOM,
familiar, estado (ACT/EGR), tipo (EMITIDA/RECIBIDA), funcionario, observaciones.
We match "USUARIO HODOM" to clinical.paciente by name.

For canasta: it's a reference catalog (prestaciones MAI), not patient-level data.
We load it into canasta_valorizada as a reference row per prestación type.
"""

from __future__ import annotations
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import psycopg

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

try:
    from ..framework.category import Functor, PathEquation
    from ..framework.provenance import NaturalTransformation
    from ..framework.hash_ids import make_id
except ImportError:
    from framework.category import Functor, PathEquation  # type: ignore[no-redef]
    from framework.provenance import NaturalTransformation  # type: ignore[no-redef]
    from framework.hash_ids import make_id  # type: ignore[no-redef]


def _normalize(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().upper()


class F9Operacional(Functor):
    name = "F9_operacional"
    depends_on = ["F0_bootstrap", "F2_pacientes", "F3_estadias"]

    def objects(self, conn: psycopg.Connection, sources) -> int:
        if openpyxl is None:
            raise RuntimeError("openpyxl required for F9")

        eta = NaturalTransformation()

        # Idempotency
        conn.execute("DELETE FROM operational.registro_llamada WHERE TRUE")

        n_llamadas = self._load_llamadas(conn, sources, eta)

        return n_llamadas

    def _load_llamadas(self, conn, sources, eta) -> int:
        xlsx_path = sources.legacy_dir / "LLAMADAS" / "REGISTRO LLAMADAS.xlsx"
        if not xlsx_path.exists():
            return 0

        # Build patient name index
        patient_index: dict[str, str] = {}
        for pid, nombre in conn.execute(
            "SELECT patient_id, nombre_completo FROM clinical.paciente"
        ).fetchall():
            norm = _normalize(nombre)
            patient_index[norm] = pid

        # Build patient_id → latest stay
        stay_lookup: dict[str, str] = {}
        for stay_id, patient_id in conn.execute(
            """SELECT stay_id, patient_id FROM clinical.estadia
               ORDER BY fecha_ingreso DESC"""
        ).fetchall():
            if patient_id not in stay_lookup:
                stay_lookup[patient_id] = stay_id

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        n = 0

        for sheet_name in wb.sheetnames:
            if sheet_name == "NO MODIFICAR":
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            # Header: FECHA, HORA, DURACIÓN, NRO. TELÉFONO, MOTIVO,
            #          USUARIO HODOM, NOMBRE FAMILIAR, ACT/EGR,
            #          TIPO DE LLAMADA, FUNCIONARIO HD, OBSERVACIONES

            for row in rows[1:]:
                if len(row) < 10:
                    continue
                fecha_raw = row[0]
                if not fecha_raw:
                    continue

                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw.date()
                else:
                    continue

                hora = str(row[1]).strip() if row[1] else None
                duracion = str(row[2]).strip() if row[2] else None
                telefono = str(row[3]).strip().replace(".0", "") if row[3] else None
                usuario = str(row[5]).strip() if row[5] else ""
                familiar = str(row[6]).strip() if row[6] else None
                estado_pac_raw = str(row[7]).strip().upper() if row[7] else None
                estado_pac = {
                    "ACTIVO": "activo", "ACT": "activo",
                    "EGRESADO": "egresado", "EGR": "egresado", "EGRESADA": "egresado",
                }.get(estado_pac_raw)

                tipo_raw = str(row[8]).strip().upper() if row[8] else None
                tipo = {
                    "EMITIDA": "emitida", "RECIBIDA": "recibida",
                }.get(tipo_raw)

                motivo_raw = str(row[4]).strip().upper() if row[4] else None
                motivo = {
                    "RESULTADO EX": "resultado_examen",
                    "RESULTADO EXAMENES": "resultado_examen",
                    "OTRO": "otro",
                    "COORDINACION": "coordinacion",
                    "COORDINACIÓN": "coordinacion",
                    "SEGUIMIENTO": "seguimiento",
                    "CONSULTA": "consulta_clinica",
                    "ASISTENCIA SOCIAL": "asistencia_social",
                }.get(motivo_raw, "otro" if motivo_raw else None)
                # row[9] = funcionario
                obs = str(row[10]).strip() if len(row) > 10 and row[10] else None

                # Match patient
                patient_id = None
                if usuario:
                    norm_usuario = _normalize(usuario)
                    patient_id = patient_index.get(norm_usuario)
                    if not patient_id:
                        # Fuzzy
                        for idx_name, pid in patient_index.items():
                            if norm_usuario in idx_name or idx_name in norm_usuario:
                                patient_id = pid
                                break

                stay_id = stay_lookup.get(patient_id) if patient_id else None

                if not tipo:
                    tipo = "emitida"  # default when not specified

                llamada_id = make_id("llam", f"{fecha}|{hora or ''}|{usuario}|{motivo or ''}")

                conn.execute(
                    """
                    INSERT INTO operational.registro_llamada
                        (llamada_id, fecha, hora, duracion, telefono, motivo,
                         patient_id, stay_id, nombre_familiar,
                         estado_paciente, tipo, observaciones)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (llamada_id) DO NOTHING
                    """,
                    (
                        llamada_id, fecha, hora, duracion, telefono, motivo,
                        patient_id, stay_id, familiar,
                        estado_pac, tipo, obs,
                    ),
                )

                eta.record(
                    conn,
                    target_table="operational.registro_llamada",
                    target_pk=llamada_id,
                    source_type="legacy",
                    source_file="REGISTRO LLAMADAS.xlsx",
                    source_key=f"{fecha}|{usuario}",
                    phase="F9",
                )
                n += 1

        wb.close()
        return n

    def path_equations(self) -> list[PathEquation]:
        return [
            PathEquation(
                name="PE-F9-LLAMADA-DATE",
                sql="""SELECT llamada_id FROM operational.registro_llamada
                    WHERE fecha < '2024-01-01' OR fecha > '2027-01-01'""",
                expected="empty",
            ),
            PathEquation(
                name="PE-F9-PATIENT-FK",
                sql="""SELECT llamada_id FROM operational.registro_llamada
                    WHERE patient_id IS NOT NULL
                      AND patient_id NOT IN (SELECT patient_id FROM clinical.paciente)""",
                expected="empty",
            ),
        ]
