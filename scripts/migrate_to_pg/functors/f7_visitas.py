# scripts/migrate_to_pg/functors/f7_visitas.py
"""
F₇: Visitas — legacy PROGRAMACIÓN XLSX → operational.visita

Source:  2025 PROGRAMACIÓN.xlsx (13 sheets) + PROGRAMACIÓN 2026.xlsx (4 sheets)
Target:  operational.visita

Each sheet is a monthly schedule grid:
  Row 0: header — cols 10+ are dates
  Row 1+: patient rows
    [0]=número, [1]=nombre, [2]=edad, [3]=RUT, [4]=dirección, [5]=teléfono,
    [6]=diagnóstico, [7]=?, [8]=fecha_ingreso, [9]=fecha_egreso?,
    [10+]=daily visit type (ERTA, NPT, KTM, CA, etc.) or empty

Pullback over clinical.paciente by RUT + clinical.estadia by (patient_id, fecha):
  - Only patients with RUT in clinical.paciente
  - Only dates that fall within an active estadia

Visit types are mapped to estado 'PROGRAMADA' (we only have the schedule, not execution).

Functor Information Loss:
  - Visit type abbreviations (ERTA, NPT, KTM, CA, etc.) are stored as-is in resultado field;
    no standard prestacion_id mapping exists yet.
  - No provider assignment — the schedule doesn't specify which professional visits.
"""

from __future__ import annotations
from datetime import datetime, date
from pathlib import Path

import psycopg

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

try:
    from ..framework.category import Functor, PathEquation
    from ..framework.provenance import NaturalTransformation
    from ..framework.hash_ids import make_id, patient_id_from_rut
except ImportError:
    from framework.category import Functor, PathEquation  # type: ignore[no-redef]
    from framework.provenance import NaturalTransformation  # type: ignore[no-redef]
    from framework.hash_ids import make_id, patient_id_from_rut  # type: ignore[no-redef]


PROGRAMACION_FILES = [
    "2025 PROGRAMACIÓN.xlsx",
    "PROGRAMACIÓN 2026.xlsx",
]

# Column layout in programación sheets
COL_RUT = 3
COL_FECHA_INGRESO = 8
COL_FECHA_EGRESO = 9
COL_DATES_START = 10  # dates start at column 10


class F7Visitas(Functor):
    name = "F7_visitas"
    depends_on = ["F0_bootstrap", "F2_pacientes", "F3_estadias"]

    def objects(self, conn: psycopg.Connection, sources) -> int:
        if openpyxl is None:
            raise RuntimeError("openpyxl required for F7")

        eta = NaturalTransformation()

        # Idempotency
        conn.execute("DELETE FROM operational.visita WHERE TRUE")

        # Build patient RUT → patient_id index
        patient_ruts = {}
        for pid, rut in conn.execute(
            "SELECT patient_id, rut FROM clinical.paciente"
        ).fetchall():
            patient_ruts[rut] = pid

        # Build patient_id → active stay lookup (by date range)
        stay_index: dict[str, list[tuple[str, date, date | None]]] = {}
        for stay_id, patient_id, fi, fe in conn.execute(
            "SELECT stay_id, patient_id, fecha_ingreso, fecha_egreso FROM clinical.estadia ORDER BY fecha_ingreso"
        ).fetchall():
            stay_index.setdefault(patient_id, []).append((stay_id, fi, fe))

        n = 0
        seen_visits: set[str] = set()  # dedup by visit_id

        for fname in PROGRAMACION_FILES:
            xlsx_path = sources.legacy_dir / fname
            if not xlsx_path.exists():
                continue

            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue

                # Row 0: parse dates from column headers
                header = rows[0]
                col_dates: dict[int, date] = {}
                for ci in range(COL_DATES_START, len(header)):
                    val = header[ci]
                    if isinstance(val, datetime):
                        col_dates[ci] = val.date()

                if not col_dates:
                    continue

                # Process patient rows
                for row in rows[1:]:
                    if len(row) <= COL_RUT:
                        continue
                    rut = str(row[COL_RUT]).strip() if row[COL_RUT] else ""
                    if not rut or "-" not in rut:
                        continue

                    patient_id = patient_ruts.get(rut)
                    if not patient_id:
                        continue

                    stays = stay_index.get(patient_id, [])

                    for ci, visit_date in col_dates.items():
                        if ci >= len(row):
                            break
                        cell = row[ci]
                        if not cell:
                            continue
                        tipo_visita = str(cell).strip()
                        if not tipo_visita:
                            continue

                        # Find matching stay for this date
                        stay_id = None
                        for sid, fi, fe in stays:
                            if fi <= visit_date and (fe is None or fe >= visit_date):
                                stay_id = sid
                                break

                        if not stay_id:
                            continue

                        visit_id = make_id("vis", f"{rut}|{visit_date}|{tipo_visita[:20]}")

                        if visit_id in seen_visits:
                            continue
                        seen_visits.add(visit_id)

                        conn.execute(
                            """
                            INSERT INTO operational.visita
                                (visit_id, stay_id, patient_id, fecha, estado, rem_prestacion)
                            VALUES (%s, %s, %s, %s, 'PROGRAMADA', %s)
                            ON CONFLICT (visit_id) DO NOTHING
                            """,
                            (visit_id, stay_id, patient_id, visit_date, tipo_visita),
                        )

                        eta.record(
                            conn,
                            target_table="operational.visita",
                            target_pk=visit_id,
                            source_type="legacy",
                            source_file=fname,
                            source_key=f"{rut}|{visit_date}|{sheet_name}",
                            phase="F7",
                        )
                        n += 1

            wb.close()

        return n

    def path_equations(self) -> list[PathEquation]:
        return [
            PathEquation(
                name="PE-F7-FK-STAY",
                sql="""SELECT v.visit_id FROM operational.visita v
                    LEFT JOIN clinical.estadia e ON e.stay_id = v.stay_id
                    WHERE e.stay_id IS NULL""",
                expected="empty",
            ),
            PathEquation(
                name="PE-F7-FK-PATIENT",
                sql="""SELECT v.visit_id FROM operational.visita v
                    LEFT JOIN clinical.paciente p ON p.patient_id = v.patient_id
                    WHERE p.patient_id IS NULL""",
                expected="empty",
            ),
            PathEquation(
                name="PE-F7-VALID-ESTADO",
                sql="""SELECT visit_id FROM operational.visita
                    WHERE estado NOT IN ('PROGRAMADA','ASIGNADA','DESPACHADA','EN_RUTA',
                    'LLEGADA','EN_ATENCION','COMPLETA','PARCIAL','NO_REALIZADA',
                    'DOCUMENTADA','VERIFICADA','REPORTADA_REM','CANCELADA')""",
                expected="empty",
            ),
            PathEquation(
                name="PE-F7-DATE-IN-STAY",
                sql="""SELECT v.visit_id FROM operational.visita v
                    JOIN clinical.estadia e ON e.stay_id = v.stay_id
                    WHERE v.fecha < e.fecha_ingreso
                       OR (e.fecha_egreso IS NOT NULL AND v.fecha > e.fecha_egreso)""",
                expected="empty",
                severity="warning",
            ),
        ]

    def glue_equations(self) -> list[PathEquation]:
        return [
            PathEquation(
                name="GLUE-F6-F7-PROVIDER-COVERAGE",
                sql="""SELECT COUNT(*) FROM operational.visita
                    WHERE provider_id IS NOT NULL
                      AND provider_id NOT IN (SELECT provider_id FROM operational.profesional)""",
                expected=0,
            ),
        ]
