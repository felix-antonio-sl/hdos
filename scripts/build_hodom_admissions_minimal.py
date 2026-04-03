#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import migrate_hodom_csv as base
import openpyxl
from build_hodom_enriched import canonical_name, dedupe_form_rows, parse_form_workbook


DEFAULT_FORM_PATHS = [
    Path("/Users/felixsanhueza/Downloads/2025 FORMULARIO HODOM (1).xlsx"),
    Path("/Users/felixsanhueza/Downloads/2025 FORMULARIO HODOM (2).xlsx"),
    Path("/Users/felixsanhueza/Downloads/2025 FORMULARIO HODOM.xlsx"),
    Path("/Users/felixsanhueza/Downloads/FORMULARIO 2026 RESP (respuestas) (1).xlsx"),
    Path("/Users/felixsanhueza/Downloads/FORMULARIO 2026 RESP (respuestas) (2).xlsx"),
    Path("/Users/felixsanhueza/Downloads/FORMULARIO 2026 RESP (respuestas).xlsx"),
    Path("/Users/felixsanhueza/Developer/_workspaces/hdos/documentacion-legacy/Hodom-hsc-dia-1/2025 FORMULARIO HODOM.xlsx"),
    Path("/Users/felixsanhueza/Developer/_workspaces/hdos/documentacion-legacy/Hodom-hsc-dia-1/FORMULARIO 2026 RESP (respuestas).xlsx"),
]
LEGACY_FALLBACK_PATH = Path(
    "/Users/felixsanhueza/Developer/_workspaces/hdos/input/reference/legacy_imports/form_response_exports/formulario-hodom-2025-copia-export-2024-05-20-respuestas.csv"
)

DEFAULT_OUTPUT_PATH = Path("output/spreadsheet/hospitalizaciones/ingresos_minimos.csv")
DEFAULT_STRICT_OUTPUT_PATH = Path("output/spreadsheet/hospitalizaciones/ingresos_minimos_estrictos.csv")
DEFAULT_REJECTED_OUTPUT_PATH = Path("output/spreadsheet/hospitalizaciones/ingresos_minimos_descartados.csv")
DEFAULT_MANUAL_RESOLUTION_PATH = Path("input/manual/admissions_minimal_resolution.csv")
DEFAULT_IDENTITY_MASTER_PATH = Path("output/spreadsheet/canonical/patient_identity_master.csv")

MINIMAL_ADMISSION_COLUMNS = [
    "ingreso_id",
    "run",
    "rut_raw",
    "rut_valido",
    "nombre",
    "fecha_ingreso",
    "source_file",
    "source_row_number",
    "source_authority",
]

STRICT_MINIMAL_ADMISSION_COLUMNS = [
    "ingreso_id",
    "run",
    "fecha_ingreso",
    "source_file",
    "source_row_number",
]

REJECTED_ADMISSION_COLUMNS = [
    "ingreso_id",
    "run",
    "rut_raw",
    "rut_valido",
    "nombre",
    "fecha_ingreso",
    "discard_reason",
    "source_file",
    "source_row_number",
    "source_authority",
]

MANUAL_RESOLUTION_COLUMNS = [
    "ingreso_id",
    "action",
    "fecha_ingreso",
    "run",
    "notes",
]


def parse_form_source(path: Path) -> list[dict[str, str]]:
    _, rows = parse_form_workbook(path)
    return rows


def source_priority(path: Path) -> tuple[int, str]:
    name = path.name
    if "FORMULARIO 2026" in name:
        return (0, name)
    if name.startswith("2025 FORMULARIO HODOM - Respuestas"):
        return (1, name)
    if name.startswith("Copia de 2025 FORMULARIO HODOM"):
        return (2, name)
    return (9, name)


def parse_submission_date(value: str) -> str:
    text = base.normalize_whitespace("" if value is None else str(value))
    if not text or text in {",", "-", "None"}:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text[:10]):
        return text[:10]

    for fmt in (
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def build_legacy_timestamp_fallback(path: Path) -> dict[tuple[str, str, str, str], str]:
    if not path.exists():
        return {}

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        canonical_header = {base.canonical_text(name): name for name in fieldnames}
        grouped: defaultdict[tuple[str, str, str, str], set[str]] = defaultdict(set)

        def cell(row: dict[str, str], *keys: str) -> str:
            for key in keys:
                source_key = canonical_header.get(base.canonical_text(key))
                if source_key is not None:
                    return base.normalize_whitespace(row.get(source_key, ""))
            return ""

        for row in reader:
            nombres = cell(row, "NOMBRES", "NOMBRES ")
            apellido_paterno = cell(row, "APELLIDO PATERNO")
            apellido_materno = cell(row, "APELLIDO MATERNO")
            apellidos = cell(row, "APELLIDOS", "APELLIDOS ")
            nombre = canonical_name(nombres, apellido_paterno, apellido_materno, apellidos)
            run = base.normalize_rut(cell(row, "RUT (sin puntos con guión)", "RUT (9999999-9) sin puntos con guión "))
            service = base.normalize_servicio_origen(
                cell(row, "SERVICIO ORIGEN SOLICITUD", "SERVICIO DE ORIGEN SOLICITUD")
            )
            diag = base.canonical_text(
                cell(
                    row,
                    "DIAGNÓSTICO DE EGRESO DE HOSPITALIZACIÓN",
                    "DIAGNÓSTICO DE EGRESO ",
                )
            )
            fecha = parse_submission_date(cell(row, "Marca temporal", "Columna 1"))
            if not fecha:
                continue
            key = (run, base.canonical_text(nombre), service, diag)
            grouped[key].add(fecha)

    fallback: dict[tuple[str, str, str, str], str] = {}
    for key, dates in grouped.items():
        if len(dates) == 1:
            fallback[key] = next(iter(dates))
    return fallback


def build_positional_timestamp_fallback(paths: list[Path]) -> dict[str, str]:
    fallback: dict[str, str] = {}

    for path in paths:
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        row_dates: dict[int, str] = {}
        rows = list(ws.iter_rows(values_only=True))
        total_rows = len(rows)
        for idx, row in enumerate(rows, start=1):
            row_dates[idx] = parse_submission_date(row[0] if row else "")

        for idx, row in enumerate(rows, start=2):
            if row_dates.get(idx):
                continue
            prev_idx = next((i for i in range(idx - 1, 0, -1) if row_dates.get(i)), None)
            next_idx = next((i for i in range(idx + 1, total_rows + 1) if row_dates.get(i)), None)
            if prev_idx is None or next_idx is None:
                continue

            prev_date = row_dates[prev_idx]
            next_date = row_dates[next_idx]
            if not prev_date or not next_date:
                continue

            prev_dt = datetime.fromisoformat(prev_date)
            next_dt = datetime.fromisoformat(next_date)
            day_gap = (next_dt.date() - prev_dt.date()).days
            row_gap = next_idx - prev_idx
            if row_gap <= 1:
                continue
            if day_gap < 0 or day_gap > 7:
                continue

            if prev_date == next_date:
                inferred = prev_date
            else:
                ratio = (idx - prev_idx) / row_gap
                inferred_dt = prev_dt + (next_dt - prev_dt) * ratio
                inferred = inferred_dt.date().isoformat()

            fallback[f"{path.name}:{idx}"] = inferred

    return fallback


def load_form_rows(paths: list[Path]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in sorted(paths, key=source_priority):
        if not path.exists():
            continue
        rows.extend(parse_form_source(path))
    return dedupe_form_rows(rows)


def build_minimal_admission_rows(
    form_rows: list[dict[str, str]],
    legacy_fallback: dict[tuple[str, str, str, str], str] | None = None,
    positional_fallback: dict[str, str] | None = None,
) -> list[dict[str, str]]:
    legacy_fallback = legacy_fallback or {}
    positional_fallback = positional_fallback or {}
    rows: list[dict[str, str]] = []
    for row in form_rows:
        submission_timestamp = row.get("submission_timestamp", "") or ""
        fecha_ingreso = submission_timestamp[:10] if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", submission_timestamp) else ""
        if not fecha_ingreso:
            fallback_key = (
                row.get("rut_norm", ""),
                base.canonical_text(row.get("nombre_completo", "")),
                base.normalize_servicio_origen(row.get("servicio_origen_solicitud", "")),
                base.canonical_text(row.get("diagnostico", "")),
            )
            fecha_ingreso = legacy_fallback.get(fallback_key, "")
        if not fecha_ingreso:
            source_entries = [entry.strip() for entry in row.get("source_rows", "").split(";") if entry.strip()]
            inferred_dates = sorted({positional_fallback.get(entry, "") for entry in source_entries if positional_fallback.get(entry, "")})
            if len(inferred_dates) == 1:
                fecha_ingreso = inferred_dates[0]
        rows.append(
            {
                "ingreso_id": row.get("form_submission_id", ""),
                "run": row.get("rut_norm", ""),
                "rut_raw": row.get("rut_raw", ""),
                "rut_valido": row.get("rut_valido", ""),
                "nombre": row.get("nombre_completo", ""),
                "fecha_ingreso": fecha_ingreso,
                "source_file": row.get("source_files", ""),
                "source_row_number": row.get("source_rows", ""),
                "source_authority": row.get("source_authority", ""),
            }
        )
    rows.sort(key=lambda row: (row["fecha_ingreso"], row["run"], row["nombre"], row["ingreso_id"]))
    return rows


def split_strict_minimal_admission_rows(
    minimal_rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    accepted: list[dict[str, str]] = []
    rejected: list[dict[str, str]] = []
    for row in minimal_rows:
        if not row.get("fecha_ingreso", ""):
            rejected.append(
                {
                    "ingreso_id": row.get("ingreso_id", ""),
                    "run": row.get("run", ""),
                    "rut_raw": row.get("rut_raw", ""),
                    "rut_valido": row.get("rut_valido", ""),
                    "nombre": row.get("nombre", ""),
                    "fecha_ingreso": row.get("fecha_ingreso", ""),
                    "discard_reason": "missing_fecha_ingreso",
                    "source_file": row.get("source_file", ""),
                    "source_row_number": row.get("source_row_number", ""),
                    "source_authority": row.get("source_authority", ""),
                }
            )
            continue
        accepted.append(
            {
                "ingreso_id": row.get("ingreso_id", ""),
                "run": row.get("run", ""),
                "fecha_ingreso": row.get("fecha_ingreso", ""),
                "source_file": row.get("source_file", ""),
                "source_row_number": row.get("source_row_number", ""),
            }
        )
    accepted.sort(key=lambda row: (row["fecha_ingreso"], row["run"], row["ingreso_id"]))
    rejected.sort(key=lambda row: (row["source_file"], row["source_row_number"], row["ingreso_id"]))
    return accepted, rejected


def load_manual_resolution(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    resolved: dict[str, dict[str, str]] = {}
    for row in rows:
        ingreso_id = base.normalize_whitespace(row.get("ingreso_id"))
        if ingreso_id:
            resolved[ingreso_id] = row
    return resolved


def load_identity_run_lookup(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    grouped: defaultdict[str, set[str]] = defaultdict(set)
    for row in rows:
        run = base.normalize_whitespace(row.get("run"))
        if not run:
            continue
        names = [row.get("nombre", ""), *str(row.get("nombre_variantes", "")).split(" | ")]
        for name in names:
            normalized = base.canonical_text(name)
            if normalized:
                grouped[normalized].add(run)

    lookup: dict[str, str] = {}
    for normalized, runs in grouped.items():
        if len(runs) == 1:
            lookup[normalized] = next(iter(runs))
    return lookup


def apply_identity_run_backfill(
    minimal_rows: list[dict[str, str]],
    identity_lookup: dict[str, str],
) -> list[dict[str, str]]:
    resolved_rows: list[dict[str, str]] = []
    for row in minimal_rows:
        updated = dict(row)
        if not updated.get("run"):
            normalized_name = base.canonical_text(updated.get("nombre", ""))
            matched_run = identity_lookup.get(normalized_name, "")
            if matched_run:
                updated["run"] = matched_run
        resolved_rows.append(updated)
    return resolved_rows


def apply_manual_resolution(
    minimal_rows: list[dict[str, str]],
    resolution_by_id: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    resolved_rows: list[dict[str, str]] = []
    for row in minimal_rows:
        updated = dict(row)
        resolution = resolution_by_id.get(row.get("ingreso_id", ""))
        if resolution:
            action = base.normalize_whitespace(resolution.get("action")).lower()
            if action == "discard":
                updated["manual_discard"] = "1"
            elif action == "set_fecha_ingreso":
                updated["fecha_ingreso"] = base.normalize_whitespace(resolution.get("fecha_ingreso"))
            elif action == "set_run":
                run_value = base.normalize_whitespace(resolution.get("run"))
                if run_value:
                    updated["run"] = run_value
                    updated["rut_valido"] = "1"
        resolved_rows.append(updated)
    return resolved_rows


def write_csv(path: Path, rows: list[dict[str, str]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extrae un maestro mínimo de ingresos HODOM desde formularios 2025/2026."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help="CSV de salida con columnas mínimas de ingreso.",
    )
    parser.add_argument(
        "--strict-output",
        type=Path,
        default=DEFAULT_STRICT_OUTPUT_PATH,
        help="CSV de salida con esquema estricto: id, run validado, fecha_ingreso y source.",
    )
    parser.add_argument(
        "--rejected-output",
        type=Path,
        default=DEFAULT_REJECTED_OUTPUT_PATH,
        help="CSV de salida con ingresos descartados del esquema estricto.",
    )
    parser.add_argument(
        "--manual-resolution",
        type=Path,
        default=DEFAULT_MANUAL_RESOLUTION_PATH,
        help="CSV opcional con correcciones manuales para ingresos mínimos.",
    )
    parser.add_argument(
        "--identity-master",
        type=Path,
        default=DEFAULT_IDENTITY_MASTER_PATH,
        help="Maestro canónico de identidad para backfill seguro de run por nombre.",
    )
    args = parser.parse_args()

    form_rows = load_form_rows(DEFAULT_FORM_PATHS)
    legacy_fallback = build_legacy_timestamp_fallback(LEGACY_FALLBACK_PATH)
    positional_fallback = build_positional_timestamp_fallback(DEFAULT_FORM_PATHS)
    minimal_rows = build_minimal_admission_rows(form_rows, legacy_fallback, positional_fallback)
    identity_lookup = load_identity_run_lookup(args.identity_master)
    minimal_rows = apply_identity_run_backfill(minimal_rows, identity_lookup)
    manual_resolution = load_manual_resolution(args.manual_resolution)
    minimal_rows = apply_manual_resolution(minimal_rows, manual_resolution)
    strict_rows, rejected_rows = split_strict_minimal_admission_rows(
        [row for row in minimal_rows if row.get("manual_discard") != "1"]
    )

    write_csv(args.output, minimal_rows, MINIMAL_ADMISSION_COLUMNS)
    write_csv(args.strict_output, strict_rows, STRICT_MINIMAL_ADMISSION_COLUMNS)
    write_csv(args.rejected_output, rejected_rows, REJECTED_ADMISSION_COLUMNS)

    print(f"forms_loaded: {len(form_rows)}")
    print(f"legacy_timestamp_backfill_keys: {len(legacy_fallback)}")
    print(f"positional_timestamp_backfill_keys: {len(positional_fallback)}")
    print(f"identity_run_backfill_keys: {len(identity_lookup)}")
    print(f"manual_resolution_rows: {len(manual_resolution)}")
    print(f"output: {args.output}")
    print(f"ingresos_minimos: {len(minimal_rows)}")
    print(f"strict_output: {args.strict_output}")
    print(f"ingresos_minimos_estrictos: {len(strict_rows)}")
    print(f"rejected_output: {args.rejected_output}")
    print(f"ingresos_minimos_descartados: {len(rejected_rows)}")


if __name__ == "__main__":
    main()
