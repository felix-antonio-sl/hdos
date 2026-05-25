import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from scripts.drive_consolidation import (
    OUT_DIR,
    build_quality_flags,
    classify_item,
    extract_rem_metrics,
    month_number_from_title,
    summarize_workbook,
)


class DriveConsolidationTests(unittest.TestCase):
    def test_outputs_are_written_under_normative_specs_directory(self):
        self.assertEqual(OUT_DIR, Path("/home/felix/projects/hdos/docs/specs/metricas-hodom"))

    def test_month_number_from_title_handles_accents_and_variants(self):
        self.assertEqual(month_number_from_title("REM JULIO .xlsx"), 7)
        self.assertEqual(month_number_from_title("SEPTIEMBRE .xlsx"), 9)
        self.assertEqual(month_number_from_title("REM OCTUBRE 24.xlsx"), 10)
        self.assertEqual(month_number_from_title("INGRESOS 2026 DRIVE"), None)

    def test_classify_item_uses_path_and_title(self):
        item = {"path": "2024/REM/REM ABRIL 24.xlsx", "title": "REM ABRIL 24.xlsx"}
        self.assertEqual(classify_item(item)["dataset"], "rem")
        self.assertEqual(classify_item(item)["year"], 2024)
        self.assertEqual(classify_item(item)["month"], 4)

        item = {"path": "2023/PRESTACIONES ENFERMERIA/JULIO.xlsx", "title": "JULIO.xlsx"}
        self.assertEqual(classify_item(item)["dataset"], "prestaciones_enfermeria")
        self.assertEqual(classify_item(item)["year"], 2023)
        self.assertEqual(classify_item(item)["month"], 7)

        item = {"path": "2025/INGRESOS 2025 (3).xlsx", "title": "INGRESOS 2025 (3).xlsx"}
        self.assertEqual(classify_item(item)["dataset"], "ingresos_nominales")
        self.assertEqual(classify_item(item)["year"], 2025)

        item = {"path": "HOMOM 2025.xlsx", "title": "HOMOM 2025.xlsx"}
        self.assertEqual(classify_item(item)["dataset"], "resumen_hodom")
        self.assertEqual(classify_item(item)["year"], 2025)

    def test_summarize_workbook_counts_non_empty_rows_and_extracts_rem(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rem_sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "REM A21"
            ws["A8"] = "COMPONENTES"
            ws["C10"] = 12
            ws["A10"] = "INGRESOS"
            ws["C11"] = 18
            ws["A11"] = "PERSONAS ATENDIDAS"
            ws["C12"] = 77
            ws["A12"] = "DÍAS PERSONAS ATENDIDAS (TOTAL DE DIAS DE TODOS LOS USUARIOS)"
            ws["A17"] = "SECCIÓN C.2: VISITAS REALIZADAS"
            ws["A19"] = "MÉDICO"
            ws["B19"] = 4
            ws["A20"] = "ENFERMERA"
            ws["B20"] = 20
            ws["A21"] = "Técnico en Enfermería"
            ws["B21"] = 6
            ws["A22"] = "Fonoaudiólogo/a"
            ws["B22"] = 5
            ws["A23"] = "Trabajador/a Social"
            ws["B23"] = 4
            ws["A28"] = "SECCIÓN C.3: CUPOS DISPONIBLES EN UNIDAD DE HOSPITALIZACIÓN DOMICILIARIA"
            ws["A30"] = "NÚMERO CUPOS PROGRAMADOS"
            ws["B30"] = 25
            wb.save(path)

            summary = summarize_workbook(path)

        self.assertEqual(summary["sheet_count"], 1)
        self.assertEqual(summary["sheets"][0]["non_empty_rows"], 12)
        rem = extract_rem_metrics(summary)
        self.assertEqual(rem["personas"]["ingresos"], 12)
        self.assertEqual(rem["personas"]["personas_atendidas"], 18)
        self.assertEqual(rem["personas"]["dias_persona"], 77)
        self.assertEqual(rem["visitas"]["medico"], 4)
        self.assertEqual(rem["visitas"]["enfermera"], 20)
        self.assertEqual(rem["visitas"]["tecnico_paramedico"], 6)
        self.assertEqual(rem["visitas"]["fonoaudiologo"], 5)
        self.assertEqual(rem["visitas"]["trabajador_social"], 4)
        self.assertEqual(rem["cupos"]["cupos_programados"], 25)

    def test_build_quality_flags_marks_coverage_gaps_and_rem_review_items(self):
        metrics = {
            "periods_by_dataset": {
                "prestaciones_enfermeria": ["2024-01", "2024-02", "2024-04"],
                "rem": [f"2024-{month:02d}" for month in range(1, 13)],
            },
            "rem_metrics": [
                {
                    "period": "2024-02",
                    "ingresos": 59,
                    "personas_atendidas": 656,
                    "dias_persona": 629,
                    "visitas_trabajador_social": 102,
                },
                {
                    "period": "2024-11",
                    "ingresos": 53,
                    "personas_atendidas": 70,
                    "dias_persona": 402,
                    "visitas_trabajador_social": None,
                },
            ],
        }

        flags = build_quality_flags(metrics)

        self.assertTrue(any(flag["code"] == "coverage_gap" and "2024-03" in flag["periods"] for flag in flags))
        self.assertTrue(any(flag["code"] == "rem_outlier_personas_atendidas" and flag["period"] == "2024-02" for flag in flags))
        self.assertTrue(any(flag["code"] == "rem_missing_metric" and flag["period"] == "2024-11" for flag in flags))
        self.assertFalse(any(flag["code"] == "empty_source_folder" for flag in flags))


if __name__ == "__main__":
    unittest.main()
