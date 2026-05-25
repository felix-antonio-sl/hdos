#!/usr/bin/env python3
"""Inventory and aggregate non-PII metrics from the HODOM Google Drive folder."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import subprocess
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORK_DIR = ROOT / ".tmp" / "drive-consolidation"
RAW_DIR = WORK_DIR / "raw"
OUT_DIR = ROOT / "docs" / "specs" / "metricas-hodom"
MANIFEST_PATH = OUT_DIR / "manifiesto-fuentes-drive-2026-05-25.json"
METRICS_JSON = OUT_DIR / "consolidacion-fuentes-drive-2026-05-25.json"
METRICS_MD = OUT_DIR / "consolidacion-fuentes-drive-2026-05-25.md"
INVENTORY_CSV = OUT_DIR / "inventario-fuentes-drive-2026-05-25.csv"

DRIVE_ROOT_URL = "https://drive.google.com/drive/folders/1P6ur0dHwmECADUluzwfVvA5g9uTO9oVe?usp=sharing"


EXPECTED_MONTHS_BY_DATASET_YEAR = {
    ("prestaciones_enfermeria", 2024): set(range(1, 13)),
    ("rem", 2024): set(range(1, 13)),
    ("rem", 2025): set(range(1, 13)),
}

REQUIRED_REM_METRIC_FIELDS = [
    "ingresos",
    "personas_atendidas",
    "dias_persona",
    "altas",
    "fallecidos_esperados",
    "reingresos_hospitalizacion",
    "visitas_medico",
    "visitas_enfermera",
    "visitas_kinesiologo",
    "visitas_fonoaudiologo",
    "visitas_trabajador_social",
    "cupos_programados",
    "cupos_utilizados",
    "cupos_disponibles",
]


MANIFEST: list[dict[str, Any]] = [
    {"path": "INVENTARIO INSUMOS.xlsx", "id": "14BHghSF78PzgmztpLbDEZedZfk_lBv42", "title": "INVENTARIO INSUMOS.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "INGRESOS 2026 DRIVE", "id": "1qynoVzgF5a5qMdTVXhfM35zQC4QCtVqh0MNaSHD2_aQ", "title": "INGRESOS 2026 DRIVE", "mime_type": "application/vnd.google-apps.spreadsheet", "native_google": True},
    {"path": "2023/PACIENTES 2023.docx", "id": "1qxXDDohR2_tIFn4bKRfvevkYyChSVsdP", "title": "PACIENTES 2023.docx", "mime_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document"},
    {"path": "2023/PRESTACIONES ENFERMERIA/JULIO.xlsx", "id": "1ATVCG6kWiPQeYIcl09PERdla7Eqn5_bD", "title": "JULIO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/PRESTACIONES ENFERMERIA/AGOSTO.xlsx", "id": "19Nm0_Em-Gbg9XNYwXuVfxOwyG57swSgw", "title": "AGOSTO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/PRESTACIONES ENFERMERIA/SEPTIEMBRE.xlsx", "id": "1_ZCtEQW8ks_1C1EKhqpMN7bw5Tk_KvXl", "title": "SEPTIEMBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/PRESTACIONES ENFERMERIA/OCTUBRE.xlsx", "id": "1BGQSqyMvsQooPXEoHy30xojVRhHpp_B5", "title": "OCTUBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/PRESTACIONES ENFERMERIA/NOVIEMBRE.xlsx", "id": "1gE4-Rd7_tgG2E4AFXYxN4Gz49EJ1yO5t", "title": "NOVIEMBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/REM/REM JULIO.xlsx", "id": "1YqqaWrx5ktmOozPsPgow-Vj_n277pT8r", "title": "REM JULIO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/REM/REM AGOSTO .xlsx", "id": "16EhYHzj79ejkkcbv-U-2AkpqWWyyLozL", "title": "REM AGOSTO .xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/REM/REM SEPTIEMBRE.xlsx", "id": "1KLqIZst4Io8XPpCbjbMZoQ3CQQjWD_r2", "title": "REM SEPTIEMBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/REM/REM OCTUBRE.xlsx", "id": "1zJkabCxZf-tmcVh3QjCg8a8pVm3Enrbt", "title": "REM OCTUBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2023/REM/REM NOVIEMBRE.xlsx", "id": "1wx-I3In7JF4zsjst_e3lLi7jQ-ttQxno", "title": "REM NOVIEMBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/ENERO.xlsx", "id": "17s1fuW2TxP_qVqPiD-mKwD2fLCswvCu5", "title": "ENERO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/FEBRERO.xlsx", "id": "1D2KUCdC-tIoUYgimmstzj-GelMyJWzXM", "title": "FEBRERO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/MARZO.xlsx", "id": "19JTaSuOnVn0OXKoFkVF5FNkM4_67anJX", "title": "MARZO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/ABRIL.xlsx", "id": "1-RdUr7XJbrv6IRegjl3usOYCzQqa3TbE", "title": "ABRIL.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/MAYO.xlsx", "id": "1ANBJfUklBw74Grs6Nja3QEbgr8n-ZcVJ", "title": "MAYO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/JUNIO.xlsx", "id": "1hrCPdomrUa7ay0HAMFfTzNnXxU9SC7yd", "title": "JUNIO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/AGOSTO.xlsx", "id": "1ClJKiA3pdNe3wBtF829Po0bTFr1lIRx0", "title": "AGOSTO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/SEPTIEMBRE .xlsx", "id": "1fhY2hZms5yUWLAyatFME-42krk22MMwd", "title": "SEPTIEMBRE .xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/OCTUBRE.xlsx", "id": "19XlEES5GtMbF9p0-y8kCTZ8KuL8gK_cK", "title": "OCTUBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/NOVIEMBRE.xlsx", "id": "1v8e0i4kPU7qTrQXiWjD5kyR9OL0EORbw", "title": "NOVIEMBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/PRESTACIONES ENFERMERIA/DICIEMBRE.xlsx", "id": "1KoeVrc_FOsNteiW7AVFt259iPYkA7lGx", "title": "DICIEMBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM ENERO.xlsx", "id": "1mzhxOD6oZNKOVZhmtyOAUaG9oCf1EA_3", "title": "REM ENERO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM FEBRERO.xlsx", "id": "1nRSfmUDKCDAL9i1itUpuLaZTGszTtJWD", "title": "REM FEBRERO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM MARZO.xlsx", "id": "1bX7VftfBpK0sKFvL1_-TYj5zKsCtX-l6", "title": "REM MARZO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM ABRIL 24.xlsx", "id": "1TJprQcKtmepC7T3dinRhNhuLb8sZSSVk", "title": "REM ABRIL 24.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM MAYO.xlsx", "id": "1o89epL0PR0uARpLos8La4djWu6XemGCE", "title": "REM MAYO.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM JUNIO 24.xlsx", "id": "1O0vZ-rS6zmXYc3hiO2uJctAuSAGn6qHy", "title": "REM JUNIO 24.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM JULIO .xlsx", "id": "1Y4Z2qhJUH5ac_2TBwyMK3gJEBQQqtU8h", "title": "REM JULIO .xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM AGOSTO .xlsx", "id": "1gJmNAnkjixRWd9iUS2hMzrf1PE7saon5", "title": "REM AGOSTO .xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM SEPTIEMBRE.xlsx", "id": "1lcSUJDJZY7smocpthV4U4C_gDoRlY7Jr", "title": "REM SEPTIEMBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM OCTUBRE 24.xlsx", "id": "1tUk3kjN3wMZUrcrfuHOxSS2KeXmTsc8E", "title": "REM OCTUBRE 24.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM NOVIEMBRE 2024.xlsx", "id": "1QvpN7L-sDHzv832n6wIeG9BF-DZ-Xw11", "title": "REM NOVIEMBRE 2024.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    {"path": "2024/REM/REM DICIEMBRE.xlsx", "id": "16a0d9oKZY1CLbNMrI31hRsNLtQB3rspS", "title": "REM DICIEMBRE.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
]

MANIFEST.extend(
    [
        {"path": "HOMOM 2025.xlsx", "id": "1Sr27xmW0V7eqBZd4tjAJ2AgcID0VSNze", "title": "HOMOM 2025.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "HOMOM 2026.xlsx", "id": "13VFdvVY76YtgbTukf3RCsARmgyNsZaOh", "title": "HOMOM 2026.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2024/INGRESOS ENE-OCT 2024.xlsx", "id": "1g-ojvSZh2uuYHwKz7ch7So9p0V2yLbtp", "title": "INGRESOS ENE-OCT 2024.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2024/PLANILLA NOV DIC 2024.xlsx", "id": "15s5FU9UKBjgTQxz277_M7zSq9KGavLXN", "title": "PLANILLA NOV DIC 2024.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/INGRESOS 2025 (3).xlsx", "id": "1EmbYGViWjZKe_uQoEi9OiGhgXiMdzFtO", "title": "INGRESOS 2025 (3).xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM ENERO 2025.xlsx", "id": "1ZDjH03MGi1-6IXBzvnoMUh4hNwZTNkU2", "title": "REM ENERO 2025.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM FEBRERO 2025.xlsx", "id": "17dnhkUd_livEcTq20R5d4_vVoSK7YdLz", "title": "REM FEBRERO 2025.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM MARZO 25.xlsx", "id": "1o4Gzm6gIRIUB0NPk9qpkk2npgnGC-BOm", "title": "REM MARZO 25.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM ABRIL 25.xlsx", "id": "1GjYTueXURnPif_ROkNqSbwJwyZwfVqLi", "title": "REM ABRIL 25.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM MAYO 25.xlsx", "id": "10LPwBU_IBAzbLy3WqVkiInjdcgoaomQN", "title": "REM MAYO 25.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM JUNIO 25.xlsx", "id": "1G8Irz-5jRvYnITkwqi6bRRfItYH3mk8m", "title": "REM JUNIO 25.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM JULIO 25.xlsx", "id": "1C4aQqgQLwEUDQeIOWETki5oe6Kg9gSmD", "title": "REM JULIO 25.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM AGOSTO 25.xlsx", "id": "1Sn8gTpfkgYlt9cHaqIaOfZQO65fPs77N", "title": "REM AGOSTO 25.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM SEPTIEMBRE 2025.xlsx", "id": "1-AALBIe9XfeIv74dKJfILmcwgsPHR9vY", "title": "REM SEPTIEMBRE 2025.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM OCTUBRE 2025.xlsx", "id": "1egEOppOHucpPKbjJ-JwmEm4nzuZN0r62", "title": "REM OCTUBRE 2025.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM NOVIEMBRE 2025.xlsx", "id": "1Cllu661Xrnv6HOu3-XC1o0nYELk-GjKI", "title": "REM NOVIEMBRE 2025.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        {"path": "2025/REM DICIEMBRE 2025.xlsx", "id": "1o33DqlEBKjxuHvaUvtkuyPASKG_k_6Lj", "title": "REM DICIEMBRE 2025.xlsx", "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    ]
)


MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "SETIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.upper().split())


def month_number_from_title(title: str) -> int | None:
    normalized = normalize_text(title)
    for name, month in MONTHS.items():
        if re.search(rf"\b{name}\b", normalized):
            return month
    return None


def classify_item(item: dict[str, Any]) -> dict[str, Any]:
    path = item.get("path", "")
    parts = path.split("/")
    year = next((int(part) for part in parts if part.isdigit() and len(part) == 4), None)
    if year is None:
        year_match = re.search(r"\b(20\d{2})\b", f"{path} {item.get('title', '')}")
        if year_match:
            year = int(year_match.group(1))
    normalized_path = normalize_text(path)
    if "REM" in normalized_path:
        dataset = "rem"
    elif "PRESTACIONES ENFERMERIA" in normalized_path:
        dataset = "prestaciones_enfermeria"
    elif "HOMOM" in normalized_path:
        dataset = "resumen_hodom"
    elif "INGRESOS 2025" in normalized_path or "INGRESOS ENE-OCT 2024" in normalized_path or "PLANILLA NOV DIC 2024" in normalized_path:
        dataset = "ingresos_nominales"
    elif "INGRESOS 2026" in normalized_path:
        dataset = "ingresos_2026"
        year = 2026
    elif "INVENTARIO INSUMOS" in normalized_path:
        dataset = "inventario_insumos"
    elif path.endswith(".docx"):
        dataset = "pacientes_docx"
    else:
        dataset = "otro"
    return {"dataset": dataset, "year": year, "month": month_number_from_title(item.get("title", ""))}


def safe_name(path: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", path)


def local_path_for(item: dict[str, Any]) -> Path:
    ext = ".xlsx"
    if item["mime_type"].endswith("wordprocessingml.document"):
        ext = ".docx"
    elif item.get("native_google"):
        ext = ".xlsx"
    name = safe_name(item["path"])
    if not name.lower().endswith(ext):
        name += ext
    return RAW_DIR / name


def download_file(item: dict[str, Any], force: bool = False) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    target = local_path_for(item)
    if target.exists() and target.stat().st_size > 0 and not force:
        return target
    if item.get("native_google"):
        url = f"https://docs.google.com/spreadsheets/d/{item['id']}/export?format=xlsx"
    else:
        url = f"https://drive.google.com/uc?export=download&id={item['id']}"
    subprocess.run(["curl", "-L", "-s", "-o", str(target), url], check=True)
    return target


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def cell_value(cell: Any) -> Any:
    if cell is None:
        return None
    if isinstance(cell, str):
        text = cell.strip()
        return text if text else None
    return cell


def numeric(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def summarize_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        non_empty_rows = 0
        non_empty_cells = 0
        sampled_rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            values = [cell_value(value) for value in row]
            if any(value is not None for value in values):
                non_empty_rows += 1
                non_empty_cells += sum(1 for value in values if value is not None)
                if len(sampled_rows) < 80:
                    sampled_rows.append(values[:30])
        sheets.append(
            {
                "title": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "non_empty_rows": non_empty_rows,
                "non_empty_cells": non_empty_cells,
                "sampled_rows": sampled_rows,
            }
        )
    return {"sheet_count": len(sheets), "sheets": sheets}


def extract_docx_metrics(path: Path) -> dict[str, Any]:
    text = ""
    with zipfile.ZipFile(path) as archive:
        if "word/document.xml" in archive.namelist():
            xml = archive.read("word/document.xml")
            root = ElementTree.fromstring(xml)
            text = " ".join(node.text or "" for node in root.iter() if node.tag.endswith("}t"))
    return {"characters": len(text), "word_count": len(text.split()), "line_like_breaks": text.count("\n")}


def find_row(sheet: dict[str, Any], label: str) -> list[Any] | None:
    wanted = normalize_text(label)
    for row in sheet["sampled_rows"]:
        if normalize_text(" ".join(str(value) for value in row[:3] if value is not None)).startswith(wanted):
            return row
    return None


def extract_rem_metrics(workbook_summary: dict[str, Any]) -> dict[str, Any]:
    metrics: dict[str, Any] = {"personas": {}, "visitas": {}, "cupos": {}}
    sheet = next((item for item in workbook_summary["sheets"] if any("HOSPITALIZACION DOMICILIARIA" in normalize_text(row[0] if row else "") for row in item["sampled_rows"])), None)
    if sheet is None and workbook_summary["sheets"]:
        sheet = workbook_summary["sheets"][0]
    if sheet is None:
        return metrics

    row_map: dict[str, list[Any]] = {}
    for row in sheet["sampled_rows"]:
        label = normalize_text(row[0] if row else "")
        if label:
            row_map[label] = row

    def get_from_row(label: str, index: int) -> int | float | None:
        row = find_row(sheet, label)
        if row is None or index >= len(row):
            return None
        return numeric(row[index])

    metrics["personas"] = {
        "ingresos": get_from_row("INGRESOS", 2),
        "personas_atendidas": get_from_row("PERSONAS ATENDIDAS", 2),
        "dias_persona": get_from_row("DIAS PERSONAS ATENDIDAS", 2),
        "altas": get_from_row("ALTAS", 2),
        "fallecidos_esperados": get_from_row("FALLECIDOS", 2),
        "reingresos_hospitalizacion": get_from_row("REINGRESOS A HOSPITALIZACION TRADICIONAL", 2),
    }

    profession_labels = {
        "MEDICO": "medico",
        "ENFERMERA": "enfermera",
        "TECNICO PARAMEDICO": "tecnico_paramedico",
        "TECNICO EN ENFERMERIA": "tecnico_paramedico",
        "KINESIOLOGO": "kinesiologo",
        "FONOAUDIOLOGO": "fonoaudiologo",
        "FONOAUDIOLOGO/A": "fonoaudiologo",
        "TRABAJADOR SOCIAL": "trabajador_social",
        "TRABAJADOR/A SOCIAL": "trabajador_social",
    }
    for raw_label, key in profession_labels.items():
        row = row_map.get(raw_label)
        if row and len(row) > 1:
            metrics["visitas"][key] = numeric(row[1])

    metrics["cupos"] = {
        "cupos_programados": get_from_row("NUMERO CUPOS PROGRAMADOS", 1),
        "cupos_utilizados": get_from_row("NUMERO CUPOS UTILIZADOS", 1),
        "cupos_disponibles": get_from_row("NUMERO DE CUPOS DISPONIBLES", 1),
    }
    return metrics


def analyze_item(item: dict[str, Any], force_download: bool = False) -> dict[str, Any]:
    classification = classify_item(item)
    result = {**item, **classification}
    path = download_file(item, force=force_download)
    result.update({"local_file": str(path), "size_bytes": path.stat().st_size, "sha256": sha256(path)})
    if path.suffix.lower() == ".xlsx":
        workbook = summarize_workbook(path)
        result["workbook"] = {key: value for key, value in workbook.items() if key != "sheets"}
        result["sheets"] = [
            {key: sheet[key] for key in ("title", "max_row", "max_column", "non_empty_rows", "non_empty_cells")}
            for sheet in workbook["sheets"]
        ]
        if classification["dataset"] == "rem":
            result["rem_metrics"] = extract_rem_metrics(workbook)
    elif path.suffix.lower() == ".docx":
        result["docx_metrics"] = extract_docx_metrics(path)
    return result


def build_quality_flags(metrics: dict[str, Any]) -> list[dict[str, Any]]:
    flags: list[dict[str, Any]] = []
    periods_by_dataset = metrics.get("periods_by_dataset", {})
    for (dataset, year), expected_months in sorted(EXPECTED_MONTHS_BY_DATASET_YEAR.items()):
        present = {
            period
            for period in periods_by_dataset.get(dataset, [])
            if isinstance(period, str) and period.startswith(f"{year}-")
        }
        expected = {f"{year}-{month:02d}" for month in expected_months}
        missing = sorted(expected - present)
        if missing:
            flags.append(
                {
                    "code": "coverage_gap",
                    "severity": "review",
                    "dataset": dataset,
                    "year": year,
                    "periods": missing,
                    "detail": "Expected monthly source file was not found in the Drive inventory.",
                }
            )

    for row in metrics.get("rem_metrics", []):
        period = row.get("period")
        missing_fields = [field for field in REQUIRED_REM_METRIC_FIELDS if row.get(field) is None]
        if missing_fields:
            flags.append(
                {
                    "code": "rem_missing_metric",
                    "severity": "review",
                    "period": period,
                    "fields": missing_fields,
                    "detail": "Metric was blank or not located in the source workbook sample.",
                }
            )

        ingresos = numeric(row.get("ingresos"))
        personas = numeric(row.get("personas_atendidas"))
        dias_persona = numeric(row.get("dias_persona"))
        outlier_checks = []
        if ingresos and personas and personas > max(ingresos * 3, ingresos + 100):
            outlier_checks.append(f"personas_atendidas={personas} versus ingresos={ingresos}")
        if dias_persona and personas and personas > dias_persona:
            outlier_checks.append(f"personas_atendidas={personas} exceeds dias_persona={dias_persona}")
        if outlier_checks:
            flags.append(
                {
                    "code": "rem_outlier_personas_atendidas",
                    "severity": "review",
                    "period": period,
                    "checks": outlier_checks,
                    "detail": "Validate this aggregate against the original REM workbook before migration.",
                }
            )
    return flags


def aggregate(results: list[dict[str, Any]]) -> dict[str, Any]:
    by_dataset = Counter(item["dataset"] for item in results)
    by_year = Counter(str(item["year"] or "sin_anio") for item in results)
    bytes_by_dataset: defaultdict[str, int] = defaultdict(int)
    periods_by_dataset: defaultdict[str, list[str]] = defaultdict(list)
    rem_rows = []
    for item in results:
        bytes_by_dataset[item["dataset"]] += item.get("size_bytes", 0)
        if item.get("year") and item.get("month"):
            periods_by_dataset[item["dataset"]].append(f"{item['year']}-{item['month']:02d}")
        if item["dataset"] == "rem" and "rem_metrics" in item:
            rem_rows.append(
                {
                    "period": f"{item['year']}-{item['month']:02d}" if item.get("year") and item.get("month") else item["title"],
                    **item["rem_metrics"]["personas"],
                    **{f"visitas_{key}": value for key, value in item["rem_metrics"]["visitas"].items()},
                    **item["rem_metrics"]["cupos"],
                }
            )

    metrics = {
        "total_files": len(results),
        "by_dataset": dict(by_dataset),
        "by_year": dict(by_year),
        "bytes_by_dataset": dict(bytes_by_dataset),
        "periods_by_dataset": {key: sorted(values) for key, values in periods_by_dataset.items()},
        "rem_periods_extracted": sorted(row["period"] for row in rem_rows),
        "rem_metrics": sorted(rem_rows, key=lambda row: row["period"]),
    }
    metrics["quality_flags"] = build_quality_flags(metrics)
    return metrics


def write_inventory(results: list[dict[str, Any]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    columns = ["path", "id", "title", "dataset", "year", "month", "mime_type", "size_bytes", "sha256", "sheet_count"]
    with INVENTORY_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, lineterminator="\n")
        writer.writeheader()
        for item in results:
            writer.writerow({column: item.get(column) or item.get("workbook", {}).get(column) for column in columns})


def write_markdown(metrics: dict[str, Any], results: list[dict[str, Any]]) -> None:
    lines = [
        "# Especificacion de consolidacion de fuentes Drive HODOM",
        "",
        f"Fuente: {DRIVE_ROOT_URL}",
        f"Generado: {datetime.now(timezone.utc).isoformat(timespec='seconds')}",
        "",
        "## Inventario",
        "",
        f"- Archivos inventariados: {metrics['total_files']}",
        f"- Datasets: {json.dumps(metrics['by_dataset'], ensure_ascii=False, sort_keys=True)}",
        f"- Años cubiertos: {json.dumps(metrics['by_year'], ensure_ascii=False, sort_keys=True)}",
        "",
        "## Cobertura por dataset",
        "",
    ]
    for dataset, periods in sorted(metrics["periods_by_dataset"].items()):
        lines.append(f"- {dataset}: {', '.join(periods) if periods else 'sin periodo mensual'}")
    lines.extend(["", "## Banderas de calidad", ""])
    if metrics.get("quality_flags"):
        for flag in metrics["quality_flags"]:
            context = []
            for key in ("dataset", "year", "period", "path"):
                if flag.get(key) is not None:
                    context.append(f"{key}={flag[key]}")
            for key in ("periods", "fields", "checks"):
                if flag.get(key):
                    context.append(f"{key}={json.dumps(flag[key], ensure_ascii=False)}")
            detail = f" - {flag['detail']}" if flag.get("detail") else ""
            lines.append(f"- [{flag['severity']}] {flag['code']} · {'; '.join(context)}{detail}")
    else:
        lines.append("- Sin banderas de calidad detectadas en esta pasada.")
    lines.extend(["", "## REM A21 C.1 extraido", ""])
    if metrics["rem_metrics"]:
        headers = ["period", "ingresos", "personas_atendidas", "dias_persona", "altas", "fallecidos_esperados", "reingresos_hospitalizacion", "visitas_medico", "visitas_enfermera", "visitas_kinesiologo", "visitas_fonoaudiologo", "visitas_trabajador_social", "cupos_programados", "cupos_utilizados", "cupos_disponibles"]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for row in metrics["rem_metrics"]:
            lines.append("| " + " | ".join(str(row.get(header, "")) for header in headers) + " |")
    lines.extend(
        [
            "",
            "## Observaciones de migracion",
            "",
            "- Los archivos crudos se descargan a `.tmp/drive-consolidation/raw` y no se versionan.",
            "- Esta especificacion vive bajo `docs/specs/metricas-hodom` y define el insumo normativo para staging previo a migracion.",
            "- Los libros Office se tratan como fuentes primarias; el paso siguiente debe crear staging tables antes de tocar tablas clinicas.",
            "- La planilla nativa `INGRESOS 2026 DRIVE` se exporta localmente como XLSX para inspeccion, manteniendo el ID de Drive como clave de provenance.",
            "- Esta pasada produce metricas agregadas; no serializa filas nominales ni datos identificables.",
            "",
            "## Archivos fuente",
            "",
        ]
    )
    for item in sorted(results, key=lambda value: value["path"]):
        sheet_count = item.get("workbook", {}).get("sheet_count", "")
        lines.append(f"- `{item['path']}` · dataset={item['dataset']} · sheets={sheet_count} · bytes={item.get('size_bytes', 0)}")
    METRICS_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(force_download: bool = False) -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = [{**item, **classify_item(item)} for item in MANIFEST]
    MANIFEST_PATH.write_text(json.dumps({"root_url": DRIVE_ROOT_URL, "files": manifest}, ensure_ascii=False, indent=2), encoding="utf-8")
    results = [analyze_item(item, force_download=force_download) for item in MANIFEST]
    metrics = aggregate(results)
    METRICS_JSON.write_text(json.dumps({"metrics": metrics, "files": results}, ensure_ascii=False, indent=2), encoding="utf-8")
    write_inventory(results)
    write_markdown(metrics, results)
    return metrics


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--force-download", action="store_true")
    args = parser.parse_args()
    metrics = run(force_download=args.force_download)
    print(json.dumps(metrics, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
